from odoo import http
from odoo.http import request
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment
from datetime import datetime

class DayReportExcelController(http.Controller):

    @http.route(['/download/day_report_excel/<int:report_id>'], type='http', auth='user')
    def download_day_report_excel(self, report_id=None, **kwargs):
        report = request.env['day.report'].browse(report_id)

        if not report.exists():
            return request.not_found()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Walkin vs Sale"

        headers = ['Date', 'Walk-ins', 'Count Bills', 'Sale Amount']
        header_font = Font(bold=True, color="FFFFFF")
        fill = openpyxl.styles.PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

        for col_num, title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        for row_idx, line in enumerate(report.line_ids, start=2):
            ws.cell(row=row_idx, column=1, value=line.date.strftime('%Y-%m-%d') if line.date else '')
            ws.cell(row=row_idx, column=2, value=line.walkins)
            ws.cell(row=row_idx, column=3, value=line.pos_lines)
            ws.cell(row=row_idx, column=4, value=line.total_amount)

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        file_data = fp.read()
        fp.close()

        filename = f'Day_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return request.make_response(
            file_data,
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"')
            ]
        )
