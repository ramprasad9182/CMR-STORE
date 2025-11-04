import base64
from io import BytesIO
from openpyxl import load_workbook
from odoo import models, fields, _
from odoo.exceptions import UserError, ValidationError
import logging

_logger = logging.getLogger(__name__)

class PurchasePTUploadWizard(models.TransientModel):
    _name = "purchase.pt.upload.wizard"
    _description = "Purchase PT Upload Wizard"

    order_id = fields.Many2one(
        "purchase.order", string="Purchase Order", required=True
    )
    file_data = fields.Binary("Upload Excel", required=True)
    # upload_file = fields.Binary("Upload Excel File", required=True)
    file_name = fields.Char("File Name")

    # def action_import_file(self):
    #     if not self.file_data:
    #         raise UserError(_("Please upload an Excel file."))
    #
    #     data = base64.b64decode(self.file_data)
    #     try:
    #         wb = load_workbook(filename=BytesIO(data), read_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Error reading Excel file: %s") % e)
    #
    #     errors = []
    #     created_lines = 0
    #
    #     # First clear existing order lines
    #     self.order_id.order_line.unlink()
    #
    #     for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         try:
    #             article_code = ''
    #             article_name = ''
    #
    #             if row[1]:
    #                 article_name = str(row[1]).strip()
    #
    #             # Extract code if format is like [CODE] NAME
    #             if article_name.startswith('[') and ']' in article_name:
    #                 article_code = article_name.split(']')[0].strip('[').strip()
    #
    #             if not article_code:
    #                 errors.append("Row %s: No article code found in '%s'" % (row_index, article_name))
    #                 continue
    #
    #             # Search product strictly by default_code
    #             product = self.env['product.product'].search([
    #                 ('default_code', '=', article_code)
    #             ], limit=1)
    #
    #             if not product:
    #                 _logger.warning("Row %s: Product with code '%s' not found", row_index, article_code)
    #
    #                 errors.append("Row %s: Product with code '%s' not found" % (row_index, article_code))
    #                 continue
    #
    #             # Prepare values for purchase.order.line
    #             line_vals = {
    #                 'order_id': self.order_id.id,
    #                 'product_id': product.id,
    #                 'name': article_name or product.display_name,  # ARTICLE_NAME
    #                 'icode_barcode': row[0] or '',
    #                 'brand': row[2] or '',
    #                 'size': row[3] or '',
    #                 'design': row[4] or '',
    #                 'fit': row[5] or '',
    #                 'colour': row[6] or '',
    #                 'product_excel': row[7] or '',
    #                 'standard_rate': float(row[8]) if row[8] else 0.0,
    #                 'mrp': float(row[9]) if row[9] else 0.0,
    #                 'rsp': float(row[10]) if row[10] else 0.0,
    #                 'item_vendor_id': row[11] or '',
    #                 'hsn_sac_code': row[12] or '',
    #                 'product_qty': float(row[13]) if row[13] else 0.0,
    #                 'des5': row[15] or '',
    #                 'des6': row[16] or '',
    #             }
    #
    #             self.env['purchase.order.line'].create(line_vals)
    #             created_lines += 1
    #
    #         except Exception as e:
    #             _logger.error("Row %s failed: %s", row_index, e)
    #             errors.append("Row %s failed: %s" % (row_index, e))
    #             continue
    #
    #     # Reset file_data
    #     self.sudo().write({'file_data': False})
    #
    #
    #     message = _("Successfully created %s purchase order lines.") % created_lines
    #     if errors:
    #         message += "\n\nErrors:\n" + "\n".join(errors)
    #
    #     return {
    #         'effect': {
    #             'fadeout': 'slow',
    #             'message': message,
    #             'type': 'rainbow_man',
    #         }
    #     }

    def action_import_file(self):
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        data = base64.b64decode(self.file_data)
        try:
            wb = load_workbook(filename=BytesIO(data), read_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % e)

        created_lines = 0

        # First clear existing order lines
        self.order_id.order_line.unlink()

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            article_code = ''
            article_name = ''

            if row[1]:
                article_name = str(row[1]).strip()

            # Extract code if format is like [CODE] NAME
            if article_name.startswith('[') and ']' in article_name:
                article_code = article_name.split(']')[0].strip('[').strip()

            if not article_code:
                raise ValidationError(_("Row %s: No article code found in '%s'") % (row_index, article_name))

            # Search product strictly by default_code
            product = self.env['product.product'].search([('default_code', '=', article_code)], limit=1)

            if not product:
                raise ValidationError(
                    _("Row %s: Product with code '%s' not found in Product Master") % (row_index, article_code)
                )

            # Prepare values for purchase.order.line
            line_vals = {
                'order_id': self.order_id.id,
                'product_id': product.id,
                'name': article_name or product.display_name,  # ARTICLE_NAME
                'icode_barcode': row[0] or '',
                'brand': row[2] or '',
                'size': row[3] or '',
                'design': row[4] or '',
                'fit': row[5] or '',
                'colour': row[6] or '',
                'product_excel': row[7] or '',
                'standard_rate': float(row[8]) if row[8] else 0.0,
                'mrp': float(row[9]) if row[9] else 0.0,
                'rsp': float(row[10]) if row[10] else 0.0,
                'item_vendor_id': row[11] or '',
                'hsn_sac_code': row[12] or '',
                'product_qty': float(row[13]) if row[13] else 0.0,
                'des5': row[15] or '',
                'des6': row[16] or '',
            }

            self.env['purchase.order.line'].create(line_vals)
            created_lines += 1

        # Reset file_data
        self.sudo().write({'file_data': False})

        return {
            'effect': {
                'fadeout': 'slow',
                'message': _("Successfully created %s purchase order lines.") % created_lines,
                'type': 'rainbow_man',
            }
        }


