import openpyxl
from io import BytesIO
from odoo import fields, models, _
from odoo.exceptions import ValidationError
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class order_line_wizard(models.TransientModel):
    _name = 'order.line.wizard'
    _description = "Order Line Wizard"

    sale_order_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='xls')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='barcode')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS File'),
         ('from_pricelist', 'Take Details With Adapted Pricelist')], default='from_product')

    sample_option = fields.Selection([('csv', 'CSV'), ('xls', 'XLS')], string='Sample Type', default='xls')
    down_samp_file = fields.Boolean(string='Download Sample Files')

    def import_sol(self):
        res = False
        counter = 0
        if self.import_option == 'csv':
            keys = ['product', 'quantity', 'price']
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.sale_order_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):

                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        product_barcode = ''
                        lot_name = ''
                        if row_no[0] == None:
                            continue
                        else:
                            counter += 1
                            if len(row_no[0]) > 13:
                                if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                    17] == '1':
                                    for i in range(0, len(row_no[0])):
                                        if i > 1 and i < 16:
                                            product_barcode += row_no[0][i]
                                        elif i > 17 and i < len(row_no[0]):
                                            lot_name += row_no[0][i]
                                            continue
                            else:
                                product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({'product': product_barcode,
                                           'serial_no': lot_name,
                                           'quantity': row_no[1],
                                           'price': row_no[2],
                                           })
                        else:
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                            })
                        res = self.create_order_line(values)
        else:
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.sale_order_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):

                product_barcode = ''
                lot_name = ''
                if row_no[0] == None:
                    continue
                else:
                    counter += 1
                    if len(row_no[0]) > 13:
                        if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                            17] == '1':
                            for i in range(0, len(row_no[0])):
                                if i > 1 and i < 16:
                                    product_barcode += row_no[0][i]
                                elif i > 17 and i < len(row_no[0]):
                                    lot_name += row_no[0][i]
                                    continue
                    else:
                        product_barcode = row_no[0]
                if row_no[1] == None or row_no[1] <= 0:
                    raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                if self.product_details_option == 'from_product':
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name
                    })
                elif self.product_details_option == 'from_xls':
                    values.update({'product': product_barcode,
                                   'quantity': row_no[1],
                                   'serial_no': lot_name,
                                   'price': row_no[2],
                                   })
                else:
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name,
                    })
                res = self.create_order_line(values)
        view_id = self.env.ref('nhcl_customizations.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_order_line(self, values):
        sale_order_brw = self.env['sale.order'].browse(self._context.get('active_id'))
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        main_company = self.env['res.company'].search([('nhcl_company_bool', '=', False)])
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13 and values['serial_no']:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    domain = [('name', '=', values['serial_no']), ('product_qty', '>', 0),
                              ('company_id', '=', main_company.id), ('type_product', '=', 'un_brand'), ]
                    if product_obj_search.tracking == 'serial':
                        domain.append(('is_uploaded', '=', False))
                    serial_no = self.env['stock.lot'].search(domain, order='id asc')
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the databasedddddddd.'))
                    serial_id_list = []

                    for serial in serial_no:
                        if serial.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial.name), ('company_id', '=', main_company.id)
                            ,('order_id.state', 'not in', ['cancel','sale'])
                        ])
                        sale_qty = sum(sale_lines.mapped('product_uom_qty'))
                        if serial.id in serial_id_list:
                            raise ValidationError(
                                _('Duplicate serial number detected: ID %s (Name: %s)') % (serial.id, serial.name))
                        if sale_qty < serial.product_qty:
                            serial_id_list.append(serial.id)
                        else:
                            raise ValidationError(
                                _('Serial Number already used in Another Sale Order Line: (Lot Name: %s)') % (
                                    serial.name))
                else:
                    product_barcodes = self.env['product.barcode'].sudo().search([('barcode', '=', values['product'])])

                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
                    if values['product']:
                        domain = [('ref', '=', values['product']), ('product_qty', '>', 0),
                                  ('company_id', '=', main_company.id), ('type_product', '=', 'brand'), ]
                        # serial_no = self.env['stock.lot'].search(
                        #     [('ref', '=', values['product']), ('product_qty', '>', 0), ('is_uploaded','=',False),
                        #      ('company_id', '=', main_company.id), ('type_product', '=', 'brand')], order='id asc'
                        # )
                        if product_obj_search.tracking == 'serial':
                            domain.append(('is_uploaded', '=', False))
                        serial_no = self.env['stock.lot'].search(domain, order='id asc')
                        if not serial_no:
                            raise ValidationError(
                                _('The serial number for this is not found in the database.'))
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            product_id = self.env['product.product']
            if product_obj_search:
                product_id = product_obj_search
            elif values['product']:
                raise ValidationError(_('%s Product was not found in the Database.') % values.get('product'))

            if sale_order_brw.state == 'draft':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)
                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise error if duplicate found

                # Check if there is a branded_barcode field in the values
                if len(values['product']) == 13 and product_obj_search.tracking == 'serial':
                    # Iterate for the number of lines you need
                    for i in range(values['quantity']):
                        lot = serial_no[
                            i % len(serial_no)]
                        if lot.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        if values['product']:
                            sale_lines = self.env['sale.order.line'].search([
                                ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id)
                                ,('order_id.state', 'not in', ['cancel','sale'])
                            ])
                            if sale_lines and lot.serial_type == 'regular':
                                raise ValidationError(
                                    _('%s Qty is exceeded Serial No:%s, available qty:0.Sale Order No:%s') % (
                                        lot.name, values['serial_no'],
                                        sale_lines[0].order_id))
                            else:

                                self.env['sale.order.line'].create({
                                    'order_id': sale_order_brw.id,
                                    'product_id': product_id.id,
                                    'lot_ids': [(6, 0, [lot.id])],
                                    'branded_barcode': values['product'],
                                    'name': product_id.display_name,
                                    'product_uom_qty': 1,
                                    'product_uom': product_id.uom_id.id,
                                    'price_unit': lot.cost_price,  # Assuming you want the cost price from serial_no
                                    'type_product': lot.type_product,
                                    'sale_serial_type': lot.serial_type,
                                })
                            sale_lines = self.env['sale.order.line'].search([
                                ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id)
                                ,('order_id.state', 'not in', ['cancel'])
                            ])
                            sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                            lot_product_qty = lot.product_qty - sale_line_qty
                            if lot_product_qty <= 0:
                                lot.write({'is_uploaded': True})
                else:
                    qty = 0
                    pending_qty = 0

                    if 'serial_no' in values and values['serial_no'] != '':
                        total_serail_no = self.env['stock.lot'].search(
                            [('ref', '=', values['product']), ('product_qty', '>', 0),
                             ('name', '=', values['serial_no']),
                             ('company_id', '=', main_company.id), ('type_product', '!=', 'brand')], order='id asc'
                        )
                        existing_sale_lines = self.env['sale.order.line'].search([
                            ('branded_barcode', '=', values['product']), ('lot_ids.name', '=', values['serial_no']),
                            ('order_id.state', 'not in', ['cancel']), ('company_id', '=', main_company.id)
                        ])
                    else:
                        total_serail_no = self.env['stock.lot'].search(
                            [('ref', '=', values['product']), ('product_qty', '>', 0),
                             ('company_id', '=', main_company.id), ('type_product', '=', 'brand')], order='id asc'
                        )
                        existing_sale_lines = self.env['sale.order.line'].search([
                            ('branded_barcode', '=', values['product']), ('company_id', '=', main_company.id)
                            ,('order_id.state', 'not in', ['cancel','sale'])
                        ])
                    existing_sale_lines_qty = sum(existing_sale_lines.mapped('product_uom_qty'))
                    total_lot_qty = sum(total_serail_no.mapped('product_qty'))
                    available_lot_product_qty = total_lot_qty - existing_sale_lines_qty
                    if len(total_serail_no) > 0 and  total_serail_no[0].serial_type == 'regular' and values[
                        'quantity'] > available_lot_product_qty and 'serial_no' in values and values['serial_no'] != '':
                        raise ValidationError(
                            _('%s Qty is exceeded Serial No:%s, available qty:%s. Sale Order No:%s') % (
                                total_serail_no[0].ref, values['serial_no'], available_lot_product_qty,
                                existing_sale_lines[0].order_id))
                    if len(total_serail_no) > 0 and total_serail_no[0].serial_type == 'regular' and values[
                        'quantity'] > available_lot_product_qty and 'serial_no' in values and values['serial_no'] == '':
                        raise ValidationError(
                            _('%s Qty is exceeded Serial No:%s, available qty:%s.Sale Order No:%s') % (
                                total_serail_no[0].ref, values['serial_no'], available_lot_product_qty,
                                existing_sale_lines[0].order_id))
                    total_qty = 0
                    for lot in serial_no:
                        if lot.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        existing_line = self.env['sale.order.line'].search([
                            ('order_id', '=', sale_order_brw.id), ('company_id', '=', main_company.id),
                            ('lot_ids', 'in', lot.id)
                        ], limit=1)

                        if existing_line and existing_line.product_id.tracking != 'lot':
                            raise ValidationError(
                                _('The serial number "%s" (ID: %s) is already used in this Sale Order.') % (
                                    lot.name, lot.id)
                            )

                        if lot.ref:
                            sale_lines = self.env['sale.order.line'].search([
                                ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id),
                                ('order_id.state', 'not in', ['cancel','sale'])
                            ])
                            sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                            lot_product_qty = lot.product_qty - sale_line_qty
                            if (values['quantity'] > lot_product_qty) or (pending_qty > 0):
                                if pending_qty > 0:
                                    if pending_qty < lot_product_qty:
                                        qty = pending_qty
                                    else:
                                        qty = lot_product_qty
                                else:
                                    qty = lot_product_qty
                                if qty > 0:
                                    if existing_line:
                                        existing_line.product_uom_qty += qty
                                        existing_line.price_unit = lot.cost_price
                                    else:
                                        self.env['sale.order.line'].create({
                                            'order_id': sale_order_brw.id,
                                            'product_id': product_id.id,
                                            'lot_ids': [(6, 0, [lot.id])],
                                            'branded_barcode': lot.ref,
                                            'name': product_id.display_name,
                                            'product_uom_qty': qty,
                                            'product_uom': product_id.uom_id.id,
                                            'price_unit': lot.cost_price,
                                            'type_product': lot.type_product,
                                            'sale_serial_type': lot.serial_type,
                                        })
                                    total_qty += qty
                                    pending_qty = values['quantity'] - total_qty
                                    if values['quantity'] == total_qty:
                                        break
                            else:
                                if existing_line:
                                    existing_line.product_uom_qty += values['quantity']
                                    existing_line.price_unit = lot.cost_price
                                else:
                                    self.env['sale.order.line'].create({
                                        'order_id': sale_order_brw.id,
                                        'product_id': product_id.id,
                                        'lot_ids': [(6, 0, [lot.id])],
                                        'branded_barcode': lot.ref,
                                        'name': product_id.display_name,
                                        'product_uom_qty': values['quantity'],
                                        'product_uom': product_id.uom_id.id,
                                        'price_unit': lot.cost_price,
                                        'type_product': lot.type_product,
                                        'sale_serial_type': lot.serial_type,
                                    })
                                break
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', lot.name), ('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = lot.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            lot.write({'is_uploaded': True})

            elif sale_order_brw.state == 'sent':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)

                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise an error if the product already exists
                if existing_line:
                    raise ValidationError(_('%s Product is already existing.') % values.get('product'))

                # Filter order line by product ID to update existing lines
                existing_order_line = sale_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)

                if existing_order_line:
                    # Update the quantity of the existing order line
                    existing_order_line.product_uom_qty += values.get('quantity')

                    # If serial numbers are present, link them to the existing order line
                    if serial_no:
                        if serial_no.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        # Add the serial numbers to lot_ids without replacing the existing ones
                        existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]

                        # Collect all the barcode references from the lot_ids (serial numbers)
                        barcodes = existing_order_line.lot_ids.mapped('ref')

                        # Filter out any None or empty values and ensure uniqueness
                        unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))

                        # Join the unique barcodes into a comma-separated string
                        branded_barcode_value = ', '.join(unique_barcodes)

                        # Update the branded_barcode field
                        existing_order_line.branded_barcode = branded_barcode_value

                        # Mark all the serial numbers as uploaded
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name), ('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

                elif values.get('product'):
                    # Create a new sale order line with the product and quantity
                    order_lines = self.env['sale.order.line'].create({
                        'order_id': sale_order_brw.id,
                        'product_id': product_id.id,
                        'lot_ids': [(6, 0, serial_no.ids)],
                        'branded_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'price_unit': product_id.lst_price,
                        'type_product': serial_no.type_product,
                        'sale_serial_type': serial_no.serial_type,

                    })
                    # If the order line was created successfully, mark all the serial numbers as uploaded
                    if order_lines:
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name), ('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

            elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        elif self.product_details_option == 'from_xls':
            if self.import_prod_option == 'barcode':
                barcode = values.get('product')
                if len(barcode) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', barcode)])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no']), ('company_id', '=', main_company.id)])
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the databasemmmmmmmmmm.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
                    serial_no = self.env['stock.lot'].search(
                        [('ref', '=', values['product']), ('company_id', '=', main_company.id),('product_qty', '>', 0), ('is_uploaded', '=', False)],
                        limit=1)
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the databasebbbbbbbbbbbb.'))
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database') % values.get(
                            'product'))

            if sale_order_brw.state == 'draft':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)

                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise an error if the product already exists
                if existing_line:
                    raise ValidationError(_('%s Product is already existing.') % values.get('product'))

                # Filter order line by product ID to update existing lines
                existing_order_line = sale_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)

                if existing_order_line:
                    # Update the quantity of the existing order line
                    existing_order_line.product_uom_qty += values.get('quantity')

                    # If serial numbers are present, link them to the existing order line
                    if serial_no:
                        if serial_no.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        # Add the serial numbers to lot_ids without replacing the existing ones
                        existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]

                        # Collect all the barcode references from the lot_ids (serial numbers)
                        barcodes = existing_order_line.lot_ids.mapped('ref')

                        # Filter out any None or empty values and ensure uniqueness
                        unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))

                        # Join the unique barcodes into a comma-separated string
                        branded_barcode_value = ', '.join(unique_barcodes)

                        # Update the branded_barcode field
                        existing_order_line.branded_barcode = branded_barcode_value

                        # Mark all the serial numbers as uploaded
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

                elif values.get('product'):
                    # Create a new sale order line with the product and quantity
                    order_lines = self.env['sale.order.line'].create({
                        'order_id': sale_order_brw.id,
                        'product_id': product_id.id,
                        'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
                        'branded_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'price_unit': product_id.lst_price,
                        'type_product': serial_no.type_product,
                        'sale_serial_type': serial_no.serial_type,

                    })
                    # If the order line was created successfully, mark all the serial numbers as uploaded
                    if order_lines:
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

            elif sale_order_brw.state == 'sent':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)

                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise an error if the product already exists
                if existing_line:
                    raise ValidationError(_('%s Product is already existing.') % values.get('product'))

                # Filter order line by product ID to update existing lines
                existing_order_line = sale_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)

                if existing_order_line:
                    # Update the quantity of the existing order line
                    existing_order_line.product_uom_qty += values.get('quantity')

                    # If serial numbers are present, link them to the existing order line
                    if serial_no:
                        if serial_no.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        # Add the serial numbers to lot_ids without replacing the existing ones
                        existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]

                        # Collect all the barcode references from the lot_ids (serial numbers)
                        barcodes = existing_order_line.lot_ids.mapped('ref')

                        # Filter out any None or empty values and ensure uniqueness
                        unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))

                        # Join the unique barcodes into a comma-separated string
                        branded_barcode_value = ', '.join(unique_barcodes)

                        # Update the branded_barcode field
                        existing_order_line.branded_barcode = branded_barcode_value

                        # Mark all the serial numbers as uploaded
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

                elif values.get('product'):
                    # Create a new sale order line with the product and quantity
                    order_lines = self.env['sale.order.line'].create({
                        'order_id': sale_order_brw.id,
                        'product_id': product_id.id,
                        'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
                        'branded_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'price_unit': product_id.lst_price,
                        'type_product': serial_no.type_product,
                        'sale_serial_type': serial_no.serial_type,

                    })
                    # If the order line was created successfully, mark all the serial numbers as uploaded
                    if order_lines:
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

            elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        else:
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no']),
                                                              ('company_id', '=', main_company.id),])
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the databasesssss.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
                    serial_no = self.env['stock.lot'].search(
                        [('ref', '=', values['product']), ('company_id', '=', main_company.id),('product_qty', '>', 0), ('is_uploaded', '=', False)],
                        limit=1)
                    if not serial_no:
                        raise ValidationError(
                            _('The serial number for this is not found in the databaseeee.'))
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database.') % values.get(
                            'product'))
            if sale_order_brw.state == 'draft':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)

                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise an error if the product already exists
                if existing_line:
                    raise ValidationError(_('%s Product is already existing.') % values.get('product'))

                # Filter order line by product ID to update existing lines
                existing_order_line = sale_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)

                if existing_order_line:
                    # Update the quantity of the existing order line
                    existing_order_line.product_uom_qty += values.get('quantity')

                    # If serial numbers are present, link them to the existing order line
                    if serial_no:
                        if serial_no.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        # Add the serial numbers to lot_ids without replacing the existing ones
                        existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]

                        # Collect all the barcode references from the lot_ids (serial numbers)
                        barcodes = existing_order_line.lot_ids.mapped('ref')

                        # Filter out any None or empty values and ensure uniqueness
                        unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))

                        # Join the unique barcodes into a comma-separated string
                        branded_barcode_value = ', '.join(unique_barcodes)

                        # Update the branded_barcode field
                        existing_order_line.branded_barcode = branded_barcode_value

                        # Mark all the serial numbers as uploaded
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

                elif values.get('product'):
                    # Create a new sale order line with the product and quantity
                    order_lines = self.env['sale.order.line'].create({
                        'order_id': sale_order_brw.id,
                        'product_id': product_id.id,
                        'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
                        'branded_barcode': values.get('product'),
                        'product_uom_qty': values.get('quantity'),
                        'type_product': serial_no.type_product,
                        'sale_serial_type': serial_no.serial_type,

                    })
                    # If the order line was created successfully, mark all the serial numbers as uploaded
                    if order_lines:
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

            elif sale_order_brw.state == 'sent':
                # Checking if the product barcode is longer than 13 characters
                if len(values['product']) > 13:
                    # Filter order line by product ID
                    existing_product_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id)

                    # Check if the serial number exists in the lot_ids of the filtered order line
                    existing_line = existing_product_line.lot_ids.filtered(lambda x: x.id in serial_no.ids)

                else:
                    # Split the 'branded_barcode' field by commas and check if the incoming product barcode exists in any of them
                    existing_line = sale_order_brw.order_line.filtered(
                        lambda x: x.product_id == product_id and
                                  values.get('product') in (x.branded_barcode or '').split(','))

                # Raise an error if the product already exists
                if existing_line:
                    raise ValidationError(_('%s Product is already existing.') % values.get('product'))

                # Filter order line by product ID to update existing lines
                existing_order_line = sale_order_brw.order_line.filtered(
                    lambda x: x.product_id == product_id)

                if existing_order_line:
                    # Update the quantity of the existing order line
                    existing_order_line.product_uom_qty += values.get('quantity')

                    # If serial numbers are present, link them to the existing order line
                    if serial_no:
                        if serial_no.is_under_plan == True:
                            raise ValidationError(_('Lot Number is Under Planned'))
                        # Add the serial numbers to lot_ids without replacing the existing ones
                        existing_order_line.lot_ids = [(4, lot.id) for lot in serial_no]

                        # Collect all the barcode references from the lot_ids (serial numbers)
                        barcodes = existing_order_line.lot_ids.mapped('ref')

                        # Filter out any None or empty values and ensure uniqueness
                        unique_barcodes = list(set([barcode for barcode in barcodes if barcode]))

                        # Join the unique barcodes into a comma-separated string
                        branded_barcode_value = ', '.join(unique_barcodes)

                        # Update the branded_barcode field
                        existing_order_line.branded_barcode = branded_barcode_value

                        # Mark all the serial numbers as uploaded
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

                elif values.get('product'):
                    # Create a new sale order line with the product and quantity
                    order_lines = self.env['sale.order.line'].create({
                        'order_id': sale_order_brw.id,
                        'product_id': product_id.id,
                        'lot_ids': [(6, 0, serial_no.ids)],  # Set the lot_ids with all serial numbers
                        'branded_barcode': values.get('product'),
                        'product_uom_qty': values.get('quantity'),
                        'type_product': serial_no.type_product,
                        'sale_serial_type': serial_no.serial_type,

                    })
                    # If the order line was created successfully, mark all the serial numbers as uploaded
                    if order_lines:
                        sale_lines = self.env['sale.order.line'].search([
                            ('lot_ids.name', '=', serial_no.name),('company_id', '=', main_company.id),
                            ('order_id.state', 'not in', ['cancel'])
                        ])
                        sale_line_qty = sum(sale_lines.mapped('product_uom_qty'))
                        lot_product_qty = serial_no.product_qty - sale_line_qty
                        if lot_product_qty == 0:
                            serial_no.write({'is_uploaded': True})

            elif sale_order_brw.state != 'sent' or sale_order_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        return True
