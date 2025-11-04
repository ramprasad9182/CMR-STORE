import base64
import io
import logging
import ast
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError, UserError
from odoo.tests import Form
import re
import requests
from collections import defaultdict
import csv
from datetime import datetime, timedelta
import xlrd
import pytz
_logger = logging.getLogger(__name__)


class Picking(models.Model):
    """Inherited stock.picking class to add fields and functions"""
    _inherit = "stock.picking"



    stock_picking_delivery_ids = fields.One2many("stock.picking.barcode", "stock_picking_delivery_id")
    nhcl_batch_number = fields.Char(string="Batch Number")

    @api.onchange('no_of_parcel')
    def _onchange_no_of_parcel(self):
        if self.no_of_parcel < 0:
            raise ValidationError(_("No of Parcels cannot be negative."))

        self.stock_picking_delivery_ids = [(5, 0, 0)]  # Clear existing lines

        for i in range(1, self.no_of_parcel + 1):
            self.stock_picking_delivery_ids += self.stock_picking_delivery_ids.new({
                'serial_no': i,
            })

    def print_barcodes_stock_picking(self):
        report_name = 'nhcl_customizations.stock_picking_delivery_barcode'
        return {
            'type': 'ir.actions.report',
            'report_name': report_name,
            'report_type': 'qweb-pdf',
            'res_id': self.id,
            'res_model': 'stock.picking',
        }


    def action_my_button(self):
        self.ensure_one()
        if self.ref_credit_note:
            report = self.env.ref('nhcl_customizations.credit_note_pos')  # This is your report ID
            return report.report_action(self.ref_credit_note)
        else:
            raise UserError("No related Credit Note found to print the report.")

    def action_import_excel(self):
        return {
            'type': 'ir.actions.act_window',
            'name': 'Import Excel',
            'res_model': 'stock.verification.import',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_stock_picking_id': self.id},
        }

    def dev_transport_entry_create(self, rec):
        data = {'partner_id': rec.partner_id and rec.partner_id.id or False,
                'picking_id': rec and rec.id or False,
                'lr_number': rec.lr_number or ' ',
                'transport_details_id': rec.transpoter_id and rec.transpoter_id.id or False,
                'contact_name': rec.transpoter_id and rec.transpoter_id.contact_name or ' ',
                'no_of_parcel': rec.no_of_parcel or 0,
                'name': rec.tracking_number or ' ',
                }
        tra_ent = self.env['dev.transport.entry'].create(data)
        if tra_ent and rec.transpoter_route_id:
            for line in rec.transpoter_route_id.location_details_ids:
                location_detail = {
                    'source_location_id': line.source_location_id and line.source_location_id.id or False,
                    'destination_location_id': line.destination_location_id and line.destination_location_id.id or False,
                    'distance': line.distance,
                    'transport_charges': line.transport_charges,
                    'time_hour': line.time_hour or ' ',
                    'tracking_number': rec.tracking_number or ' ',
                    'picking_id': rec.id,
                    'transport_entry_id': tra_ent.id,
                }
                self.env['transport.location.details'].create(location_detail)
        return tra_ent

    @api.model
    def create(self, vals):
        vals.update({
            'tracking_number': self.env['ir.sequence'].next_by_code('stock.picking.tracking') or '/'
        })
        return super(Picking, self).create(vals)

    @api.depends('move_ids_without_package.nhcl_tax_ids', 'move_ids_without_package.nhcl_price_total',
                 'nhcl_amount_total', 'nhcl_amount_untaxed')
    def _compute_nhcl_tax_totals_json(self):
        for order in self:
            order.nhcl_tax_totals_json = self.env['account.tax']._prepare_tax_totals(
                [x._convert_to_tax_base_line_dict() for x in order.move_ids_without_package],
                order.currency_id or order.company_id.currency_id,
            )

    @api.depends('move_ids_without_package.nhcl_price_total')
    def _amount_all(self):
        for order in self:
            order_lines = order.move_ids_without_package
            amount_untaxed = amount_tax = 0.00
            if order_lines:
                tax_results = self.env['account.tax']._compute_taxes([
                    line._convert_to_tax_base_line_dict()
                    for line in order_lines
                ])
                totals = tax_results['totals']
                amount_untaxed = totals.get(order.currency_id, {}).get('nhcl_amount_untaxed', 0.0)
                amount_tax = totals.get(order.currency_id, {}).get('nhcl_amount_tax', 0.0)
            order.nhcl_amount_untaxed = amount_untaxed
            order.nhcl_amount_tax = amount_tax
            order.nhcl_amount_total = order.nhcl_amount_untaxed + order.nhcl_amount_tax

    nhcl_last_serial_number = fields.Char('Last Serial Number')
    is_confirm = fields.Boolean('Is Confirm', copy=False)
    stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'), ('inter_state', 'Inter State'),
         ('intra_state', 'Intra State'),
         ('others', 'Others')], string='Stock Type', tracking=True)
    dummy_stock_type = fields.Selection(
        [('advertisement', 'Advertisement'), ('ho_operation', 'HO Operation'),
         ('others', 'Others')], string='Dummy Stock Type', compute='_compute_dummy_stock_type')
    stock_barcode = fields.Char(string='Barcode Scan')
    label_click_count = fields.Integer(string="Label Click Count", default=0)
    transpoter_id = fields.Many2one('dev.transport.details', string='Transport by')
    transpoter_route_id = fields.Many2one('dev.routes.details', string='Transporter Route')
    no_of_parcel = fields.Integer(string='No Of Parcels')
    nhcl_credit_note_count = fields.Integer(string='CN Count')
    tracking_number = fields.Char(string='Tracking Number')
    nhcl_tracking_number = fields.Char(string='Source Number')

    lr_number = fields.Char(string='LR Number')
    vehicle_number = fields.Char(string='Vehicle Number')
    driver_name = fields.Char(string='Driver Name')
    transport_location_line = fields.One2many('transport.location.details', 'picking_id', string='Transport Routes')
    transport_entry_ids = fields.One2many('dev.transport.entry', 'picking_id', string='Transport Entry')
    nhcl_replication_status = fields.Boolean(string='Replication Status')
    stock_verification_ids = fields.One2many('stock.verification', 'stock_picking_id')
    verify_barcode = fields.Char(string='Verification Scan')
    stock_picking_type = fields.Selection(
        [('exchange', 'Customer-Return'), ('receipt', 'Receipt'), ('regular', 'Regular'), ('damage', 'Damage'),
         ('return', 'Return'),
         ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
         ('return_main', 'Return-Main'), ('delivery', 'Delivery'), ('pos_order', 'POS Order'),
         ('manufacturing', 'Manufacturing')], string='Type',
        tracking=True, related='picking_type_id.stock_picking_type')
    nhcl_pos_order = fields.Many2one('pos.order', string="POS Order", copy=False)
    nhcl_purchased_store = fields.Char(string="Purchased Store", copy=False)
    nhcl_invoice_date = fields.Date(string="Bill Date", copy=False)
    currency_id = fields.Many2one('res.currency', 'Currency', required=True, readonly=True,
                                  default=lambda self: self.env.company.currency_id.id, copy=False)
    nhcl_tax_totals_json = fields.Binary(compute='_compute_nhcl_tax_totals_json', copy=False)
    nhcl_amount_untaxed = fields.Monetary(string='Untaxed Amount', store=True, readonly=True, compute='_amount_all',
                                          tracking=True, copy=False)
    nhcl_amount_tax = fields.Monetary(string='Taxes', store=True, readonly=True, compute='_amount_all', copy=False)
    nhcl_amount_total = fields.Monetary(string='Total', store=True, readonly=True, compute='_amount_all', copy=False)
    stock_operation_type = fields.Selection([('scan', 'Scan'), ('import', 'Import')
                                                , ('document', 'Document')], string='Operation Type',
                                            tracking=True, default='scan')
    stock_document = fields.Many2one('stock.picking', string="Document", copy=False,
                                     domain=[('stock_picking_type', '=', 'receipt'), ('state', '=', 'done')])
    company_type = fields.Selection([('same', 'Same'), ('other', 'Other')], string="Company Type", copy=False)
    store_pos_order = fields.Char('Pos Order', copy=False)
    store_name = fields.Many2one('nhcl.ho.store.master', string='Store Name', copy=False)
    ref_credit_note = fields.Many2one('account.move', string="Ref. Credit Note",
                                      domain="[('move_type', '=', 'out_refund')]")
    nhcl_phone = fields.Char(string="Phone", related='partner_id.phone')
    scan_or_import = fields.Selection([
        ('scan', 'Scan'),
        ('import', 'Import')
    ], string="Scan or Import", default='scan')
    lot_qty = fields.Float(string="LOT Qty.")
    receipt_lot_qty = fields.Float(string="LOT Qty.")
    exchange_reason = fields.Text(string='Reason', tracking=True)

    # is_picking_type = fields.Boolean(string="Is Picking Type", compute="_compute_is_picking_type", store=True)

    # @api.depends('state', 'stock_picking_type')
    # def _compute_is_picking_type(self):
    #     for record in self:
    #         if record.stock_picking_type == 'exchange' and record.state == 'assigned':
    #             record.is_picking_type = True
    #         if record.stock_type == False :
    #             record.is_picking_type = True
    #         else:
    #             record.is_picking_type = False

    # def action_log_unmatched_serials(self):
    #     """Move remaining unmatched lines into Last Scanned Serial Number"""
    #     LastScanned = self.env['last.scanned.serial.number']
    #     for picking in self:
    #         unmatched_lines = picking.stock_verification_ids.filtered(
    #             lambda l: l.stock_status == 'un_matched'
    #         )
    #         if not unmatched_lines:
    #             continue
    #
    #         # Get all serials already logged in LastScanned
    #         existing_serials = set(
    #             LastScanned.search([('stock_serial', 'in', unmatched_lines.mapped('stock_serial'))]).mapped(
    #                 'stock_serial')
    #         )
    #
    #         vals_list = []
    #         for line in unmatched_lines:
    #             if line.stock_serial in existing_serials:
    #                 continue  # Skip if already exists
    #
    #             vals_list.append({
    #                 'Receipt_number': picking.name,
    #                 'document_number': picking.origin,
    #                 'stock_product_id': line.stock_product_id.id,
    #                 'stock_serial': line.stock_serial,
    #                 'stock_product_barcode': line.stock_product_barcode,
    #                 'stock_qty': line.stock_qty,
    #                 'store_name': picking.company_id.name,
    #                 'date': line.create_date.date().strftime("%Y-%m-%d"),
    #             })
    #
    #         if vals_list:
    #             LastScanned.create(vals_list)

    def get_receipt_picking_lines(self):
        self.ensure_one()
        if self.stock_document:
            # Check if this stock document is already used in another picking
            existing_usage = self.search([
                ('id', '!=', self.id),
                ('stock_document', '=', self.stock_document.id)
            ])
            if existing_usage:
                raise ValidationError("This document has already been used in another picking.")
            # Clear existing lines before generating new ones
            self.move_ids_without_package.unlink()
            self.move_line_ids_without_package.unlink()
            # Get stock moves from the selected document
            moves = self.stock_document.move_ids_without_package.filtered(lambda m: m.state == 'done')
            for move in moves:
                # Create new move for the current picking
                new_move = self.move_ids_without_package.create({
                    'name': move.name,
                    'product_id': move.product_id.id,
                    'product_uom_qty': move.product_uom_qty,
                    'product_uom': move.product_uom.id,
                    'picking_id': self.id,
                    'location_id': move.location_id.id,
                    'location_dest_id': move.location_dest_id.id,
                })
                # Get stock move lines
                for move_line in move.move_line_ids:
                    # Create new move line for the current move
                    self.env['stock.move.line'].create({
                        'move_id': new_move.id,
                        'product_id': move_line.product_id.id,
                        'lot_id': move_line.lot_id.id,
                        'quantity': move_line.quantity,
                        'product_uom_id': move_line.product_uom_id.id,
                        'location_id': move_line.location_id.id,
                        'location_dest_id': move_line.location_dest_id.id,
                        'picking_id': self.id,
                    })

    @api.constrains('nhcl_pos_order')
    def check_nhcl_pos_order(self):
        for rec in self:
            if rec.nhcl_pos_order:
                used_same = self.env['stock.picking'].search_count([('nhcl_pos_order', '=', rec.nhcl_pos_order.id)])
                if used_same > 1:
                    raise ValidationError(f"Already Used in another Exchange.")
            if rec.nhcl_pos_order and rec.nhcl_pos_order.is_pos_order_used == True and rec.state != 'done':
                raise ValidationError("Already Exchanged...")

    @api.constrains('store_pos_order')
    def check_store_pos_order(self):
        for rec in self:
            if rec.store_pos_order:
                used_other = self.env['stock.picking'].search_count([('store_pos_order', '=', rec.store_pos_order)])
                if used_other > 1:
                    raise ValidationError(f"Already Used in another Exchange.")

    # @api.onchange('company_type')
    # def set_stock_type(self):
    #     for rec in self:
    #         rec.stock_type = 'ho_operation'

    @api.onchange('nhcl_pos_order')
    def nhcl_get_pos_order(self):
        for rec in self:
            # Reset fields when no pos order is set
            rec.partner_id = False
            rec.nhcl_invoice_date = False
            rec.nhcl_purchased_store = False
            rec.move_ids_without_package = [(5, 0, 0)]
            rec.move_line_ids_without_package = [(5, 0, 0)]
            if rec.nhcl_pos_order:
                rec.partner_id = rec.nhcl_pos_order.partner_id.id
                rec.nhcl_invoice_date = rec.nhcl_pos_order.date_order.date()
                rec.nhcl_purchased_store = rec.nhcl_pos_order.company_id.name
                rec.stock_type = 'ho_operation'
                move_lines = []
                # Iterate over each line in the pos order
                for line in rec.nhcl_pos_order.lines:
                    # Check if pack_lot_ids exists and get lot name
                    lot_name = line.pack_lot_ids.lot_name if line.pack_lot_ids else False
                    if lot_name:
                        # Search for the corresponding lot record (limit search to one)
                        lots = self.env['stock.lot'].search([('name', '=', lot_name)], limit=1)
                        if lots:
                            lot_ids = [(4, lots.id)]
                            type_product = lots.type_product
                        else:
                            lot_ids = []
                            type_product = False
                    else:
                        lot_ids = []
                        type_product = False
                    # Prepare values for the move line including lot_ids directly
                    move_line_vals = {
                        'location_dest_id': rec.location_dest_id.id,
                        'location_id': rec.location_id.id,
                        'name': line.full_product_name,
                        'product_id': line.product_id.id,
                        'product_uom_qty': line.qty,
                        'quantity': line.qty,
                        'nhcl_rsp': line.price_unit,
                        'nhcl_tax_ids': line.tax_ids_after_fiscal_position,
                        'nhcl_discount': line.discount,
                        'nhcl_gdiscount': line.gdiscount,
                        'nhcl_disc_lines': line.disc_lines,
                        'dummy_lot_ids': lot_ids,
                        'serial_no': lot_name,
                        'pos_order_lines': line.id,
                        'type_product': type_product,
                    }
                    move_lines.append((0, 0, move_line_vals))
                rec.move_ids_without_package = move_lines
                rec.move_ids.move_line_ids.write({'lot_name':lot_name})

    @api.onchange('company_type')
    def _onchange_company_type(self):
        self.store_name = False
        self.store_pos_order = False

    def nhcl_get_pos_order_from_diff_store(self):
        for rec in self:
            if rec.state == 'done':
                raise ValidationError(_("Picking in Done State"))
            rec.partner_id = False
            rec.nhcl_invoice_date = False
            rec.nhcl_purchased_store = False
            rec.move_ids_without_package = [(5, 0, 0)]
            rec.move_line_ids_without_package = [(5, 0, 0)]
            if rec.store_pos_order != False:
                store_ip = rec.store_name.nhcl_terminal_ip
                store_port = rec.store_name.nhcl_port_no
                store_api_key = rec.store_name.nhcl_api_key
                headers_source = {'api-key': f"{store_api_key}", 'Content-Type': 'application/json'}
                pos_data = f"http://{store_ip}:{store_port}/api/pos.order/search"
                pos_data_domain = [('pos_reference', '=', rec.store_pos_order), ('is_pos_order_used', '=', False)]
                pos_data_url = f"{pos_data}?domain={pos_data_domain}"
                pos = requests.get(pos_data_url, headers=headers_source).json()
                if pos.get("data"):
                    pos_partner = False
                    line_partner_data = f"http://{store_ip}:{store_port}/api/res.partner/search"
                    if pos.get("data")[0]['partner_id']:
                        line_partner_data_domain = [('id', '=', pos.get("data")[0]['partner_id'][0]['id'])]
                        line_partner_data_url = f"{line_partner_data}?domain={line_partner_data_domain}"
                        pos_partner = requests.get(line_partner_data_url, headers=headers_source).json()
                    partner_id = False
                    if pos_partner == False:
                        raise ValidationError(_('Invalid Customer for %s in %s store', rec.store_pos_order,
                                                rec.store_name.nhcl_store_name))
                    if pos_partner and pos_partner.get("data"):
                        if pos_partner.get("data")[0]['phone'] == False:
                            raise ValidationError(_('Customer Invalid Phone Number'))
                        phone = pos_partner.get("data")[0]['phone']

                        partner_id = self.env['res.partner'].search([('phone', '=', phone)])
                        if not partner_id:
                            partner_category = self.env['res.partner.category'].search([('name', '=', 'Customer')])
                            if not partner_category:
                                partner_category = self.env['res.partner.category'].create({'name': 'Customer'})
                            partner_id = self.env['res.partner'].create(
                                {'name': pos_partner.get("data")[0]['name'], 'phone': phone,
                                 'vat': "1234567890123Z1", 'group_contact': partner_category.id})
                    pos_order_id = pos.get("data")[0]["id"]
                    rec.partner_id = partner_id
                    rec.nhcl_invoice_date = pos.get("data")[0]["date_order"]
                    rec.nhcl_purchased_store = pos.get("data")[0]["company_id"][0]['name']
                    rec.stock_type = 'ho_operation'
                    for line_data in pos.get("data")[0]["lines"]:
                        pos_line_data = f"http://{store_ip}:{store_port}/api/pos.order.line/search"
                        pos_line_data_domain = [('id', '=', line_data['id']), ('order_id', '=', pos_order_id)]
                        pos_line_data_url = f"{pos_line_data}?domain={pos_line_data_domain}"
                        pos_line = requests.get(pos_line_data_url, headers=headers_source).json()
                        line = False
                        product_id = False
                        pack_lot_ids = False
                        if pos_line.get("data"):
                            line = pos_line.get("data")[0]
                            line_product_data = f"http://{store_ip}:{store_port}/api/product.product/search"
                            line_product_data_domain = [('id', '=', pos_line.get("data")[0]['product_id'][0]['id'])]
                            line_product_data_url = f"{line_product_data}?domain={line_product_data_domain}"
                            pos_line_product = requests.get(line_product_data_url, headers=headers_source).json()
                            if pos_line_product.get("data"):
                                product_nhcl_id = pos_line_product.get("data")[0]['nhcl_id']
                                product_id = self.env['product.product'].search([('nhcl_id', '=', product_nhcl_id)],
                                                                                limit=1)
                        pack_lot_name = False
                        nhcl_categ_1 = False
                        nhcl_categ_2 = False
                        nhcl_categ_3 = False
                        nhcl_categ_4 = False
                        nhcl_categ_5 = False
                        nhcl_categ_6 = False
                        nhcl_categ_7 = False
                        nhcl_categ_8 = False
                        nhcl_descrip_1 = False
                        nhcl_descrip_2 = False
                        nhcl_descrip_3 = False
                        nhcl_descrip_4 = False
                        nhcl_descrip_5 = False
                        nhcl_descrip_6 = False
                        nhcl_descrip_7 = False
                        nhcl_descrip_8 = False
                        if "pack_lot_ids" in line:
                            pack_lot_ids = line["pack_lot_ids"]
                            lot_data = f"http://{store_ip}:{store_port}/api/pos.pack.operation.lot/search"
                            lot_data_domain = [('id', '=', pack_lot_ids[0]['id'])]
                            lot_data_url = f"{lot_data}?domain={lot_data_domain}"
                            pack_lot_data = requests.get(lot_data_url, headers=headers_source).json()
                            pack_line_lot_name = pack_lot_data.get("data")[0]["lot_name"]
                            lot_id_data = f"http://{store_ip}:{store_port}/api/stock.lot/search"
                            lot_id_data_domain = [('name', '=', pack_line_lot_name)]
                            lot_id_data_url = f"{lot_id_data}?domain={lot_id_data_domain}"
                            product_lot = requests.get(lot_id_data_url, headers=headers_source).json()
                            pack_lot_name = product_lot.get("data")[0]
                            if pack_lot_name["category_1"]:
                                nhcl_categ_1 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_1"][0]["name"])])
                            if pack_lot_name["category_2"]:
                                nhcl_categ_2 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_2"][0]["name"])])
                            if pack_lot_name["category_3"]:
                                nhcl_categ_3 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_3"][0]["name"])])
                            if pack_lot_name["category_4"]:
                                nhcl_categ_4 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_4"][0]["name"])])
                            if pack_lot_name["category_5"]:
                                nhcl_categ_5 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_5"][0]["name"])])
                            if pack_lot_name["category_6"]:
                                nhcl_categ_6 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_6"][0]["name"])], limit=1)
                            if pack_lot_name["category_7"]:
                                nhcl_categ_7 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_7"][0]["name"])])
                            if pack_lot_name["category_8"]:
                                nhcl_categ_8 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["category_8"][0]["name"])])
                            if pack_lot_name["description_1"]:
                                nhcl_descrip_1 = self.env['product.aging.line'].search([('name','=',pack_lot_name["description_1"][0]["name"])])
                            if pack_lot_name["description_2"]:
                                nhcl_descrip_2 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_2"][0]["name"])])
                            if pack_lot_name["description_3"]:
                                nhcl_descrip_3 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_3"][0]["name"])])
                            if pack_lot_name["description_4"]:
                                nhcl_descrip_4 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_4"][0]["name"])], limit=1)
                            if pack_lot_name["description_5"]:
                                nhcl_descrip_5 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_5"][0]["name"])])
                            if pack_lot_name["description_6"]:
                                nhcl_descrip_6 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_6"][0]["name"])])
                            if pack_lot_name["description_7"]:
                                nhcl_descrip_7 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_7"][0]["name"])])
                            if pack_lot_name["description_8"]:
                                nhcl_descrip_8 = self.env['product.attribute.value'].search([('name','=',pack_lot_name["description_8"][0]["name"])])

                        tax_id = False
                        if line["tax_ids"]:
                            tax = self.env['account.tax'].search([('name', '=', line["tax_ids"][0]["name"])], limit=1)
                            tax_id = tax.id
                        move_id = self.env['stock.move'].create({
                            'location_dest_id': rec.location_dest_id.id,
                            'location_id': rec.location_id.id,
                            'name': product_id.display_name,
                            'picking_id': rec.id,
                            'product_id': product_id.id,
                            'product_uom_qty': line["qty"],
                            'quantity': line["qty"],
                            'nhcl_rsp': pack_lot_name["rs_price"] if pack_lot_name != False else line['price_unit'],
                            'nhcl_tax_ids': [(6, 0, [tax_id])],
                            'nhcl_discount': line['discount'],
                            'nhcl_gdiscount': line['gdiscount'],
                            'nhcl_disc_lines': line['disc_lines'],
                            'type_product': pack_lot_name["type_product"] if pack_lot_name != False else False,
                            # 'dummy_lot_ids': lot_ids,
                            "serial_no": pack_lot_name["name"] if pack_lot_name != False else False,
                            'ref_pos_order_line_id': line['id'],
                            'nhcl_cost_price': line['nhcl_cost_price'],
                        })
                        for move_line in move_id.move_line_ids:
                            move_line.write(
                                {"internal_ref_lot": pack_lot_name["ref"] if pack_lot_name != False else False,
                                 'rs_price': move_id.nhcl_rsp,
                                 'cost_price': move_id.nhcl_cost_price,
                                 'type_product': move_id.type_product,
                                 "lot_name": pack_lot_name["name"] if pack_lot_name != False else False,
                                 'categ_1': nhcl_categ_1.id if nhcl_categ_1 else False,
                                 'categ_2': nhcl_categ_2.id if nhcl_categ_2 else False,
                                 'categ_3': nhcl_categ_3.id if nhcl_categ_3 else False,
                                 'categ_4': nhcl_categ_4.id if nhcl_categ_4 else False,
                                 'categ_5': nhcl_categ_5.id if nhcl_categ_5 else False,
                                 'categ_6': nhcl_categ_6.id if nhcl_categ_6 else False,
                                 'categ_7': nhcl_categ_7.id if nhcl_categ_7 else False,
                                 'categ_8': nhcl_categ_8.id if nhcl_categ_8 else False,
                                 'descrip_1': nhcl_descrip_1.id if nhcl_descrip_1 else False,
                                 'descrip_2': nhcl_descrip_2.id if nhcl_descrip_2 else False,
                                 'descrip_3': nhcl_descrip_3.id if nhcl_descrip_3 else False,
                                 'descrip_4': nhcl_descrip_4.id if nhcl_descrip_4 else False,
                                 'descrip_5': nhcl_descrip_5.id if nhcl_descrip_5 else False,
                                 'descrip_6': nhcl_descrip_6.id if nhcl_descrip_6 else False,
                                 'descrip_7': nhcl_descrip_7.id if nhcl_descrip_7 else False,
                                 'descrip_8': nhcl_descrip_8.id if nhcl_descrip_8 else False,
                                 })
                else:
                    raise ValidationError(
                        _('Pos Order Number does not exist in %s or Already Used.', rec.store_name.nhcl_store_name))

    def stock_create_credit_note(self):
        """Creates and posts a credit note from stock picking"""
        for picking in self:
            if picking.nhcl_credit_note_count >= 1:
                raise ValidationError("Credit Note is already generated against " + picking.name)
            if not picking.partner_id:
                raise ValidationError("Customer is missing for this picking!")
            picking.nhcl_credit_note_count += 1
            journal = self.env['account.journal'].search([('type', '=', 'cash'), ('name', '=', 'Credit Note Issue')],
                                                         limit=1)
            if not journal:
                raise ValidationError(
                    "No journal found with type 'cash' and name 'Credit Note Issue'. Please configure it.")
            credit_note_vals = {
                'move_type': 'out_refund',
                'partner_id': picking.partner_id.id,
                'invoice_origin': picking.name,
                'picking_ref': picking.name,
                'currency_id': self.env.company.currency_id.id,
                'journal_id': journal.id,
                'invoice_date': fields.Date.context_today(self),
                'invoice_line_ids': [],
            }
            credit_note = self.env['account.move'].create(credit_note_vals)
            has_lines = False
            for move in picking.move_ids_without_package:
                if move.quantity > 0:
                    product = move.product_id
                    account = product.property_account_income_id
                    if not account:
                        raise ValidationError(f"Income account is not defined for product {product.display_name}.")
                    price_unit = move.nhcl_rsp
                    if move.nhcl_discount:
                        price_unit -= (price_unit * move.nhcl_discount / 100)
                    if move.nhcl_gdiscount:
                        price_unit -= (price_unit * move.nhcl_gdiscount / 100)

                    self.env['account.move.line'].create({
                        'move_id': credit_note.id,
                        'product_id': product.id,
                        'quantity': move.quantity,
                        'price_unit': price_unit,
                        'name': "Credit Note",
                        'account_id': account.id,
                        'tax_ids': [(6, 0, move.nhcl_tax_ids.ids)],
                    })
                    has_lines = True
            if not has_lines:
                raise ValidationError("No credit note lines created. Credit note cannot be posted.")
            credit_note.action_post()
            if credit_note.state == 'posted':
                picking.partner_id.wallet_amount += credit_note.amount_total
                picking.ref_credit_note = credit_note.id
                picking.partner_id.credit_note_ids = [(0, 0, {
                    'voucher_number': picking.ref_credit_note.name,
                    'pos_bill_number': picking.nhcl_pos_order.pos_reference,
                    'pos_bill_date': picking.nhcl_invoice_date,
                    'total_amount': credit_note.amount_total,
                })]

    def update_pos_order_status_other_store(self):
        for rec in self:
            store_ip = rec.store_name.nhcl_terminal_ip
            store_port = rec.store_name.nhcl_port_no
            store_api_key = rec.store_name.nhcl_api_key
            headers_source = {'api-key': f"{store_api_key}", 'Content-Type': 'application/json'}
            pos_data = f"http://{store_ip}:{store_port}/api/pos.order/search"
            pos_data_domain = [('pos_reference', '=', rec.store_pos_order)]
            pos_data_url = f"{pos_data}?domain={pos_data_domain}"
            try:
                response = requests.get(pos_data_url, headers=headers_source)
                response.raise_for_status()  # Raises an HTTPError for bad responses

                # Parse the JSON response
                data = response.json()  # Now `data` is a dictionary
                pos_order_data = data.get("data", [])
                if pos_order_data:
                    pos_order_id = pos_order_data[0]['id']

                    # Step 3: Prepare payload
                    pos_list = {
                        'is_pos_order_used': True,
                    }

                    # Step 4: Update the POS order
                    store_url_data1 = f"http://{store_ip}:{store_port}/api/pos.order/{pos_order_id}"
                    update_response = requests.put(store_url_data1, headers=headers_source, json=pos_list)
                    update_response.raise_for_status()

                    _logger.info(
                        f"Successfully updated POS order '{rec.store_pos_order}' "
                        f"in store '{rec.store_name.nhcl_store_name}' as used."
                    )
            except requests.exceptions.RequestException as e:
                _logger.error(
                    f"'{rec.name}' Failed to update POS order '{rec.store_pos_order}' "
                    f"in '{rec.store_name.nhcl_store_name}' store. Error: {e}"
                )


    @api.onchange('partner_id')
    def get_stock_type(self):
        if self.picking_type_id.code == 'outgoing':
            if self.partner_id and self.env.company.state_id:
                if self.partner_id.state_id.id == self.env.company.state_id.id:
                    self.stock_type = 'intra_state'
                else:
                    self.stock_type = 'inter_state'
            else:
                self.stock_type = ''
        elif self.picking_type_id.code == 'internal':
            if self.partner_id.name == 'CMR TEXTILES PRIVATE LIMITED':
                self.stock_type = 'ho_operation'
    # Store button
    def button_validate(self):
        res = None
        for rec in self:
            # Check if stock_picking_type is 'exchange'
            if rec.stock_picking_type == 'exchange' and rec.picking_type_code == 'incoming':
                if not self.env.context.get('bypass_exchange_wizard') and any(
                        rec.move_ids_without_package.filtered(lambda x: x.nhcl_exchange == True)) == True:
                    return {
                        'name': _('Exchange Confirmation'),
                        'type': 'ir.actions.act_window',
                        'target': 'new',
                        'res_model': 'pos.exchange.wizard',
                        'view_mode': 'form',
                        'view_id': self.env.ref('nhcl_customizations.view_pos_exchange_wizard_wizard').id,
                        'context': {'default_nhcl_picking_id': rec.id,
                                    'bypass_exchange_wizard': True,
                                    },
                    }
                discount_lines = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and '% on your order' in (
                            line.product_id.name or '').lower()
                )
                if discount_lines:
                    valid_lines = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    order_total_qty = sum(valid_lines.mapped('quantity'))
                    for disc_line in discount_lines:
                        if valid_lines:
                            share = abs(disc_line.nhcl_rsp) / order_total_qty
                            for valid_line in valid_lines:
                                valid_line.nhcl_rsp -= share
                gift_card_line = rec.move_ids_without_package.filtered(
                    lambda line: not line.serial_no and line.nhcl_rsp < 0 and 'Gift Card' in (
                            line.product_id.name or '')
                )
                if gift_card_line:
                    valid_lines_list = rec.move_ids_without_package.filtered(
                        lambda line: line.serial_no and line.nhcl_rsp >= 0)
                    total_qty = sum(valid_lines_list.mapped('quantity'))
                    if total_qty > 0:
                        for gift_line in gift_card_line:
                            gift_share_per_unit = abs(gift_line.nhcl_rsp) / total_qty
                            for valid_line in valid_lines_list:
                                valid_line.nhcl_rsp -= gift_share_per_unit
                rec.stock_pos_exchange()
                # Filter out move lines where nhcl_exchange is False and unlink them
                rec.move_ids_without_package.filtered(lambda line: not line.nhcl_exchange).unlink()
                rec.nhcl_pos_order.is_pos_order_used = True
            # Update lot_ids from dummy_lot_ids before validation
            for move_line in rec.move_ids_without_package:
                if move_line.dummy_lot_ids:
                    move_line.lot_ids = [(6, 0, move_line.dummy_lot_ids.ids)]
                    if rec.company_type == "same":
                        move_line.move_line_ids.update({'cost_price': move_line.dummy_lot_ids.cost_price})
                    elif rec.company_type == "other":
                        move_line.move_line_ids.lot_id.update({
                            'cost_price': move_line.nhcl_cost_price,
                            'category_1': move_line.categ_1.id if move_line.categ_1 else False,
                            'category_2': move_line.categ_2.id if move_line.categ_2 else False,
                            'category_3': move_line.categ_3.id if move_line.categ_3 else False,
                            'category_4': move_line.categ_4.id if move_line.categ_4 else False,
                            'category_5': move_line.categ_5.id if move_line.categ_5 else False,
                            'category_6': move_line.categ_6.id if move_line.categ_6 else False,
                            'category_7': move_line.categ_7.id if move_line.categ_7 else False,
                            'category_8': move_line.categ_8.id if move_line.categ_8 else False,
                            'description_1': move_line.descrip_1.id if move_line.descrip_1 else False,
                            'description_2': move_line.descrip_2.id if move_line.descrip_2 else False,
                            'description_3': move_line.descrip_3.id if move_line.descrip_3 else False,
                            'description_4': move_line.descrip_4.id if move_line.descrip_4 else False,
                            'description_5': move_line.descrip_5.id if move_line.descrip_5 else False,
                            'description_6': move_line.descrip_6.id if move_line.descrip_6 else False,
                            'description_7': move_line.descrip_7.id if move_line.descrip_7 else False,
                            'description_8': move_line.descrip_8.id if move_line.descrip_8 else False,
                        })
            if rec.stock_picking_type in ('return','damage') and rec.picking_type_code == 'outgoing':
                if not all([
                    rec.transpoter_id,
                    rec.transpoter_route_id,
                    rec.lr_number,
                    rec.driver_name,
                    rec.no_of_parcel,
                    rec.vehicle_number,
                ]):
                    raise UserError("Please fill all transporter details for Return type before confirming.")
            if rec.stock_picking_type == 'receipt' and rec.picking_type_code == 'incoming':
                for stock_move_id in rec.move_ids:
                    stock_move_id.product_uom_qty = sum(self.stock_verification_ids.filtered(
                            lambda x: x.stock_product_id == stock_move_id.product_id).mapped('stock_qty'))
                assigned_serials = set(rec.move_line_ids_without_package.filtered(lambda x: x.lot_name).mapped('lot_name'))
                attribute_values = self.env['product.attribute.value'].search([])
                for dummy in self.stock_verification_ids:
                    if dummy.stock_status != 'matched' and dummy.stock_product_id.tracking == 'serial':
                        continue
                    for move_line in self.move_line_ids_without_package.filtered(
                            lambda x: x.product_id.id == dummy.stock_product_id.id):
                        if (not move_line.lot_name
                                and move_line.type_product == dummy.type_product
                                and dummy.stock_serial not in assigned_serials):
                            names_to_match = [
                                dummy.nhcl_categ_1, dummy.nhcl_categ_2, dummy.nhcl_categ_3, dummy.nhcl_categ_4,
                                dummy.nhcl_categ_5, dummy.nhcl_categ_6, dummy.nhcl_categ_7, dummy.nhcl_categ_8,
                                 dummy.nhcl_descrip_2, dummy.nhcl_descrip_3,
                                dummy.nhcl_descrip_4, dummy.nhcl_descrip_5, dummy.nhcl_descrip_6,
                            ]

                            names_to_match = [n for n in names_to_match if n]

                            attribute_ids = self.env['product.attribute.value'].search([
                                ('name', 'in', names_to_match)
                            ])
                            description_1 = self.env['product.aging.line'].search([('name','=',dummy.nhcl_descrip_1)])
                            categ_1 = categ_2 = categ_3 = categ_4 = categ_5 = categ_6= categ_7 = categ_8 = False
                            description_2 = description_3 = description_4 = description_5 = description_6 = False
                            for attribute_id in attribute_ids:
                                if attribute_id.name ==  dummy.nhcl_categ_1:
                                    categ_1 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_2:
                                    categ_2 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_3:
                                    categ_3 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_4:
                                    categ_4 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_5:
                                    categ_5 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_6:
                                    categ_6 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_7:
                                    categ_7 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_categ_8:
                                    categ_8 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_descrip_2:
                                    description_2 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_descrip_3:
                                    description_3 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_descrip_4:
                                    description_4 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_descrip_5:
                                    description_5 = attribute_id.id
                                elif attribute_id.name == dummy.nhcl_descrip_6:
                                    description_6 = attribute_id.id

                            move_line.update({
                                'lot_name': dummy.stock_serial,
                                'quantity': dummy.stock_actual_qty,
                                'internal_ref_lot': dummy.stock_product_barcode,
                                'mr_price': dummy.mr_price,
                                'rs_price': dummy.rs_price or 0,
                                'cost_price': dummy.cost_price or 0,
                                'segment': dummy.segment,
                                'categ_1': categ_1 or False,
                                'categ_2': categ_2 or False,
                                'categ_3': categ_3 or False,
                                'categ_4': categ_4 or False,
                                'categ_5': categ_5 or False,
                                'categ_6': categ_6 or False,
                                'categ_7': categ_7 or False,
                                'categ_8': categ_8 or False,
                                'descrip_1': description_1 or False,
                                'descrip_2': description_2 or False,
                                'descrip_3': description_3 or False,
                                'descrip_4': description_4 or False,
                                'descrip_5': description_5 or False,
                                'descrip_6': description_6 or False,
                            })
                            assigned_serials.add(dummy.stock_serial)
                to_zero = rec.move_line_ids_without_package.filtered(lambda x: not x.lot_name)
                if to_zero:
                    to_zero.write({'quantity': 0.0})
            res = super(Picking, self).button_validate()
            if rec.state == 'done' and rec.transpoter_id and rec.stock_type in ['inter_state', 'intra_state']:
                rec.dev_transport_entry_create(rec)
            if rec.company_type == 'other':
                rec.update_pos_order_status_other_store()

        return res


    def stock_pos_exchange(self):
        for picking in self:
            for move in picking.move_ids_without_package:
                if move.nhcl_disc_lines:
                    try:
                        prod_list = ast.literal_eval(move.nhcl_disc_lines or "[]")
                    except Exception as e:
                        print(f"Error parsing nhcl_disc_lines: {e}")
                        continue
                    if not prod_list:
                        continue
                    # Normalize and clean prod names from discount line
                    prod_list_clean = [str(p).strip().lower() for p in prod_list]
                    # Find all matching lines
                    matching_moves = picking.move_ids_without_package.filtered(
                        lambda m: m != move and m.product_id.nhcl_display_name.strip().split(']', 1)[
                            -1].strip().lower() in prod_list_clean)
                    if matching_moves:
                        share = move.nhcl_rsp / len(matching_moves)
                        for matched_move in matching_moves:
                            print(f" - Reducing {share} from {matched_move.product_id.nhcl_display_name}")
                            matched_move.nhcl_rsp += share

    # # naseer
    # @api.model
    # def create(self, vals):
    #     rec = super(Picking, self).create(vals)
    #     if rec.transpoter_id and rec.stock_type in ['inter_state','intra_state']:
    #         rec.dev_transport_entry_create(rec)
    #     return rec

    @api.depends('stock_type')
    def _compute_dummy_stock_type(self):
        for i in self:
            if i.stock_type == 'ho_operation':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'advertisement':
                i.dummy_stock_type = 'advertisement'
            elif i.stock_type == 'others':
                i.dummy_stock_type = 'others'
            elif i.stock_type == 'inter_state':
                i.dummy_stock_type = 'ho_operation'
            elif i.stock_type == 'intra_state':
                i.dummy_stock_type = 'ho_operation'
            else:
                i.dummy_stock_type = ''

    @api.onchange('verify_barcode')
    def _onchange_verify_barcode(self):
        if self.verify_barcode:
            barcode = self.verify_barcode
            # Patterns for barcode formats
            gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
            ean13_pattern = r'(\d{13})'
            custom_serial_pattern = r'^(R\d+)'
            if re.match(gs1_pattern, barcode):
                # GS1 Barcode
                product_barcode = re.match(gs1_pattern, barcode).group(1)
                code = re.match(gs1_pattern, barcode).group(2)
                matched_line = self.stock_verification_ids.filtered(lambda x: x.stock_serial == code)
                if not matched_line:
                    if not matched_line:
                        existing = self.env['last.scanned.serial.number'].search([
                            ('stock_serial', '=', code)
                        ], limit=1)
                        if not existing:
                            self.env['last.scanned.serial.number'].create({
                                'stock_serial': code,
                                'stock_product_barcode': barcode,
                                'receipt_number': self.name,
                                'document_number': self.origin,
                                'store_name': self.company_id.name
                            })
                            self.env.cr.commit()
                    raise ValidationError('No matching product or serial/lot number found.')
                tracking_type = matched_line.stock_product_id.tracking
                product_id = matched_line.stock_product_id
                if tracking_type == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("You not should enter qty.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = 1.0
                elif tracking_type == 'lot':
                    if self.receipt_lot_qty == 0 or self.receipt_lot_qty < 0:
                        raise ValidationError("You should enter qty.")
                    if matched_line.stock_qty < self.receipt_lot_qty:
                        raise ValidationError(
                            f"You have given {self.receipt_lot_qty} more than {matched_line.stock_qty}.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    if matched_line.stock_qty == self.receipt_lot_qty:
                        matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = self.receipt_lot_qty

            elif re.match(ean13_pattern, barcode):
                # EAN-13 Barcode
                ean13_barcode = re.match(ean13_pattern, barcode).group(1)
                matched_line = self.stock_verification_ids.filtered(lambda x: x.stock_product_barcode == ean13_barcode and x.stock_status == 'un_matched' )[:1]
                if not matched_line:
                    if not matched_line:
                        existing = self.env['last.scanned.serial.number'].search([
                            ('stock_product_barcode', '=', ean13_barcode)
                        ], limit=1)
                        if not existing:
                            self.env['last.scanned.serial.number'].create({
                                'stock_product_barcode': ean13_barcode,
                                'receipt_number': self.name,
                                'document_number': self.origin,
                                'store_name': self.company_id.name
                            })
                            self.env.cr.commit()
                    raise ValidationError('No matching product or serial/lot number found.')
                tracking_type = matched_line.stock_product_id.tracking
                if tracking_type == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("You not should enter qty.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = 1.0
                elif tracking_type == 'lot':
                    if self.receipt_lot_qty == 0 or self.receipt_lot_qty < 0:
                        raise ValidationError("You should enter qty.")
                    if matched_line.stock_qty < self.receipt_lot_qty:
                        raise ValidationError(
                            f"You have given {self.receipt_lot_qty} more than {matched_line.stock_qty}.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    if matched_line.stock_qty == self.receipt_lot_qty:
                        matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = self.receipt_lot_qty
            elif re.match(custom_serial_pattern, barcode):
                # Custom Serial Barcode
                code = re.match(custom_serial_pattern, barcode).group(1)
                matched_line = self.stock_verification_ids.filtered(
                    lambda x: x.stock_serial == code and x.type_product == 'un_brand')
                if not matched_line:
                    # Fallback to internal_ref_lot check
                    matched_line = self.stock_verification_ids.filtered(
                        lambda x: x.stock_product_barcode == code and x.type_product == 'brand')[:1]
                    if not matched_line:
                        existing = self.env['last.scanned.serial.number'].search([
                            ('stock_serial', '=', code)
                        ], limit=1)
                        if not existing:
                            self.env['last.scanned.serial.number'].create({
                                'stock_serial': code,
                                'receipt_number': self.name,
                                'document_number': self.origin,
                                'store_name': self.company_id.name
                            })
                            self.env.cr.commit()
                        raise ValidationError('No matching product or serial/lot number found.')
                tracking_type = matched_line.stock_product_id.tracking
                if tracking_type == 'serial':
                    if self.receipt_lot_qty != 0:
                        raise ValidationError("You not should enter qty.")
                    if matched_line.stock_status == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = 1.0
                elif tracking_type == 'lot':
                    if self.receipt_lot_qty == 0 or self.receipt_lot_qty < 0:
                        raise ValidationError("You should enter qty.")
                    if matched_line.stock_qty < self.receipt_lot_qty:
                        raise ValidationError(
                            f"You have given {self.receipt_lot_qty} more than {matched_line.stock_qty}.")
                    if matched_line.stock_qty == 'matched':
                        raise ValidationError(f"Already Matched {matched_line.stock_serial}.")
                    if matched_line.stock_qty == self.receipt_lot_qty:
                        matched_line.stock_status = 'matched'
                    matched_line.stock_actual_qty = self.receipt_lot_qty
            else:
                raise ValidationError('Invalid barcode format.')
            self.verify_barcode = False
            self.receipt_lot_qty = 0.0

    def action_open_label_type(self):
        # Increment the click count before or after calling super
        self.label_click_count += 1

        # Call the original method using super()
        return super(Picking, self).action_open_label_type()

    # @api.onchange('stock_barcode')
    # def _onchange_stock_barcode(self):
    #     if not self.stock_type:
    #         if self.stock_barcode:
    #             raise ValidationError('Please select a Stock Type before scanning a barcode.')
    #         return
    #
    #     barcode = self.stock_barcode
    #
    #     # Barcode patterns
    #     gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
    #     ean13_pattern = r'(\d{13})'
    #     custom_serial_pattern = r'^(R\d+)'
    #
    #     def search_product(field, value):
    #         """Helper: Search for product in product.product and product.template."""
    #         product = self.env['product.product'].search([(field, '=', value)], limit=1)
    #         if not product:
    #             template = self.env['product.template'].search([(field, '=', value)], limit=1)
    #             if template:
    #                 product = template.product_variant_id
    #         return product
    #
    #     def update_lot_tracked(product, lot):
    #         """
    #         Helper: For products tracked by lot.
    #         - If the lot exists in the move line, increase its quantity.
    #         - If the lot is new, add it to the existing move line.
    #         """
    #         line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
    #         if line:
    #             lot_line = line.dummy_lot_ids.filtered(lambda l: l.id == lot.id)
    #             if lot_line:
    #                 # If the lot exists, increase its quantity
    #                 line.product_uom_qty += 1
    #             else:
    #                 # If it's a new lot, add it and increase quantity
    #                 line.dummy_lot_ids = [(4, lot.id)]
    #                 line.product_uom_qty += 1
    #         else:
    #             # Create new move line with the lot
    #             self.move_ids_without_package = [(0, 0, {
    #                 'product_id': product.id,
    #                 'product_uom_qty': 1,
    #                 'location_id': self.location_id.id,
    #                 'location_dest_id': self.location_dest_id.id,
    #                 'name': product.display_name,
    #                 'dummy_lot_ids': [(4, lot.id)],
    #             })]
    #
    #     # --------------------- GS1 Barcode branch ---------------------
    #     if re.match(gs1_pattern, barcode):
    #         match = re.match(gs1_pattern, barcode)
    #         product_barcode = match.group(1)
    #         lot_or_serial = match.group(2)
    #         product = search_product('barcode', product_barcode)
    #         if not product:
    #             raise ValidationError(f'No product found with barcode {product_barcode}')
    #
    #         if product.tracking == 'lot':
    #             # Find the lot for lot-tracked products
    #             lot = self.env['stock.lot'].search(
    #                 [('product_id', '=', product.id), ('name', '=', lot_or_serial)], limit=1)
    #             if not lot:
    #                 raise ValidationError(f'No lot found with number: {lot_or_serial}')
    #             update_lot_tracked(product, lot)
    #         else:
    #             # Serial-tracked: perform serial uniqueness validations.
    #             pickings = self.env['stock.picking'].search([
    #                 ('picking_type_code', 'in', ['internal', 'outgoing']),
    #                 ('state', '!=', 'cancel'),
    #                 ('move_line_ids.lot_id.name', '=', lot_or_serial),
    #             ])
    #             if pickings:
    #                 picking_names = ', '.join(pickings.mapped('name'))
    #                 raise ValidationError(
    #                     f'Serial number {lot_or_serial} is already assigned in pickings: {picking_names}')
    #
    #             line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
    #             if line:
    #                 lot = self.env['stock.lot'].search(
    #                     [('product_id', '=', product.id), ('name', '=', lot_or_serial)], limit=1)
    #                 if not lot:
    #                     raise ValidationError(f'No serial number found: {lot_or_serial}')
    #                 if lot not in line.dummy_lot_ids:
    #                     line.dummy_lot_ids = [(4, lot.id)]
    #                 line.product_uom_qty = len(line.dummy_lot_ids)
    #             else:
    #                 lot = self.env['stock.lot'].search(
    #                     [('product_id', '=', product.id), ('name', '=', lot_or_serial)], limit=1)
    #                 if not lot:
    #                     raise ValidationError(f'No serial number found: {lot_or_serial}')
    #                 self.move_ids_without_package = [(0, 0, {
    #                     'product_id': product.id,
    #                     'product_uom_qty': 1,
    #                     'dummy_lot_ids': [(4, lot.id)],
    #                     'location_id': self.location_id.id,
    #                     'location_dest_id': self.location_dest_id.id,
    #                     'name': product.display_name,
    #                 })]
    #
    #     # --------------------- EAN-13 Barcode branch ---------------------
    #     elif re.match(ean13_pattern, barcode):
    #         match = re.match(ean13_pattern, barcode)
    #         ean13_barcode = match.group(1)
    #         lots = self.env['stock.lot'].search([('ref', '=', ean13_barcode), ('product_qty', '>', 0)])
    #         if not lots:
    #             raise ValidationError(f'No lots found with EAN-13 barcode {ean13_barcode} or insufficient quantity')
    #         product = lots[0].product_id
    #         if not product:
    #             raise ValidationError(f'No product associated with lots for barcode {ean13_barcode}')
    #
    #         if product.tracking == 'lot':
    #             available_lots = lots.filtered(
    #                 lambda l: l.name not in self.move_ids_without_package.mapped('dummy_lot_ids.name'))
    #             if not available_lots:
    #                 raise ValidationError(f'All lot numbers for barcode {ean13_barcode} have been assigned or used.')
    #             next_lot = available_lots[0]
    #             update_lot_tracked(product, next_lot)
    #         else:
    #             available_lots = lots.filtered(
    #                 lambda l: l.name not in self.move_ids_without_package.mapped('dummy_lot_ids.name'))
    #             if not available_lots:
    #                 raise ValidationError(f'All serial numbers for barcode {ean13_barcode} have been assigned or used.')
    #             next_lot = available_lots[0]
    #             line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
    #             if line:
    #                 line.dummy_lot_ids = [(4, next_lot.id)]
    #                 line.product_uom_qty = len(line.dummy_lot_ids)
    #             else:
    #                 self.move_ids_without_package = [(0, 0, {
    #                     'product_id': product.id,
    #                     'product_uom_qty': 1,
    #                     'location_id': self.location_id.id,
    #                     'location_dest_id': self.location_dest_id.id,
    #                     'name': product.display_name,
    #                     'dummy_lot_ids': [(4, next_lot.id)],
    #                 })]
    #
    #     # --------------------- Custom Serial branch ---------------------
    #     elif re.match(custom_serial_pattern, barcode):
    #         match = re.match(custom_serial_pattern, barcode)
    #         prefix = match.group(1)
    #         lot = self.env['stock.lot'].search([('name', '=like', f'{prefix}%')], limit=1)
    #         if not lot:
    #             raise ValidationError(f'No lot found with serial number prefix {prefix}')
    #         product = lot.product_id
    #         if product.tracking == 'lot':
    #             update_lot_tracked(product, lot)
    #         else:
    #             line = self.move_ids_without_package.filtered(lambda l: l.product_id == product)
    #             if line:
    #                 line.dummy_lot_ids = [(4, lot.id)]
    #                 print("line.dummy_lot_ids", line.dummy_lot_ids)
    #                 line.product_uom_qty = len(line.dummy_lot_ids)
    #             else:
    #                 self.move_ids_without_package = [(0, 0, {
    #                     'product_id': product.id,
    #                     'product_uom_qty': 1,
    #                     'dummy_lot_ids': [(4, lot.id)],
    #                     'location_id': self.location_id.id,
    #                     'location_dest_id': self.location_dest_id.id,
    #                     'name': product.display_name,
    #                 })]
    #     else:
    #         raise ValidationError('Invalid barcode format')
    #
    #     self.stock_barcode = False
    @api.onchange('stock_barcode')
    def _onchange_stock_barcode(self):
        if not self.stock_type:
            if self.stock_barcode:
                raise ValidationError('Please select a Stock Type before scanning a barcode.')
            return

        barcode = str(self.stock_barcode or '').strip()

        # Barcode patterns
        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'
        ean13_pattern = r'(\d{13})'
        custom_serial_pattern = r'^(R\d+)'

        def search_product(field, value):
            """Helper: Search for product in product.product and product.template."""
            product = self.env['product.product'].search([(field, '=', value)], limit=1)
            if not product:
                template = self.env['product.template'].search([(field, '=', value)], limit=1)
                if template:
                    product = template.product_variant_id
            return product

        def create_move_line(product, lot, qty):
            """Helper: Create a new move line with the given quantity and lot.
            Uses the .new() method and union operator to avoid mixing recordset with command lists.
            """
            new_line_vals = {
                'product_id': product.id,
                'product_uom_qty': qty,
                'location_id': self.location_id.id,
                'location_dest_id': self.location_dest_id.id,
                'name': product.display_name,
                'dummy_lot_ids': [(4, lot.id)],
            }
            new_line = self.env['stock.move'].new(new_line_vals)
            self.move_ids_without_package |= new_line

        # --------------------- GS1 Barcode branch ---------------------
        if re.match(gs1_pattern, barcode):
            match = re.match(gs1_pattern, barcode)
            product_barcode = match.group(1)
            lot_or_serial = match.group(2)
            product = search_product('barcode', product_barcode)
            if not product:
                raise ValidationError(f'No product found with barcode {product_barcode}')

            if product.tracking == 'lot':
                # For lot-tracked products, ensure lot_qty is provided and > 0.
                if not hasattr(self, 'lot_qty') or self.lot_qty <= 0:
                    raise ValidationError('Please provide a valid quantity greater than 0 for lot tracked products.')
                location = False
                if self.stock_picking_type in ['main_damage', 'return']:
                    location = self.env.ref('stock.stock_location_stock').id
                elif self.stock_picking_type in ['damage_main', 'damage']:
                    location = self.env['stock.location'].search([('name', 'ilike', '-DM')], limit=1).id
                elif self.stock_picking_type in ['return_main']:
                    location = self.env['stock.location'].search([('name', 'ilike', '-RE')], limit=1).id
                # Search lots by EAN reference in stock.quants
                lots = self.env['stock.quant'].search([
                    ('lot_id.ref', '=', lot_or_serial),
                    ('quantity', '>', 0),
                    ('location_id', '=', location)
                ])
                lot = lots.lot_id
                if not lot:
                    raise ValidationError(f'No lot found with number: {lot_or_serial}')
                create_move_line(product, lot, self.lot_qty)
            else:
                # Serial-tracked: validate uniqueness first.
                pickings = self.env['stock.picking'].search([
                    ('picking_type_code', 'in', ['internal', 'outgoing']),
                    ('state', '!=', 'cancel'),
                    ('move_line_ids.lot_id.name', '=', lot_or_serial),
                ])
                if pickings:
                    picking_names = ', '.join(pickings.mapped('name'))
                    raise ValidationError(
                        f'Serial number {lot_or_serial} is already assigned in pickings: {picking_names}')
                location = False
                if self.stock_picking_type in ['main_damage', 'return']:
                    location = self.env.ref('stock.stock_location_stock').id
                elif self.stock_picking_type in ['damage_main', 'damage']:
                    location = self.env['stock.location'].search([('name', 'ilike', '-DM')], limit=1).id
                elif self.stock_picking_type in ['return_main']:
                    location = self.env['stock.location'].search([('name', 'ilike', '-RE')], limit=1).id
                # Search lots by EAN reference in stock.quants
                lots = self.env['stock.quant'].search([
                    ('lot_id.ref', '=', lot_or_serial),
                    ('quantity', '>', 0),
                    ('location_id', '=', location)
                ])
                lot = lots.lot_id
                if not lot:
                    raise ValidationError(f'No serial number found: {lot_or_serial}')
                create_move_line(product, lot, 1)

        # --------------------- EAN-13 Barcode branch ---------------------
        elif re.match(ean13_pattern, barcode):
            match = re.match(ean13_pattern, barcode)
            ean13_barcode = match.group(1)

            # Determine location based on stock_picking_type
            location = False
            if self.stock_picking_type in ['main_damage', 'return']:
                location = self.env.ref('stock.stock_location_stock').id
            elif self.stock_picking_type in ['damage_main', 'damage']:
                location = self.env['stock.location'].search([('name', 'ilike', '-DM')], limit=1).id
            elif self.stock_picking_type in ['return_main']:
                location = self.env['stock.location'].search([('name', 'ilike', '-RE')], limit=1).id

            # Search lots by EAN reference in stock.quants
            lots = self.env['stock.quant'].search([
                ('lot_id.ref', '=', ean13_barcode),
                ('quantity', '>', 0),
                ('location_id', '=', location)
            ])
            if not lots:
                raise ValidationError(f'No lots found with EAN-13 barcode {ean13_barcode} or insufficient quantity.')

            product = lots[0].product_id
            if not product:
                raise ValidationError(f'No product associated with lots for barcode {ean13_barcode}.')

            # Separate logic based on tracking type
            if product.tracking == 'serial':
                used_serials = set(self.move_ids_without_package.mapped('dummy_lot_ids.name'))
                used_serials |= set(self.env['stock.picking'].search([
                    ('stock_picking_type', '=', self.stock_picking_type),
                    ('state', '!=', 'cancel'),
                    ('move_ids_without_package.dummy_lot_ids.name', '!=', False)
                ]).mapped('move_ids_without_package.dummy_lot_ids.name'))

                available_lots = lots.lot_id.filtered(lambda l: l.name not in used_serials)
                if not available_lots:
                    raise ValidationError(f'All serial numbers for barcode {ean13_barcode} have been assigned or used.')
                create_move_line(product, available_lots[0], 1)

            elif product.tracking == 'lot':
                if not hasattr(self, 'lot_qty') or self.lot_qty <= 0:
                    raise ValidationError('Please provide a valid quantity greater than 0 for lot tracked products.')

                remaining_qty = self.lot_qty
                for quant in lots:
                    lot = quant.lot_id
                    available_qty = quant.quantity
                    # Quantity already used in this picking
                    current_qty = sum(
                        ml.product_uom_qty for ml in self.move_ids_without_package
                        if lot.name in ml.dummy_lot_ids.name
                    )
                    # Quantity already used in other open pickings of same type
                    other_qty = sum(
                        ml.product_uom_qty for ml in self.env['stock.move'].search([
                            ('dummy_lot_ids.name', '=', lot.name),
                            ('move_picking_type', '=', self.stock_picking_type),
                            ('picking_id.state', 'not in', ['cancel', 'done']),
                        ])
                    )
                    available_qty -= (current_qty + other_qty)
                    if available_qty <= 0:
                        continue
                    assign_qty = min(available_qty, remaining_qty)
                    if assign_qty > 0:
                        create_move_line(product, lot, assign_qty)
                        remaining_qty -= assign_qty
                    if remaining_qty <= 0:
                        break
                if remaining_qty > 0:
                    raise ValidationError(
                        f"Not enough quantity available across lots for barcode {ean13_barcode}. "
                        f"Remaining qty needed: {remaining_qty}"
                    )

        # --------------------- Custom Serial branch ---------------------
        elif re.match(custom_serial_pattern, barcode):
            StockLot = self.env['stock.lot']
            # STEP 1: Search for lot by ref = scanned barcode
            location_id = False
            if self.stock_picking_type in ['main_damage', 'return']:
                location_id = self.env.ref('stock.stock_location_stock').id
            elif self.stock_picking_type in ['damage_main', 'damage']:
                location_id = self.env['stock.location'].search([('name', 'ilike', '-DM')], limit=1).id
            elif self.stock_picking_type in ['return_main']:
                location_id = self.env['stock.location'].search([('name', 'ilike', '-RE')], limit=1).id
            lots_with_ref = self.env['stock.quant'].search([
                '|',
                ('lot_id.ref', '=', barcode),
                ('lot_id.name', '=', barcode),
                ('location_id', '=', location_id),
                ('quantity', '>', 0),
            ])
            if not lots_with_ref:
                raise ValidationError(f"No lot found with ref '{barcode}'.")
            product = lots_with_ref[0].product_id
            if not product:
                raise ValidationError(f"No product associated with lot ref '{barcode}'.")
            if product.tracking not in ('serial', 'lot'):
                raise ValidationError(f"Product '{product.display_name}' is not tracked.")
            # In this case, since ref is unique per lot, we just reuse the original search result
            all_matching_lots = lots_with_ref
            # STEP 3: Exclude already used lots (for serials)
            used_names = set(self.move_ids_without_package.mapped('dummy_lot_ids.name'))
            used_names |= set(self.env['stock.picking'].search([
                ('stock_picking_type', '=', self.stock_picking_type), ('state', '!=', 'cancel'),
                '|',
                ('move_ids_without_package.dummy_lot_ids.ref', '=', barcode),
                ('move_ids_without_package.dummy_lot_ids.name', '=', barcode),
            ]).mapped('move_ids_without_package.dummy_lot_ids.name'))
            available_lot = None
            for lot in all_matching_lots:
                if product.tracking == 'serial':
                    if lot.lot_id.name not in used_names:
                        available_lot = lot.lot_id
                        qty = 1
                        break
                elif product.tracking == 'lot':
                    if not self.lot_qty or self.lot_qty <= 0:
                        raise ValidationError("Please enter a valid quantity for lot tracked product.")
                    existing_qty = sum(
                        ml.product_uom_qty
                        for ml in self.move_ids_without_package
                        for l in ml.dummy_lot_ids
                        if l.name == lot.lot_id.name
                    )
                    other_qty = sum(
                        ml.product_uom_qty
                        for ml in self.env['stock.move'].search([
                            ('dummy_lot_ids.name', '=', lot.lot_id.name),
                            ('move_picking_type', '=', self.stock_picking_type),
                            ('picking_id.state', 'not in', ['cancel']),  # include draft, assigned, confirmed, done
                        ])
                    )

                    total_qty = existing_qty + other_qty

                    if total_qty + self.lot_qty > lot.quantity:
                        raise ValidationError(
                            f"Lot {lot.lot_id.name} exceeds available quantity "
                            f"(Given {total_qty + self.lot_qty} Available {lot.quantity}).")
                    # if (existing_qty + self.lot_qty) <= lot.quantity:
                    else:
                        available_lot = lot.lot_id
                        qty = self.lot_qty
                        break
            if not available_lot:
                raise ValidationError(f"All lots with ref '{barcode}' are either used or exceed available quantity.")

            create_move_line(product, available_lot, qty)
        else:
            raise ValidationError('Invalid barcode format')
        self.stock_barcode = False
        self.lot_qty = False


    def reset_product_lines(self):
        for rec in self:
            rec.move_ids_without_package.unlink()


class PickingType(models.Model):
    _inherit = "stock.picking.type"

    stock_picking_type = fields.Selection([('exchange', 'Customer-Return'), ('receipt', 'Receipt')
                                              , ('delivery', 'Delivery'), ('pos_order', 'POS Order'),
                                           ('manufacturing', 'Manufacturing'),
                                           ('regular', 'Regular'), ('damage', 'Damage'), ('return', 'Return'),
                                           ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
                                           ('return_main', 'Return-Main')], string='Type', tracking=True)


class StockMove(models.Model):
    """Inherited stock.move class to add fields and functions"""
    _inherit = "stock.move"

    serial_no = fields.Char(string="Serial No")
    dummy_lot_ids = fields.Many2many('stock.lot', string="Ref S.No")
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    pos_order_lines = fields.Many2one('pos.order.line', string='pos order lines', copy=False)
    nhcl_tax_ids = fields.Many2many('account.tax', 'nhcl_tax',
                                    domain=[('type_tax_use', '=', 'sale'), ('active', '=', True)],
                                    string="Tax")
    nhcl_total = fields.Float(string="Total", copy=False)
    nhcl_rsp = fields.Float(string="RSP", copy=False)
    nhcl_exchange = fields.Boolean(string="Exchange", copy=False)
    nhcl_discount = fields.Float(string="Discount (%)", copy=False)
    nhcl_gdiscount = fields.Float(string="Global Discount (%)", copy=False)
    nhcl_disc_lines = fields.Char(string="Disc", copy=False)
    nhcl_price_total = fields.Monetary(compute='_compute_amount', string='Total', store=True)
    nhcl_price_subtotal = fields.Monetary(compute='_compute_amount', string='Subtotal', store=True)
    nhcl_price_tax = fields.Float(compute='_compute_amount', string='Tax', store=True)
    currency_id = fields.Many2one("res.currency", string="Currency", required=True,
                                  related='picking_id.currency_id')
    ref_pos_order_line_id = fields.Integer('Pos Order Line Id', default="0", copy=False)
    move_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')
    s_no = fields.Integer(string="S.No", compute="_compute_s_no")
    nhcl_cost_price = fields.Float(string="Cost Price", copy=False)


    @api.depends('picking_id')
    def _compute_s_no(self):
        for rec in self:
            if rec.picking_id and rec.id in rec.picking_id.move_ids_without_package.ids:
                rec.s_no = rec.picking_id.move_ids_without_package.ids.index(rec.id)+1
            else:
                rec.s_no = 0



    def _get_new_picking_values(self):
        res = super(StockMove, self)._get_new_picking_values()
        if res.get('origin'):
            sale_id = self.env['sale.order'].search([('name', '=', res.get('origin'))], limit=1)
            if sale_id:
                res.update({
                    'stock_type': sale_id.so_type,
                })
        return res

    def picking_unlink(self):
        for move in self:
            move.unlink()

    @api.model
    def _prepare_merge_moves_distinct_fields(self):
        distinct_fields = super(StockMove, self)._prepare_merge_moves_distinct_fields()
        distinct_fields.append('pos_order_lines')
        distinct_fields.append('ref_pos_order_line_id')
        distinct_fields.append('dummy_lot_ids')
        return distinct_fields

    @api.depends('quantity', 'nhcl_rsp', 'nhcl_tax_ids')
    def _compute_amount(self):
        for line in self:
            tax_results = self.env['account.tax']._compute_taxes([line._convert_to_tax_base_line_dict()])
            totals = next(iter(tax_results['totals'].values()))
            amount_untaxed = totals['amount_untaxed']
            amount_tax = totals['amount_tax']
            line.update({
                'nhcl_price_subtotal': amount_untaxed,
                'nhcl_price_tax': amount_tax,
                'nhcl_price_total': amount_untaxed + amount_tax,
            })

    # updating the price unit,currency,req qty,product,partner
    def _convert_to_tax_base_line_dict(self):
        # Hook method to returns the different argument values for the
        # compute_all method, due to the fact that discounts mechanism
        # is not implemented yet on the purchase orders.
        # This method should disappear as soon as this feature is
        # also introduced like in the sales module.
        self.ensure_one()
        return self.env['account.tax']._convert_to_tax_base_line_dict(
            self,
            price_unit=self.nhcl_rsp * (1 - (self.nhcl_discount or 0.0) / 100.0) * (
                        1 - (self.nhcl_gdiscount or 0.0) / 100.0),
            currency=self.picking_id.currency_id,
            quantity=self.quantity,
            product=self.product_id,
            taxes=self.nhcl_tax_ids,
            partner=self.picking_id.partner_id,
            price_subtotal=self.nhcl_price_subtotal,
        )

    @api.onchange('product_id')
    def _onchange_product_id_stock(self):
        if self.picking_id and not self.picking_id.stock_type:
            # Clear the product_id and raise an error if no stock_type is selected
            self.product_id = False
            raise ValidationError(
                "You must select a Stock Type before selecting a product."
            )

    def action_assign_serial(self):
        if self.picking_id.picking_type_id.code == 'incoming' and self.picking_id.stock_type == 'ho_operation':
            raise ValidationError(
                _("You are not allowed to Assign the Serial Number For Product %s") % (self.product_id.name))
        else:
            return super(StockMove, self).action_assign_serial()

    def _update_reserved_quantity(self, need, location_id, lot_id=None, quant_ids=None, package_id=None, owner_id=None,
                                  strict=True):
        if self._context.get("sol_lot_id"):
            # Use sale line's lots if available, otherwise fallback to dummy lots
            if self.sale_line_id and self.sale_line_id.lot_ids:
                lot_id = self.sale_line_id.lot_ids
            elif self.dummy_lot_ids:
                lot_id = self.dummy_lot_ids
        return super()._update_reserved_quantity(
            need, location_id,
            lot_id=lot_id,
            quant_ids=quant_ids,
            package_id=package_id,
            owner_id=owner_id,
            strict=strict
        )

    def _prepare_move_line_vals(self, quantity=None, reserved_quant=None):
        vals = super()._prepare_move_line_vals(quantity=quantity, reserved_quant=reserved_quant)
        sale_line = self.sale_line_id
        if sale_line:
            if self.product_id.tracking == 'lot' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids.id
            elif self.product_id.tracking == 'serial' and sale_line.lot_ids:
                vals['lot_id'] = sale_line.lot_ids[0].id
        if self.serial_no and self.picking_id.stock_picking_type == 'exchange' and self.picking_id.company_type == 'same':
            vals['lot_name'] = self.serial_no
            lot = self.dummy_lot_ids[0]
            vals['cost_price'] = lot.cost_price
        if self.picking_id.move_type == 'direct' and self.dummy_lot_ids:
            vals['lot_id'] =  self.dummy_lot_ids.id
        return vals



class StockMoveLine(models.Model):
    _inherit = "stock.move.line"


    internal_ref_lot = fields.Char(string="Barcode", copy=False, tracking=True)

    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    descrip_9 = fields.Many2one('product.attribute.value', string='Discount', copy=False,
                                domain=[('attribute_id.name', '=', 'Discount')])
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, tracking=True, store=True)
    move_line_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')

    s_no = fields.Integer(string="S.No", compute="_compute_s_no")

    @api.depends('batch_id')
    def _compute_s_no(self):
        for rec in self:
            if rec.batch_id and rec.id in rec.batch_id.move_line_ids.ids:
                rec.s_no = rec.batch_id.move_line_ids.ids.index(rec.id) + 1
            else:
                rec.s_no = 0


    def compute_get_unit_price(self):
        for rec in self:
            if rec.picking_id.picking_type_id.stock_picking_type != 'receipt':
                if rec.lot_id:
                    rec.cost_price = rec.lot_id.cost_price
                    rec.internal_ref_lot = rec.lot_id.ref
                    rec.type_product = rec.lot_id.type_product
                    rec.mr_price = rec.lot_id.mr_price
                    rec.rs_price = rec.lot_id.rs_price
                    rec.segment = rec.lot_id.segment
                    rec.categ_1 = rec.lot_id.category_1
                    rec.categ_2 = rec.lot_id.category_2
                    rec.categ_3 = rec.lot_id.category_3
                    rec.categ_4 = rec.lot_id.category_4
                    rec.categ_5 = rec.lot_id.category_5
                    rec.categ_6 = rec.lot_id.category_6
                    rec.categ_7 = rec.lot_id.category_7
                    rec.categ_8 = rec.lot_id.category_8
                    rec.descrip_1 = rec.lot_id.description_1
                    rec.descrip_2 = rec.lot_id.description_2
                    rec.descrip_3 = rec.lot_id.description_3
                    rec.descrip_4 = rec.lot_id.description_4
                    rec.descrip_5 = rec.lot_id.description_5
                    rec.descrip_6 = rec.lot_id.description_6
                    rec.descrip_7 = rec.lot_id.description_7
                    rec.descrip_8 = rec.lot_id.description_8
                else:
                    rec.cost_price = 0.0

    def get_product_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_attribute_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith('Color'):
                        rec.categ_1 = j.id

    @api.model
    def create(self, vals):
        record = super(StockMoveLine, self).create(vals)
        record.get_product_attributes()
        record.compute_get_unit_price()
        if record.move_id and record.move_id.product_id.id == record.product_id.id and record.picking_id.stock_picking_type == 'receipt':
            record.move_id.type_product = record.type_product
        return record

    def write(self, vals):
        res = super(StockMoveLine, self).write(vals)
        for move_line in self:
            if move_line.lot_id:
                lot_values = {}
                for categ_field in ['categ_1', 'categ_2', 'categ_3', 'categ_4', 'categ_5', 'categ_6', 'categ_7',
                                    'categ_8']:
                    if categ_field in vals:
                        lot_values[categ_field.replace('categ', 'category')] = vals[categ_field]
                for desc_field in ['descrip_1', 'descrip_2', 'descrip_3', 'descrip_4', 'descrip_5', 'descrip_6',
                                   'descrip_7', 'descrip_8']:
                    if desc_field in vals:
                        lot_values[desc_field.replace('descrip', 'description')] = vals[desc_field]
                for price_field in ['cost_price', 'mr_price', 'rs_price']:
                    if price_field in vals:
                        lot_values[price_field] = vals[price_field]
                if 'internal_ref_lot' in vals:
                    lot_values['ref'] = vals['internal_ref_lot']
                if lot_values:
                    move_line.lot_id.write(lot_values)
        return res

    def _create_and_assign_production_lot(self):
        res = super(StockMoveLine, self)._create_and_assign_production_lot()
        if self.picking_id.picking_type_id.code == 'incoming':
            for rec in self:
                rec.lot_id.write({
                    'type_product': rec.type_product,
                    'picking_id': rec.picking_id.id,
                    'ref': rec.internal_ref_lot,
                    'category_1': rec.categ_1,
                    'category_2': rec.categ_2,
                    'category_3': rec.categ_3,
                    'category_4': rec.categ_4,
                    'category_5': rec.categ_5,
                    'category_6': rec.categ_6,
                    'category_7': rec.categ_7,
                    'category_8': rec.categ_8,
                    'description_1': rec.descrip_1,
                    'description_2': rec.descrip_2,
                    'description_3': rec.descrip_3,
                    'description_4': rec.descrip_4,
                    'description_5': rec.descrip_5,
                    'description_6': rec.descrip_6,
                    'description_7': rec.descrip_7,
                    'description_8': rec.descrip_8,
                    'mr_price': rec.mr_price,
                    'cost_price': rec.cost_price,
                    'rs_price': rec.rs_price,
                    # 'product_aging': rec.product_aging,
                    'segment': rec.segment,
                    # 'transfer_price': rec.transfer_price,
                    # 'transfer_percent': rec.transfer_percent,
                })
                if rec.type_product == 'brand':
                    # Search for existing barcode
                    existing_barcode = self.env['product.barcode'].sudo().search(
                        [('barcode', '=', rec.internal_ref_lot)],
                        limit=1)

                    if rec.picking_id.picking_type_id.code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                        if existing_barcode:
                            # Increment nhcl_inward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_inward_qty to 1
                            self.env['product.barcode'].sudo().create({
                                'barcode': rec.internal_ref_lot,
                                'product_id': rec.lot_id.product_id.id,
                                'nhcl_inward_qty': 1,
                            })

                    elif rec.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                        if existing_barcode:
                            # Increment nhcl_outward_qty for existing barcode
                            existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                        else:
                            # Create a new barcode and set nhcl_outward_qty to 1
                            self.env['product.barcode'].sudo().create({
                                'barcode': rec.internal_ref_lot,
                                'product_id': rec.lot_id.product_id.id,
                                'nhcl_outward_qty': 1,
                            })

        return res

    @api.onchange('internal_ref_lot')
    def sending_no_to_lot(self):
        for rec in self:
            if rec.lot_id and rec.internal_ref_lot:
                # Set the reference on the lot
                # rec.lot_id.ref = rec.internal_ref_lot
                # Search for existing barcode
                existing_barcode = self.env['product.barcode'].sudo().search([('barcode', '=', rec.internal_ref_lot)],
                                                                             limit=1)

                if rec.picking_id.picking_type_id.code == 'incoming':  # If the picking code is for incoming transfers (receipts)
                    if existing_barcode:
                        # Increment nhcl_inward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_inward_qty': existing_barcode.nhcl_inward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_inward_qty to 1
                        self.env['product.barcode'].sudo().create({
                            'barcode': rec.internal_ref_lot,
                            'product_id': rec.lot_id.product_id.id,
                            'nhcl_inward_qty': 1,
                        })

                elif rec.picking_id.picking_type_id.code == 'outgoing':  # If the picking code is for outgoing transfers (deliveries)
                    if existing_barcode:
                        # Increment nhcl_outward_qty for existing barcode
                        existing_barcode.sudo().write({'nhcl_outward_qty': existing_barcode.nhcl_outward_qty + 1})
                    else:
                        # Create a new barcode and set nhcl_outward_qty to 1
                        self.env['product.barcode'].sudo().create({
                            'barcode': rec.internal_ref_lot,
                            'product_id': rec.lot_id.product_id.id,
                            'nhcl_outward_qty': 1,
                        })


class StockLot(models.Model):
    """Inherited stock.lot class to add fields and functions"""
    _inherit = 'stock.lot'

    category_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                                 domain=[('attribute_id.name', '=', 'Color')])
    category_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                                 domain=[('attribute_id.name', '=', 'Fit')])
    category_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                                 domain=[('attribute_id.name', '=', 'Brand')])
    category_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                                 domain=[('attribute_id.name', '=', 'Pattern')])
    category_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Type')])
    category_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Border Size')])
    category_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                                 domain=[('attribute_id.name', '=', 'Size')])
    category_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                                 domain=[('attribute_id.name', '=', 'Design')])
    description_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    description_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                    domain=[('attribute_id.name', '=', 'Range')])
    description_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                    domain=[('attribute_id.name', '=', 'Collection')])
    description_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                    domain=[('attribute_id.name', '=', 'Fabric')])
    description_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                    domain=[('attribute_id.name', '=', 'Exclusive')])
    description_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                    domain=[('attribute_id.name', '=', 'Print')])
    description_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                    domain=[('attribute_id.name', '=', 'Days Ageing')])
    description_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    description_9 = fields.Many2one('product.attribute.value', string='Discount', copy=False,
                                    domain=[('attribute_id.name', '=', 'Discount')])
    product_description = fields.Html(string="Product Description", copy=False)
    web_product = fields.Char(string="Website Product Name", copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    # transfer_price = fields.Float(string='TRP', copy=False)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)
    picking_id = fields.Many2one('stock.picking', string="GRC No", copy=False)
    # transfer_percent = fields.Float(string="TP %", copy=False)
    is_used = fields.Boolean(string='POS Posted')
    # product_aging = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, tracking=True, store=True)
    ref = fields.Char('Barcode',
                      help="Internal reference number in case it differs from the manufacturer's lot/serial number")
    ho_grc_no = fields.Char(string='HO GRC NO.')
    serial_type = fields.Selection([('regular', 'Regular'), ('return', 'Returned')],
                                   string='Serial Type', copy=False, tracking=True, default='regular')
    is_uploaded = fields.Boolean('Is Uploaded', copy=False)

    @api.onchange('category_1')
    def updating_line_to_lot_category_1(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_1 = self.category_1

    @api.onchange('category_2')
    def updating_line_to_lot_category_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_2 = self.category_2

    @api.onchange('category_3')
    def updating_line_to_lot_category_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_3 = self.category_3

    @api.onchange('category_4')
    def updating_line_to_lot_category_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_4 = self.category_4

    @api.onchange('category_5')
    def updating_line_to_lot_category_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_5 = self.category_5

    @api.onchange('category_6')
    def updating_line_to_lot_category_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_6 = self.category_6

    @api.onchange('category_7')
    def updating_line_to_lot_category_7(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_7 = self.category_7

    @api.onchange('category_8')
    def updating_line_to_lot_category_8(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.categ_8 = self.category_8

    @api.onchange('description_1')
    def updating_line_to_lot_description_1(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_1 = self.description_1

    @api.onchange('description_2')
    def updating_line_to_lot_description_2(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_2 = self.description_2

    @api.onchange('description_3')
    def updating_line_to_lot_description_3(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_3 = self.description_3

    @api.onchange('description_4')
    def updating_line_to_lot_description_4(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_4 = self.description_4

    @api.onchange('description_5')
    def updating_line_to_lot_description_5(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_5 = self.description_5

    @api.onchange('description_6')
    def updating_line_to_lot_description_6(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_6 = self.description_6

    @api.onchange('description_7')
    def updating_line_to_lot_description_7(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_7 = self.description_7

    @api.onchange('description_8')
    def updating_line_to_lot_description_8(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.descrip_8 = self.description_8

    @api.onchange('mr_price')
    def updating_line_to_lot_mr_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.mr_price = self.mr_price

    @api.onchange('rs_price')
    def updating_line_to_lot_rs_price(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.rs_price = self.rs_price

    # @api.onchange('transfer_price')
    # def updating_line_to_lot_transfer_price(self):
    # detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
    # detail.transfer_price = self.transfer_price

    @api.onchange('transfer_percent')
    def updating_line_to_lot_transfer_percent(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.transfer_percent = self.transfer_percent

    @api.onchange('ref')
    def updating_line_to_lot_ref(self):
        detail = self.env['stock.move.line'].search([('lot_id.name', '=', self.name)])
        detail.internal_ref_lot = self.ref

    def get_attributes(self):
        for rec in self:
            val = rec.product_id.product_template_attribute_value_ids
            for i in val:
                attribute = self.env['product.attribute.value'].search([('name', '=', i.name)])
                for j in attribute:
                    if j and j.attribute_id.name == i.attribute_id.name and i.attribute_id.name.startswith(
                            'Color'):
                        rec.category_1 = j.id

    @api.model
    def create(self, vals):
        record = super(StockLot, self).create(vals)
        record.get_attributes()
        return record

    def search_by_loyalty_rule(self, loyalty_rule):
        domain = []
        today = datetime.today()

        # Get the user's timezone
        user_tz = self.env.user.tz or pytz.utc
        local = pytz.timezone(user_tz)

        # Get the selected ageing slab range
        if loyalty_rule.day_ageing_slab:
            # Mapping for ageing slabs
            slab_mapping = {
                '1': (0, 30),
                '2': (30, 60),
                '3': (60, 90),
                '4': (90, 120),
                '5': (120, 150),
                '6': (150, 180),
                '7': (180, 210),
                '8': (210, 240),
                '9': (240, 270),
                '10': (270, 300),
                '11': (300, 330),
                '12': (330, 360)
            }
            # Get the start and end days for the slab
            slab_start, slab_end = slab_mapping.get(str(loyalty_rule.day_ageing_slab), (0, 360))

            # Calculate the lower and upper bounds for the matching date range
            ageing_date_start = today - timedelta(days=slab_end)
            ageing_date_end = today - timedelta(days=slab_start)

            # Ensure the start date is earlier than the end date
            if ageing_date_start > ageing_date_end:
                ageing_date_start, ageing_date_end = ageing_date_end, ageing_date_start

            # Localize the dates to the user's timezone
            from_date_local = ageing_date_start.replace(hour=0, minute=0, second=0, microsecond=0)
            to_date_local = ageing_date_end.replace(hour=23, minute=59, second=59, microsecond=999999)

            from_date_local = local.localize(from_date_local)
            to_date_local = local.localize(to_date_local)

            # Convert the localized dates to UTC
            from_date_utc = from_date_local.astimezone(pytz.utc)
            to_date_utc = to_date_local.astimezone(pytz.utc)

            # Format the dates into ISO 8601 format
            from_date_str = from_date_utc.strftime("%Y-%m-%dT%H:%M:%S")
            to_date_str = to_date_utc.strftime("%Y-%m-%dT%H:%M:%S")

            # Add create_date range condition to the domain
            domain.append(('create_date', '>=', from_date_str))
            domain.append(('create_date', '<=', to_date_str))
        # Loop through all category and description fields
        for i in range(1, 8):
            # Dynamically construct field names
            category_field = f'category_{i}'
            description_field = f'description_{i}'
            # Get the corresponding many2many fields in loyalty.rule
            category_rule_field = f'category_{i}_ids'
            description_rule_field = f'description_{i}_ids' if i != 7 else None

            # Add to domain if the loyalty rule fields have values
            if getattr(loyalty_rule, category_rule_field):
                domain.append((category_field, 'in', getattr(loyalty_rule, category_rule_field).ids))
            if description_rule_field and getattr(loyalty_rule, description_rule_field):
                domain.append((description_field, 'in', getattr(loyalty_rule, description_rule_field).ids))

        # Add product filtering if ref_product_ids is defined in the loyalty rule
        if loyalty_rule.ref_product_ids:
            domain.append(('product_id', 'in', loyalty_rule.ref_product_ids.ids))
        # Add category filtering if product_category_id is defined in the loyalty rule
        if loyalty_rule.product_category_ids:
            selected_categories = loyalty_rule.product_category_id.ids
            for category in loyalty_rule.product_category_ids:
                selected_categories += category.search([('id', 'child_of', category.id)]).ids
            selected_categories = list(set(selected_categories))
            domain.append(('product_id.categ_id', 'in', selected_categories))
        # Add product tag filtering if product_tag_id is defined in the loyalty rule
        if loyalty_rule.product_tag_id:
            domain.append(('product_id.product_tag_ids', '=', loyalty_rule.product_tag_id.id))
        # Additional checks for stock.lot records
        domain.append(('product_qty', '>=', 1))
        domain.append(('product_id.item_type', '=', 'inventory'))
        # Add company filtering if company_id is defined
        if self.company_id:
            domain.append(('company_id', '=', self.company_id.id))
        lots = self.env['stock.lot'].search(domain)
        return lots

    @api.model
    def _get_next_serial(self, company, product):
        """Return the next serial number to be attributed to the product."""
        if product.tracking == "serial":
            auto_generate_sequence = self.env['nhcl.master.sequence'].search(
                [('nhcl_code', '=', 'Auto Serial Number'), ('nhcl_state', '=', 'activate')])
            if auto_generate_sequence:
                if auto_generate_sequence.nhcl_next_number == 1:
                    last_serial = auto_generate_sequence.nhcl_prefix + '0'
                else:
                    last_serial = auto_generate_sequence.nhcl_prefix + str(auto_generate_sequence.nhcl_next_number)
                if last_serial:
                    return self.env['stock.lot'].generate_lot_names(last_serial, 2)[1]['lot_name']
            else:
                raise ValidationError(
                    "Serial sequence is not configured in the Sequence Master. Please configure it.")
        else:
            return super(StockLot, self)._get_next_serial(company, product)


class StockBackorderConfirmation(models.TransientModel):
    """Inherited stock.backorder.confirmation class to override existing functions"""
    _inherit = 'stock.backorder.confirmation'

    def process(self):
        res = super(StockBackorderConfirmation, self).process()
        for pick_id in self.pick_ids:
            if pick_id.picking_type_id.code == 'incoming' and pick_id.stock_picking_type == 'receipt' and pick_id.is_confirm == False:
                backorder = self.env['stock.picking'].search([('backorder_id', '=', pick_id.id)])
                un_matched_ids = pick_id.stock_verification_ids.filtered(lambda x: x.stock_status == 'un_matched')
                if backorder:
                    backorder.move_ids.move_line_ids.filtered(lambda x: x.quantity == 0).unlink()
                    for un_matched_id in un_matched_ids:
                        move_lines = backorder.move_ids.move_line_ids.filtered(lambda x: x.product_id == un_matched_id.stock_product_id and x.type_product != un_matched_id.type_product)
                        for move_line in move_lines:
                            move_line.type_product = un_matched_id.type_product
                        vals = {
                            'stock_product_id': un_matched_id.stock_product_id.id,
                            'stock_serial': un_matched_id.stock_serial,
                            'stock_qty': un_matched_id.stock_qty - un_matched_id.stock_actual_qty,
                            # 'stock_actual_qty': un_matched_id.stock_actual_qty,
                            'stock_status': un_matched_id.stock_status,
                            'type_product': un_matched_id.type_product,
                            'stock_product_barcode': un_matched_id.stock_product_barcode,
                            'stock_picking_id': backorder.id,
                            'mr_price': un_matched_id.mr_price,
                            'rs_price': un_matched_id.rs_price if un_matched_id.rs_price else 0,
                            'cost_price': un_matched_id.cost_price if un_matched_id.cost_price else 0,
                            'segment': un_matched_id.segment,
                            'categ_1': un_matched_id.categ_1.id if un_matched_id.categ_1 else False,
                            'categ_2': un_matched_id.categ_2.id if un_matched_id.categ_2 else False,
                            'categ_3': un_matched_id.categ_3.id if un_matched_id.categ_3 else False,
                            'categ_4': un_matched_id.categ_4.id if un_matched_id.categ_4 else False,
                            'categ_5': un_matched_id.categ_5.id if un_matched_id.categ_5 else False,
                            'categ_6': un_matched_id.categ_6.id if un_matched_id.categ_6 else False,
                            'categ_7': un_matched_id.categ_7.id if un_matched_id.categ_7 else False,
                            'categ_8': un_matched_id.categ_8.id if un_matched_id.categ_8 else False,
                            'descrip_1': un_matched_id.descrip_1.id if un_matched_id.descrip_1 else False,
                            'descrip_2': un_matched_id.descrip_2.id if un_matched_id.descrip_2 else False,
                            'descrip_3': un_matched_id.descrip_3.id if un_matched_id.descrip_3 else False,
                            'descrip_4': un_matched_id.descrip_4.id if un_matched_id.descrip_4 else False,
                            'descrip_5': un_matched_id.descrip_5.id if un_matched_id.descrip_5 else False,
                            'descrip_6': un_matched_id.descrip_6.id if un_matched_id.descrip_6 else False,
                            'nhcl_categ_1': un_matched_id.nhcl_categ_1 if un_matched_id.nhcl_categ_1 else False,
                            'nhcl_categ_2': un_matched_id.nhcl_categ_2 if un_matched_id.nhcl_categ_2 else False,
                            'nhcl_categ_3': un_matched_id.nhcl_categ_3 if un_matched_id.nhcl_categ_3 else False,
                            'nhcl_categ_4': un_matched_id.nhcl_categ_4 if un_matched_id.nhcl_categ_4 else False,
                            'nhcl_categ_5': un_matched_id.nhcl_categ_5 if un_matched_id.nhcl_categ_5 else False,
                            'nhcl_categ_6': un_matched_id.nhcl_categ_6 if un_matched_id.nhcl_categ_6 else False,
                            'nhcl_categ_7': un_matched_id.nhcl_categ_7 if un_matched_id.nhcl_categ_7 else False,
                            'nhcl_categ_8': un_matched_id.nhcl_categ_8 if un_matched_id.nhcl_categ_8 else False,
                            'nhcl_descrip_1': un_matched_id.nhcl_descrip_1 if un_matched_id.nhcl_descrip_1 else False,
                            'nhcl_descrip_2': un_matched_id.nhcl_descrip_2 if un_matched_id.nhcl_descrip_2 else False,
                            'nhcl_descrip_3': un_matched_id.nhcl_descrip_3 if un_matched_id.nhcl_descrip_3 else False,
                            'nhcl_descrip_4': un_matched_id.nhcl_descrip_4 if un_matched_id.nhcl_descrip_4 else False,
                            'nhcl_descrip_5': un_matched_id.nhcl_descrip_5 if un_matched_id.nhcl_descrip_5 else False,
                            'nhcl_descrip_6': un_matched_id.nhcl_descrip_6 if un_matched_id.nhcl_descrip_6 else False,
                            'nhcl_descrip_7': un_matched_id.nhcl_descrip_7 if un_matched_id.nhcl_descrip_7 else False,
                            'nhcl_descrip_8': un_matched_id.nhcl_descrip_8 if un_matched_id.nhcl_descrip_8 else False,
                        }
                        self.env['stock.verification'].create(vals)
        return res


class StockVerification(models.Model):
    _name = 'stock.verification'

    stock_picking_id = fields.Many2one('stock.picking', copy=False)
    stock_product_id = fields.Many2one('product.product', string='Product', copy=False)
    stock_serial = fields.Char(string="Serial's", copy=False)
    stock_product_barcode = fields.Char(string="Barcode", copy=False)
    stock_qty = fields.Float(string='Qty', copy=False)
    stock_actual_qty = fields.Float(string='Act Qty', copy=False)
    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'Un Brand'),
                                     ('others', 'Others')], string='Type Product')
    stock_status = fields.Selection([('matched', 'Matched'), ('un_matched', 'Un Matched')], string='Status',
                                    default='un_matched')
    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, tracking=True, store=True)

    nhcl_categ_1 = fields.Char(string='Color', copy=False,
                               )
    nhcl_categ_2 = fields.Char(string='Fit', copy=False,
                               )
    nhcl_categ_3 = fields.Char(string='Brand', copy=False,
                               )
    nhcl_categ_4 = fields.Char(string='Pattern', copy=False,
                               )
    nhcl_categ_5 = fields.Char(string='Border Type', copy=False,
                               )
    nhcl_categ_6 = fields.Char(string='Border Size', copy=False,
                               )
    nhcl_categ_7 = fields.Char(string='Size', copy=False,
                               )
    nhcl_categ_8 = fields.Char(string='Design', copy=False,
                               )
    nhcl_descrip_1 = fields.Char(string="Product Aging", copy=False)
    nhcl_descrip_2 = fields.Char(string='Range', copy=False,
                                 )
    nhcl_descrip_3 = fields.Char(string='Collection', copy=False,
                                 )
    nhcl_descrip_4 = fields.Char(string='Fabric', copy=False,
                                 )
    nhcl_descrip_5 = fields.Char(string='Exclusive', copy=False,
                                 )
    nhcl_descrip_6 = fields.Char(string='Print', copy=False,
                                 )
    nhcl_descrip_7 = fields.Char(string='Days Ageing', copy=False)
    nhcl_descrip_8 = fields.Char(string='Description 8', copy=False)

    # def write(self, vals):
    #     res = super(StockVerification,self).write(vals)
    #     if res.product_id.tracking == 'serial' and res.stock_qty > 1:
    #         raise ValidationError("You cannot add more than 1.")

    @api.model
    def create(self, vals):
        product = self.env['product.product'].browse(vals.get('stock_product_id'))
        serial_or_barcode = vals.get('stock_serial')
        qty = vals.get('stock_qty', 0)

        # Only raise error if product is serial-tracked and input is a serial (not 13-digit barcode)
        if product and product.tracking == 'serial' and qty > 1:
            if not re.match(r'^\d{13}$', serial_or_barcode):  # not a 13-digit barcode
                raise ValidationError(
                    "You cannot add more than 1 quantity for a serial-tracked product (by serial number).")

        return super().create(vals)

    def write(self, vals):
        res = super().write(vals)
        for record in self:
            product = record.stock_product_id
            serial_or_barcode = record.stock_serial
            qty = record.stock_qty

            if product.tracking == 'serial' and qty > 1:
                if not re.match(r'^\d{13}$', serial_or_barcode):
                    raise ValidationError(
                        "You cannot set quantity > 1 for a serial-tracked product (by serial number).")
        return res

class StockPickingBatch(models.Model):
    _inherit = 'stock.picking.batch'

    nhcl_replication_status = fields.Boolean(string='Replication Status')
    stock_picking_type = fields.Selection(
        [('exchange', 'Customer-Return'), ('receipt', 'Receipt'), ('regular', 'Regular'), ('damage', 'Damage'),
         ('return', 'Return'),
         ('damage_main', 'Damage-Main'), ('main_damage', 'Main-Damage'),
         ('return_main', 'Return-Main'), ('delivery', 'Delivery'), ('pos_order', 'POS Order'),
         ('manufacturing', 'Manufacturing')], string='Type',
        tracking=True, related='picking_type_id.stock_picking_type')

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', '/') == '/':
                company_id = vals.get('company_id', self.env.company.id)
                picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'regular')])
                damage_picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'damage')])
                return_picking_type = self.env['stock.picking.type'].search([('stock_picking_type', '=', 'return')])
                if vals.get('is_wave'):
                    vals['name'] = self.env['ir.sequence'].with_company(company_id).next_by_code('picking.wave') or '/'
                elif vals.get('picking_type_id') == picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                elif vals.get('picking_type_id') == damage_picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = damage_picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                elif vals.get('picking_type_id') == return_picking_type.id:
                    sequence = self.env['ir.sequence'].with_company(company_id).next_by_code(
                        'picking.batch.return') or '/'
                    company = self.env.company.name
                    operation = return_picking_type.stock_picking_type
                    vals['name'] = f'Batch/{company}/{operation}/{sequence}'

                else:
                    vals['name'] = self.env['ir.sequence'].with_company(company_id).next_by_code('picking.batch') or '/'
        res = super().create(vals_list)
        print("batch",res)
        if res.stock_picking_type == 'receipt':
            for pick in  res.picking_ids:
                pick.nhcl_batch_number = res.name
        return res



class StockVerificationImport(models.TransientModel):
    _name = 'stock.verification.import'
    _description = 'Import Stock Verification'

    stock_picking_id = fields.Many2one('stock.picking', string="Stock Picking", required=True)
    file_type = fields.Selection([
        ('excel', 'Excel'),
        ('csv', 'CSV')
    ], string="File Type", required=True, default='excel')
    file_data = fields.Binary(string="Upload File", required=True)
    file_name = fields.Char(string="File Name", required=True)

    def is_valid_file_extension(self, file_name):
        valid_extensions = ['.xls', '.xlsx', '.ods', '.csv', '.txt']
        return any(file_name.lower().endswith(ext) for ext in valid_extensions)

    def clean_string(self, value):
        if value is None:
            return ''
        if isinstance(value, float):
            return str(int(value))
        return str(value).strip().replace('.0', '')

    def action_import(self):
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        try:
            data = base64.b64decode(self.file_data)
            wb = load_workbook(filename=BytesIO(data), read_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % e)

        gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'  # GS1 pattern for unbranded
        ean13_pattern = r'^\d{13}$'  # EAN13 for branded

        # Accumulators
        unbrand_serials_seen = {}
        unbrand_lot_qty = {}
        branded_serial_qty = {}  # barcode → total excel qty
        branded_lot_qty = {}

        # ===== FIRST PASS: Read Excel & validate =====
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw_code = str(row[0]).strip() if row[0] else ''
            qty = float(row[1]) if row[1] else 0

            if not raw_code:
                raise UserError(_("Missing Serial/Barcode in row %s.") % row_index)
            if qty <= 0:
                raise UserError(_("Quantity must be greater than 0 in row %s.") % row_index)

            # Try GS1 extraction for unbranded
            match = re.match(gs1_pattern, raw_code)
            unbrand_serial = match.group(2) if match else raw_code

            # Check if unbranded line exists
            unbrand_line = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', unbrand_serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ], limit=1)

            if unbrand_line:
                tracking_type = unbrand_line.stock_product_id.tracking
                if tracking_type == 'serial':
                    # Count total qty per serial in Excel
                    unbrand_serials_seen[unbrand_serial] = unbrand_serials_seen.get(unbrand_serial, 0) + qty
                    if unbrand_serials_seen[unbrand_serial] > 1:
                        raise UserError(
                            _("Duplicate or excess quantity for unbranded serial '%s'. Total in Excel = %s. Only 1 allowed.") %
                            (unbrand_serial, unbrand_serials_seen[unbrand_serial])
                        )
                else:  # lot
                    unbrand_lot_qty[unbrand_serial] = unbrand_lot_qty.get(unbrand_serial, 0) + qty
            else:
                # Branded product
                if not re.match(ean13_pattern, raw_code):
                    raise UserError(_("Invalid barcode for branded product in row %s.") % row_index)

                brand_line = self.env['stock.verification'].search([
                    ('stock_product_barcode', '=', raw_code),
                    ('stock_picking_id', '=', self.stock_picking_id.id),
                    ('type_product', '=', 'brand')
                ])

                if not brand_line:
                    raise UserError(
                        _("No verification line found for branded barcode '%s' in row %s.") % (raw_code, row_index))

                tracking_type = brand_line[0].stock_product_id.tracking
                if tracking_type == 'serial':
                    branded_serial_qty[raw_code] = branded_serial_qty.get(raw_code, 0) + qty
                else:  # lot
                    branded_lot_qty[raw_code] = branded_lot_qty.get(raw_code, 0) + qty

        # ===== SECOND PASS: Update Odoo lines =====

        # Unbranded Serial → qty always 1
        for serial in unbrand_serials_seen:
            line = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ], limit=1)
            if line:
                line.stock_actual_qty = 1
                line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'

        # Unbranded Lot → set summed qty (with validation)
        for serial, total_qty in unbrand_lot_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_serial', 'ilike', serial),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'un_brand')
            ])
            if lines:
                total_stock_qty = sum(lines.mapped('stock_qty'))
                if total_qty > total_stock_qty:
                    raise UserError(_(
                        "Excel quantity (%s) for unbranded lot '%s' exceeds available stock quantity (%s)."
                    ) % (total_qty, serial, total_stock_qty))

                for line in lines:
                    line.stock_actual_qty = total_qty
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'

        # Branded Serial → validate & fill first N lines
        for barcode, excel_qty in branded_serial_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ])
            if not lines:
                continue

            total_stock_qty = sum(lines.mapped('stock_qty'))
            if excel_qty > total_stock_qty:
                raise UserError(_(
                    "Excel quantity (%s) for branded serial '%s' exceeds available stock quantity (%s)."
                ) % (excel_qty, barcode, total_stock_qty))

            count_assigned = 0
            for line in lines:
                if count_assigned < excel_qty:
                    line.stock_actual_qty = 1
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'
                    count_assigned += 1
                else:
                    pass  # untouched lines

        # Branded Lot → set summed qty (with validation)
        for barcode, total_qty in branded_lot_qty.items():
            lines = self.env['stock.verification'].search([
                ('stock_product_barcode', '=', barcode),
                ('stock_picking_id', '=', self.stock_picking_id.id),
                ('type_product', '=', 'brand')
            ])
            if lines:
                total_stock_qty = sum(lines.mapped('stock_qty'))
                if total_qty > total_stock_qty:
                    raise UserError(_(
                        "Excel quantity (%s) for branded lot '%s' exceeds available stock quantity (%s)."
                    ) % (total_qty, barcode, total_stock_qty))

                for line in lines:
                    line.stock_actual_qty = total_qty
                    line.stock_status = 'matched' if line.stock_actual_qty == line.stock_qty else 'un_matched'

        self.file_data = False

        return {
            'effect': {
                'fadeout': 'slow',
                'message': _("Excel processed successfully for Brand and UnBrand products."),
                'type': 'rainbow_man',
            }
        }

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     _logger.info("File successfully read and decoded.")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_count = defaultdict(int)
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             # First pass: count barcodes and collect unique serials/lots
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial Number or Barcode must be provided.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 if serial_no:
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial/lot number in file: {serial_no}")
    #
    #                     move_line = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: Serial/Lot number '{serial_no}' not found in move lines.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial':
    #                         if qty != 1:
    #                             raise UserError(
    #                                 f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #                     elif tracking == 'lot':
    #                         pass  # Allow quantity > 1 for lot-tracked
    #                     else:
    #                         raise UserError(
    #                             f"Row {idx}: Product tracking not defined for product '{move_line.product_id.display_name}'.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #                 else:
    #                     # Barcode line, to be processed in second pass
    #                     barcode_count[barcode] += 1
    #
    #             # Second pass: process grouped barcodes
    #             for barcode, excel_qty in barcode_count.items():
    #                 matching_lines = move_lines.filtered(
    #                     lambda ml: ml.internal_ref_lot == barcode and ml.lot_name not in found_serials)
    #                 available_serials = [ml.lot_name for ml in matching_lines]
    #
    #                 if len(available_serials) < excel_qty:
    #                     raise UserError(
    #                         f"Barcode '{barcode}' mismatch: Excel shows {excel_qty}, "
    #                         f"but only {len(available_serials)} serials found in stock move lines."
    #                     )
    #
    #                 used_lines = matching_lines[:excel_qty]
    #
    #                 for ml in used_lines:
    #                     found_serials.add(ml.lot_name)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': ml.lot_name,
    #                         'stock_qty': 1,
    #                     }))
    #
    #         # Clear existing lines and write new ones
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         if operation_qty == verification_qty:
    #             self.stock_picking_id.verification_success = 'matched'
    #         else:
    #             self.stock_picking_id.verification_success = ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     _logger.info("File successfully read and decoded.")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_count = defaultdict(int)
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             # First pass: count barcodes and collect unique serials/lots
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial Number or Barcode must be provided.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 if serial_no:
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial/lot number in file: {serial_no}")
    #
    #                     move_line = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: Serial/Lot number '{serial_no}' not found in move lines.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial':
    #                         if qty != 1:
    #                             raise UserError(
    #                                 f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #                     elif tracking == 'lot':
    #                         pass  # Allow quantity > 1 for lot-tracked
    #                     else:
    #                         raise UserError(
    #                             f"Row {idx}: Product tracking not defined for product '{move_line.product_id.display_name}'.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #                 else:
    #                     # Barcode line, to be processed in second pass
    #                     barcode_count[barcode] += qty
    #
    #             # Second pass: process grouped barcodes
    #             for barcode, excel_qty in barcode_count.items():
    #                 matching_lines = move_lines.filtered(
    #                     lambda ml: ml.internal_ref_lot == barcode and ml.lot_name not in found_serials)
    #
    #                 if not matching_lines:
    #                     raise UserError(
    #                         f"Barcode '{barcode}' not found in stock move lines."
    #                     )
    #
    #                 ml = matching_lines[0]  # ✅ get first match
    #                 tracking = ml.product_id.tracking  # ✅ Get product tracking
    #
    #                 if tracking == 'serial':
    #                     if len(matching_lines) < excel_qty:
    #                         raise UserError(
    #                             f"Barcode '{barcode}' mismatch: Excel shows {excel_qty}, "
    #                             f"but only {len(matching_lines)} serials found in stock move lines."
    #                         )
    #                     used_lines = matching_lines[:excel_qty]
    #
    #                     for each_ml in used_lines:
    #                         found_serials.add(each_ml.lot_name)
    #
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': barcode,
    #                         'stock_qty': excel_qty,  # ✅ total quantity from Excel
    #                     }))
    #
    #                 elif tracking == 'lot':
    #                     # ✅ Allow single lot line with qty > 1
    #                     found_serials.add(ml.lot_name)
    #
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': ml.product_id.id,
    #                         'stock_serial': barcode,
    #                         'stock_qty': excel_qty,  # ✅ total qty from Excel for lot
    #                     }))
    #                 else:
    #                     raise UserError(f"Product '{ml.product_id.display_name}' has undefined tracking.")
    #
    #         # Clear existing lines and write new ones
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]  # Reset old lines
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         if operation_qty == verification_qty:
    #             self.stock_picking_id.verification_success = 'matched'
    #         else:
    #             self.stock_picking_id.verification_success = ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #     if self.stock_picking_id.verification_success == 'matched':
    #         raise UserError("This picking is already verified and matched. Further imports are not allowed.")
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_qty_map = defaultdict(int)
    #     barcode_product_map = {}
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #             _logger.info("Processing Excel file...")
    #
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 serial_no = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not serial_no and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial or Barcode must be provided.")
    #
    #                 if serial_no and barcode:
    #                     raise UserError(f"Row {idx}: Provide either Serial or Barcode, not both.")
    #
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 # --- Handle Unbranded via Serial ---
    #                 if serial_no:
    #                     matched_lines = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not matched_lines:
    #                         raise UserError(f"Row {idx}: Serial '{serial_no}' not found in move lines.")
    #
    #                     move_line = next((ml for ml in matched_lines if ml.product_id), None)
    #                     if not move_line:
    #                         raise UserError(f"Row {idx}: No product found for Serial '{serial_no}'.")
    #
    #                     if move_line.type_product == 'brand':
    #                         raise UserError(f"Row {idx}: '{serial_no}' is a branded product. Use Barcode.")
    #
    #                     tracking = move_line.product_id.tracking
    #                     if tracking == 'serial' and qty != 1:
    #                         raise UserError(f"Row {idx}: Quantity must be 1 for serial-tracked product '{serial_no}'.")
    #
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial '{serial_no}' in file.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': move_line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #
    #                 # --- Handle Branded via Barcode ---
    #                 elif barcode:
    #                     if not re.match(r'^\d{13}$', barcode):
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' is not valid EAN-13.")
    #
    #                     matched_lines = move_lines.filtered(lambda ml: ml.internal_ref_lot == barcode and ml.lot_name)
    #                     if not matched_lines:
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' not found in move lines.")
    #
    #                     branded_lines = matched_lines.filtered(lambda ml: ml.type_product == 'brand')
    #                     if not branded_lines:
    #                         raise UserError(f"Row {idx}: Barcode '{barcode}' refers to unbranded product. Use Serial.")
    #
    #                     available_lines = [ml for ml in branded_lines if ml.lot_name not in found_serials]
    #                     if len(available_lines) < qty:
    #                         raise UserError(
    #                             f"Row {idx}: Qty = {qty}, but only {len(available_lines)} available for barcode '{barcode}'.")
    #
    #                     used_lines = available_lines[:int(qty)]
    #                     for ml in used_lines:
    #                         found_serials.add(ml.lot_name)
    #
    #                     if barcode not in barcode_product_map:
    #                         product_line = next((ml for ml in branded_lines if ml.product_id), None)
    #                         if not product_line:
    #                             raise UserError(f"Row {idx}: No product found for barcode '{barcode}'.")
    #                         barcode_product_map[barcode] = product_line.product_id.id
    #
    #                     barcode_qty_map[barcode] += qty
    #
    #         # --- Finalize barcode-based lines (grouped) ---
    #         for barcode, total_qty in barcode_qty_map.items():
    #             verification_lines.append((0, 0, {
    #                 'stock_product_id': barcode_product_map[barcode],
    #                 'stock_serial': barcode,
    #                 'stock_qty': total_qty,
    #             }))
    #
    #         # --- Save verification lines ---
    #         self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #         _logger.info(f"Total verification lines created: {len(verification_lines)}")
    #
    #         # --- Check if fully matched ---
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #         self.stock_picking_id.verification_success = 'matched' if operation_qty == verification_qty else ''
    #
    #     except Exception as e:
    #         _logger.error(f"Error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}

    # main
    # def action_import(self):
    #     _logger.info("Starting barcode verification import process.")
    #
    #
    #     if not self.file_name or isinstance(self.file_name, bool):
    #         raise UserError("File name is missing or invalid.")
    #
    #     if not self.is_valid_file_extension(self.file_name):
    #         raise UserError("Invalid file format! Allowed: .xls, .xlsx, .csv, .txt")
    #
    #     file_content = base64.b64decode(self.file_data)
    #     if len(file_content) == 0:
    #         raise UserError("File is empty!")
    #
    #     # ✅ Reset verification lines before starting fresh
    #     self.stock_picking_id.stock_verification_ids = [(5, 0, 0)]
    #
    #     gs1_pattern = r'01(\d{14})21([A-Za-z0-9]+)'  # GS1 format
    #     ean13_pattern = r'^\d{13}$'
    #
    #     verification_lines = []
    #     found_serials = set()
    #     barcode_qty_map = defaultdict(float)
    #     barcode_row_map = {}
    #     move_lines = self.stock_picking_id.move_line_ids_without_package
    #
    #     try:
    #         if self.file_type == 'excel':
    #             workbook = openpyxl.load_workbook(io.BytesIO(file_content))
    #             sheet = workbook.active
    #
    #             for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #                 raw_serial_input = self.clean_string(row[0])
    #                 barcode = self.clean_string(row[1])
    #                 qty = row[2] or 0
    #
    #                 if not raw_serial_input and not barcode:
    #                     raise UserError(f"Row {idx}: Either Serial or Barcode must be provided.")
    #                 if raw_serial_input and barcode:
    #                     raise UserError(f"Row {idx}: Provide either Serial or Barcode, not both.")
    #                 if not isinstance(qty, (int, float)) or qty <= 0:
    #                     raise UserError(f"Row {idx}: Quantity must be a positive number.")
    #
    #                 # GS1 Serial Parse
    #                 gs1_match = re.match(gs1_pattern, raw_serial_input)
    #                 serial_no = gs1_match.group(2) if gs1_match else raw_serial_input
    #
    #                 # Unbranded: Use Serial/GS1
    #                 if serial_no:
    #                     matching_lines = move_lines.filtered(lambda ml: ml.lot_name == serial_no)
    #                     if not matching_lines:
    #                         raise UserError(f"Row {idx}: Serial/Lot '{serial_no}' not found in move lines.")
    #
    #                     line = next((ml for ml in matching_lines if ml.product_id), False)
    #                     if not line:
    #                         raise UserError(f"Row {idx}: No product found for serial '{serial_no}'.")
    #
    #                     if line.type_product == 'brand':
    #                         raise UserError(f"Row {idx}: Branded product '{serial_no}' must use barcode, not serial.")
    #
    #                     tracking = line.product_id.tracking
    #                     if tracking == 'serial' and qty != 1:
    #                         raise UserError(f"Row {idx}: Serial-tracked product '{serial_no}' must have quantity 1.")
    #
    #                     if serial_no in found_serials:
    #                         raise UserError(f"Row {idx}: Duplicate serial '{serial_no}' in file.")
    #
    #                     found_serials.add(serial_no)
    #                     verification_lines.append((0, 0, {
    #                         'stock_product_id': line.product_id.id,
    #                         'stock_serial': serial_no,
    #                         'stock_qty': qty,
    #                     }))
    #
    #                 # Branded: Use Barcode
    #                 elif barcode:
    #                     if not re.match(ean13_pattern, barcode):
    #                         raise UserError(f"Row {idx}: Invalid barcode format '{barcode}'. Must be 13 digits.")
    #                     barcode_qty_map[barcode] += qty
    #                     barcode_row_map[barcode] = idx
    #
    #         # ✅ Second pass: process barcode lines
    #         for barcode, total_qty in barcode_qty_map.items():
    #             idx = barcode_row_map.get(barcode, 0)
    #             matching_lines = move_lines.filtered(lambda ml: ml.internal_ref_lot == barcode and ml.lot_name)
    #             if not matching_lines:
    #                 raise UserError(f"Row {idx}: Barcode '{barcode}' not found in move lines.")
    #
    #             branded_lines = matching_lines.filtered(lambda ml: ml.type_product == 'brand')
    #             if not branded_lines:
    #                 raise UserError(f"Row {idx}: Barcode '{barcode}' refers to unbranded product. Use serial.")
    #
    #             product_line = next((ml for ml in branded_lines if ml.product_id), False)
    #             if not product_line:
    #                 raise UserError(f"Row {idx}: Product not found for barcode '{barcode}'.")
    #
    #             tracking = product_line.product_id.tracking
    #
    #             if tracking == 'serial':
    #                 available_lines = [ml for ml in branded_lines if ml.lot_name not in found_serials]
    #                 if len(available_lines) < total_qty:
    #                     raise UserError(
    #                         f"Row {idx}: Qty = {total_qty}, but only {len(available_lines)} available for barcode '{barcode}'.")
    #
    #                 for ml in available_lines[:int(total_qty)]:
    #                     found_serials.add(ml.lot_name)
    #
    #                 verification_lines.append((0, 0, {
    #                     'stock_product_id': product_line.product_id.id,
    #                     'stock_serial': barcode,
    #                     'stock_qty': total_qty,
    #                 }))
    #
    #             elif tracking == 'lot':
    #                 available_qty = sum(ml.quantity for ml in branded_lines if ml.lot_name not in found_serials)
    #                 if available_qty < total_qty:
    #                     raise UserError(
    #                         f"Row {idx}: Qty = {total_qty}, but only {available_qty} available for barcode '{barcode}'.")
    #
    #                 for ml in branded_lines:
    #                     found_serials.add(ml.lot_name)
    #
    #                 verification_lines.append((0, 0, {
    #                     'stock_product_id': product_line.product_id.id,
    #                     'stock_serial': barcode,
    #                     'stock_qty': total_qty,
    #                 }))
    #
    #         # ✅ Push new verification lines
    #         self.stock_picking_id.stock_verification_ids = verification_lines
    #
    #         # ✅ Check final quantity match
    #         operation_qty = sum(self.stock_picking_id.move_line_ids_without_package.mapped('quantity'))
    #         verification_qty = sum(line[2]['stock_qty'] for line in verification_lines)
    #
    #     except Exception as e:
    #         _logger.error(f"Import error: {str(e)}")
    #         raise UserError(f"Error importing file: {str(e)}")
    #
    #     return {'type': 'ir.actions.act_window_close'}



class StockPickingBarcode(models.Model):
    _name = 'stock.picking.barcode'

    @api.model
    def create(self, vals):
        picking = self.env['stock.picking'].browse(vals.get('stock_picking_delivery_id'))
        serial_no = vals.get('serial_no')
        # delivery_check_seq = self.env['nhcl.master.sequence'].search(
        #     [('nhcl_code', '=', 'cmr.delivery'), ('nhcl_state', '=', 'activate')], limit=1)
        # if not delivery_check_seq:
        #     raise ValidationError(_('The Delivery Check Sequence is not specified in the sequence master. "Please configure it!.'))

        if not picking.lr_number:
            raise ValidationError(_('LR Number is missing in the Stock Picking.'))
        final_barcode = f"{picking.lr_number}-{serial_no}"
        barcode = f"{picking.name}-{serial_no}"
        vals['barcode'] = barcode

        if vals.get('sequence', 'New') == 'New':
            vals['sequence'] = final_barcode

        res = super(StockPickingBarcode, self).create(vals)
        return res

    barcode = fields.Char(string='Barcode')
    stock_picking_delivery_id = fields.Many2one('stock.picking', string="Delivery Number")
    lr_number = fields.Char(
        string="LR Number",
        related='stock_picking_delivery_id.lr_number',
        store=False,
        readonly=True
    )
    sequence = fields.Char(string="Sequence",copy=False, default=lambda self: _("New"))
    serial_no = fields.Integer(string='S.NO')



class NhclStockMoveLine(models.Model):
    _name = "nhcl.stock.move.line"

    product_id = fields.Many2one('product.product', 'Product', ondelete="cascade", check_company=True, domain="[('type', '!=', 'service')]", index=True)
    lot_name = fields.Char('Lot/Serial Number Name')
    quantity = fields.Float(
        'Quantity', digits='Product Unit of Measure', copy=False, store=True,
         readonly=False)
    internal_ref_lot = fields.Char(string="Barcode", copy=False, tracking=True)

    type_product = fields.Selection([('brand', 'Brand'), ('un_brand', 'UnBrand'), ('others', 'Others')],
                                    string='Brand Type', copy=False)

    categ_1 = fields.Many2one('product.attribute.value', string='Color', copy=False,
                              domain=[('attribute_id.name', '=', 'Color')])
    categ_2 = fields.Many2one('product.attribute.value', string='Fit', copy=False,
                              domain=[('attribute_id.name', '=', 'Fit')])
    categ_3 = fields.Many2one('product.attribute.value', string='Brand', copy=False,
                              domain=[('attribute_id.name', '=', 'Brand')])
    categ_4 = fields.Many2one('product.attribute.value', string='Pattern', copy=False,
                              domain=[('attribute_id.name', '=', 'Pattern')])
    categ_5 = fields.Many2one('product.attribute.value', string='Border Type', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Type')])
    categ_6 = fields.Many2one('product.attribute.value', string='Border Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Border Size')])
    categ_7 = fields.Many2one('product.attribute.value', string='Size', copy=False,
                              domain=[('attribute_id.name', '=', 'Size')])
    categ_8 = fields.Many2one('product.attribute.value', string='Design', copy=False,
                              domain=[('attribute_id.name', '=', 'Design')])
    descrip_1 = fields.Many2one('product.aging.line', string="Product Aging", copy=False)
    descrip_2 = fields.Many2one('product.attribute.value', string='Range', copy=False,
                                domain=[('attribute_id.name', '=', 'Range')])
    descrip_3 = fields.Many2one('product.attribute.value', string='Collection', copy=False,
                                domain=[('attribute_id.name', '=', 'Collection')])
    descrip_4 = fields.Many2one('product.attribute.value', string='Fabric', copy=False,
                                domain=[('attribute_id.name', '=', 'Fabric')])
    descrip_5 = fields.Many2one('product.attribute.value', string='Exclusive', copy=False,
                                domain=[('attribute_id.name', '=', 'Exclusive')])
    descrip_6 = fields.Many2one('product.attribute.value', string='Print', copy=False,
                                domain=[('attribute_id.name', '=', 'Print')])
    descrip_7 = fields.Many2one('product.attribute.value', string='Days Ageing', copy=False,
                                domain=[('attribute_id.name', '=', 'Days Ageing')])
    descrip_8 = fields.Many2one('product.attribute.value', string='Description 8', copy=False)
    cost_price = fields.Float(string='CP', copy=False)
    mr_price = fields.Float(string='MRP', copy=False)
    rs_price = fields.Float(string='RSP', copy=False)
    segment = fields.Selection([('apparel', 'Apparel'), ('non_apparel', 'Non Apparel'), ('others', 'Others')],
                               string="Segment", copy=False, tracking=True, store=True)
    # move_line_picking_type = fields.Selection(related='picking_id.stock_picking_type', string='Picking Type')
