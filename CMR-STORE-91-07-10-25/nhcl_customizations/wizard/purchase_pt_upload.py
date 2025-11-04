import base64
from io import BytesIO
from openpyxl import load_workbook
from odoo import models, fields, _
from odoo.exceptions import UserError, ValidationError
import logging

_logger = logging.getLogger(__name__)

class PurchasePTUploadWizard(models.TransientModel):
    _name = "purchase.pt.upload.wizard"
    _description = "Purchase PT Upload Wizard"

    order_id = fields.Many2one(
        "purchase.order", string="Purchase Order", required=True
    )
    file_data = fields.Binary("Upload Excel", required=True)
    # upload_file = fields.Binary("Upload Excel File", required=True)
    file_name = fields.Char("File Name")

    # def action_import_file(self):
    #     if not self.file_data:
    #         raise UserError(_("Please upload an Excel file."))
    #
    #     data = base64.b64decode(self.file_data)
    #     try:
    #         wb = load_workbook(filename=BytesIO(data), read_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Error reading Excel file: %s") % e)
    #
    #     errors = []
    #     created_lines = 0
    #
    #     # First clear existing order lines
    #     self.order_id.order_line.unlink()
    #
    #     for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         try:
    #             article_code = ''
    #             article_name = ''
    #
    #             if row[1]:
    #                 article_name = str(row[1]).strip()
    #
    #             # Extract code if format is like [CODE] NAME
    #             if article_name.startswith('[') and ']' in article_name:
    #                 article_code = article_name.split(']')[0].strip('[').strip()
    #
    #             if not article_code:
    #                 errors.append("Row %s: No article code found in '%s'" % (row_index, article_name))
    #                 continue
    #
    #             # Search product strictly by default_code
    #             product = self.env['product.product'].search([
    #                 ('default_code', '=', article_code)
    #             ], limit=1)
    #
    #             if not product:
    #                 _logger.warning("Row %s: Product with code '%s' not found", row_index, article_code)
    #
    #                 errors.append("Row %s: Product with code '%s' not found" % (row_index, article_code))
    #                 continue
    #
    #             # Prepare values for purchase.order.line
    #             line_vals = {
    #                 'order_id': self.order_id.id,
    #                 'product_id': product.id,
    #                 'name': article_name or product.display_name,  # ARTICLE_NAME
    #                 'icode_barcode': row[0] or '',
    #                 'brand': row[2] or '',
    #                 'size': row[3] or '',
    #                 'design': row[4] or '',
    #                 'fit': row[5] or '',
    #                 'colour': row[6] or '',
    #                 'product_excel': row[7] or '',
    #                 'standard_rate': float(row[8]) if row[8] else 0.0,
    #                 'mrp': float(row[9]) if row[9] else 0.0,
    #                 'rsp': float(row[10]) if row[10] else 0.0,
    #                 'item_vendor_id': row[11] or '',
    #                 'hsn_sac_code': row[12] or '',
    #                 'product_qty': float(row[13]) if row[13] else 0.0,
    #                 'des5': row[15] or '',
    #                 'des6': row[16] or '',
    #             }
    #
    #             self.env['purchase.order.line'].create(line_vals)
    #             created_lines += 1
    #
    #         except Exception as e:
    #             _logger.error("Row %s failed: %s", row_index, e)
    #             errors.append("Row %s failed: %s" % (row_index, e))
    #             continue
    #
    #     # Reset file_data
    #     self.sudo().write({'file_data': False})
    #
    #
    #     message = _("Successfully created %s purchase order lines.") % created_lines
    #     if errors:
    #         message += "\n\nErrors:\n" + "\n".join(errors)
    #
    #     return {
    #         'effect': {
    #             'fadeout': 'slow',
    #             'message': message,
    #             'type': 'rainbow_man',
    #         }
    #     }

    # def action_import_file(self):
    #     if not self.file_data:
    #         raise UserError(_("Please upload an Excel file."))
    #
    #     data = base64.b64decode(self.file_data)
    #     try:
    #         wb = load_workbook(filename=BytesIO(data), read_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Error reading Excel file: %s") % e)
    #
    #     created_lines = 0
    #
    #     # First clear existing order lines
    #     self.order_id.order_line.unlink()
    #
    #     for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         article_code = ''
    #         article_name = ''
    #
    #         if row[1]:
    #             article_name = str(row[1]).strip()
    #
    #         # Extract code if format is like [CODE] NAME
    #         if article_name.startswith('[') and ']' in article_name:
    #             article_code = article_name.split(']')[0].strip('[').strip()
    #
    #         if not article_code:
    #             raise ValidationError(_("Row %s: No article code found in '%s'") % (row_index, article_name))
    #
    #         # Search product strictly by default_code
    #         product = self.env['product.product'].search([('default_code', '=', article_code)], limit=1)
    #
    #         if not product:
    #             raise ValidationError(
    #                 _("Row %s: Product with code '%s' not found in Product Master") % (row_index, article_code)
    #             )
    #
    #         # Prepare values for purchase.order.line
    #         line_vals = {
    #             'order_id': self.order_id.id,
    #             'product_id': product.id,
    #             'name': article_name or product.display_name,  # ARTICLE_NAME
    #             'icode_barcode': row[0] or '',
    #             'brand': row[2] or '',
    #             'size': row[3] or '',
    #             'design': row[4] or '',
    #             'fit': row[5] or '',
    #             'colour': row[6] or '',
    #             'product_excel': row[7] or '',
    #             'standard_rate': float(row[8]) if row[8] else 0.0,
    #             'mrp': float(row[9]) if row[9] else 0.0,
    #             'rsp': float(row[10]) if row[10] else 0.0,
    #             'item_vendor_id': row[11] or '',
    #             'hsn_sac_code': row[12] or '',
    #             'product_qty': float(row[13]) if row[13] else 0.0,
    #             'des5': row[15] or '',
    #             'des6': row[16] or '',
    #         }
    #
    #         self.env['purchase.order.line'].create(line_vals)
    #         created_lines += 1
    #
    #     # Reset file_data
    #     self.sudo().write({'file_data': False})
    #
    #     return {
    #         'effect': {
    #             'fadeout': 'slow',
    #             'message': _("Successfully created %s purchase order lines.") % created_lines,
    #             'type': 'rainbow_man',
    #         }
    #     }

    # def action_import_file(self):
    #     if not self.file_data:
    #         raise UserError(_("Please upload an Excel file."))
    #
    #     data = base64.b64decode(self.file_data)
    #     try:
    #         wb = load_workbook(filename=BytesIO(data), read_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Error reading Excel file: %s") % e)
    #
    #     created_lines = 0
    #
    #     # First clear existing order lines
    #     self.order_id.order_line.unlink()
    #
    #     # Define attribute columns in Excel: column index starts from 0
    #     attribute_columns = {
    #         'brand': 2,
    #         'size': 3,
    #         'design': 4,
    #         'fit': 5,
    #         'colour': 6,
    #     }
    #
    #     for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         for attr_name, col_index in attribute_columns.items():
    #             attr_value = row[col_index]
    #             if not attr_value:
    #                 raise ValidationError(
    #                     _("Row %s: Column '%s' must have a value") % (row_index, attr_name)
    #                 )
    #
    #             # Find the attribute by name (case-insensitive)
    #             attribute = self.env['product.attribute'].search([('name', 'ilike', attr_name)], limit=1)
    #             if not attribute:
    #                 raise ValidationError(
    #                     _("Row %s: Attribute '%s' does not exist in Odoo") % (row_index, attr_name)
    #                 )
    #
    #             # Check attribute value exists (case-insensitive)
    #             value_exists = self.env['product.attribute.value'].search([
    #                 ('attribute_id', '=', attribute.id),
    #                 ('name', 'ilike', str(attr_value).strip())
    #             ], limit=1)
    #             if not value_exists:
    #                 raise ValidationError(
    #                     _("Row %s: Value '%s' for attribute '%s' does not exist in Odoo") % (
    #                         row_index, attr_value, attr_name
    #                     )
    #                 )
    #
    #         # Prepare values for purchase.order.line
    #         line_vals = {
    #             'order_id': self.order_id.id,
    #             'icode_barcode': row[0] or '',
    #             'brand': row[2] or '',
    #             'size': row[3] or '',
    #             'design': row[4] or '',
    #             'fit': row[5] or '',
    #             'colour': row[6] or '',
    #             'product_excel': row[7] or '',
    #             'standard_rate': float(row[8]) if row[8] else 0.0,
    #             'mrp': float(row[9]) if row[9] else 0.0,
    #             'rsp': float(row[10]) if row[10] else 0.0,
    #             'item_vendor_id': row[11] or '',
    #             'hsn_sac_code': row[12] or '',
    #             'product_qty': float(row[13]) if row[13] else 0.0,
    #             'des5': row[15] or '',
    #             'des6': row[16] or '',
    #         }
    #
    #         self.env['purchase.order.line'].create(line_vals)
    #         created_lines += 1
    #
    #     # Reset file_data
    #     self.sudo().write({'file_data': False})
    #
    #     return {
    #         'effect': {
    #             'fadeout': 'slow',
    #             'message': _("Successfully created %s purchase order lines.") % created_lines,
    #             'type': 'rainbow_man',
    #         }
    #     }

    # def action_import_file(self):
    #     if not self.file_data:
    #         raise UserError(_("Please upload an Excel file."))
    #
    #     data = base64.b64decode(self.file_data)
    #     try:
    #         wb = load_workbook(filename=BytesIO(data), read_only=True)
    #         sheet = wb.active
    #     except Exception as e:
    #         raise UserError(_("Error reading Excel file: %s") % e)
    #
    #     created_lines = 0
    #
    #     # First clear existing order lines
    #     self.order_id.order_line.unlink()
    #
    #     # Map Excel columns (index starts at 0) to Odoo attributes
    #     attribute_columns = {
    #         'brand': 2,
    #         'size': 3,
    #         'design': 4,
    #         'fit': 5,
    #         'color': 6,
    #     }
    #
    #     for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
    #         article_code = ''
    #         article_name = ''
    #
    #         if row[1]:
    #             article_name = str(row[1]).strip()
    #
    #         # Extract code if format is like [CODE] NAME
    #         if article_name.startswith('[') and ']' in article_name:
    #             article_code = article_name.split(']')[0].strip('[').strip()
    #
    #         if not article_code:
    #             raise ValidationError(_("Row %s: No article code found in '%s'") % (row_index, article_name))
    #
    #         # Search product strictly by default_code
    #         product = self.env['product.product'].search([('default_code', '=', article_code)], limit=1)
    #
    #         if not product:
    #             raise ValidationError(
    #                 _("Row %s: Product with code '%s' not found in Product Master") % (row_index, article_code)
    #             )
    #
    #         # --- ATTRIBUTE VALUE VALIDATION (case-insensitive) ---
    #         for attr_name, col_index in attribute_columns.items():
    #             attr_value = row[col_index]
    #
    #             # Convert value to string and strip spaces only if not None
    #             if attr_value is not None:
    #                 attr_value_str = str(attr_value).strip()
    #
    #                 # Find attribute in Odoo (case-insensitive)
    #                 attribute = self.env['product.attribute'].search([('name', '=ilike', attr_name)], limit=1)
    #                 if not attribute:
    #                     raise ValidationError(
    #                         _("Row %s: Attribute '%s' does not exist in Odoo") % (row_index, attr_name))
    #
    #                 # Check attribute value exists (case-sensitive)
    #                 value_exists = self.env['product.attribute.value'].search([
    #                     ('attribute_id', '=', attribute.id),
    #                     ('name', '=', attr_value_str)
    #                 ], limit=1)
    #
    #                 if not value_exists:
    #                     raise ValidationError(
    #                         _("Row %s: Value '%s' for attribute '%s' does not exist in Odoo") %
    #                         (row_index, attr_value_str, attr_name)
    #                     )
    #         # --- END ATTRIBUTE VALUE VALIDATION ---
    #
    #         # Ensure 'name' field is never empty
    #         line_name = article_name or product.display_name or 'Imported Line'
    #
    #         # Prepare values for purchase.order.line
    #         line_vals = {
    #             'order_id': self.order_id.id,
    #             'product_id': product.id,
    #             'name': line_name,  # mandatory field
    #             'icode_barcode': row[0] or '',
    #             'brand': str(row[2]).strip() if row[2] else '',
    #             'size': str(row[3]).strip() if row[3] else '',
    #             'design': str(row[4]).strip() if row[4] else '',
    #             'fit': str(row[5]).strip() if row[5] else '',
    #             'color': str(row[6]).strip() if row[6] else '',
    #             'product_excel': row[7] or '',
    #             'standard_rate': float(row[8]) if row[8] else 0.0,
    #             'mrp': float(row[9]) if row[9] else 0.0,
    #             'rsp': float(row[10]) if row[10] else 0.0,
    #             'item_vendor_id': row[11] or '',
    #             'hsn_sac_code': row[12] or '',
    #             'product_qty': float(row[13]) if row[13] else 0.0,
    #             'des5': row[15] or '',
    #             'des6': row[16] or '',
    #         }
    #
    #         self.env['purchase.order.line'].create(line_vals)
    #         created_lines += 1
    #
    #     # Reset file_data
    #     self.sudo().write({'file_data': False})
    #
    #     return {
    #         'effect': {
    #             'fadeout': 'slow',
    #             'message': _("Successfully created %s purchase order lines.") % created_lines,
    #             'type': 'rainbow_man',
    #         }
    #     }


    def action_import_file(self):
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        data = base64.b64decode(self.file_data)
        try:
            wb = load_workbook(filename=BytesIO(data), read_only=True)
            sheet = wb.active
        except Exception as e:
            raise UserError(_("Error reading Excel file: %s") % e)

        created_lines = 0
        errors = []  # Collect all errors across all rows
        seen_barcodes = set()  # Track duplicate icode_barcode in Excel

        # First clear existing order lines
        self.order_id.order_line.unlink()

        # Map Excel columns (index starts at 0) to Odoo attributes
        attribute_columns = {
            'brand': 2,
            'size': 3,
            'design': 4,
            'fit': 5,
            'color': 6,
        }

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_errors = []  # Errors for this row
            barcode = row[0]  # icode_barcode column
            article_code = ''
            article_name = str(row[1]).strip() if row[1] else ''

            # Extract code if format is like [CODE] NAME
            if article_name.startswith('[') and ']' in article_name:
                article_code = article_name.split(']')[0].strip('[').strip()

            # --- PRODUCT VALIDATION ---
            if not article_code:
                row_errors.append(f"Row {row_index}: No article code found in '{article_name}'")
            else:
                product = self.env['product.product'].search([('default_code', '=', article_code)], limit=1)
                if not product:
                    row_errors.append(f"Row {row_index}: Product with code '{article_code}' not found in Product Master")

            # --- BARCODE DUPLICATE VALIDATION ---
            if barcode:
                barcode_str = str(barcode).strip()
                if barcode_str in seen_barcodes:
                    row_errors.append(f"Row {row_index}: Duplicate barcode '{barcode_str}' in Excel")
                else:
                    seen_barcodes.add(barcode_str)

            # --- ATTRIBUTE VALUE VALIDATION (CASE-SENSITIVE) ---
            for attr_name, col_index in attribute_columns.items():
                attr_value = row[col_index]

                # Skip empty values; empty is allowed
                if attr_value is not None:
                    attr_value_str = str(attr_value).strip()

                    # Find attribute in Odoo (case-insensitive for attribute name)
                    attribute = self.env['product.attribute'].search([('name', '=ilike', attr_name)], limit=1)
                    if not attribute:
                        row_errors.append(f"Row {row_index}: Attribute '{attr_name}' does not exist")
                        continue

                    # Case-sensitive exact match for value
                    value_exists = self.env['product.attribute.value'].search([
                        ('attribute_id', '=', attribute.id),
                        ('name', '=', attr_value_str)  # CASE-SENSITIVE
                    ])
                    if not value_exists:
                        row_errors.append(
                            f"Row {row_index}: Value '{attr_value_str}' for attribute '{attr_name}' does not exist"
                        )

            # If there were errors in this row, add to overall errors and skip creating purchase line
            if row_errors:
                errors.extend(row_errors)
                continue

            # Ensure 'name' field is never empty
            line_name = article_name or product.display_name or 'Imported Line'

            # Prepare values for purchase.order.line
            line_vals = {
                'order_id': self.order_id.id,
                'product_id': product.id,
                'name': line_name,  # mandatory field
                'icode_barcode': str(row[0]).strip() if row[0] else '',
                'brand': str(row[2]).strip() if row[2] else '',
                'size': str(row[3]).strip() if row[3] else '',
                'design': str(row[4]).strip() if row[4] else '',
                'fit': str(row[5]).strip() if row[5] else '',
                'color': str(row[6]).strip() if row[6] else '',
                'product_excel': row[7] or '',
                'standard_rate': float(row[8]) if row[8] else 0.0,
                'mrp': float(row[9]) if row[9] else 0.0,
                'rsp': float(row[10]) if row[10] else 0.0,
                'price_unit': float(row[8]) if row[8] else 0.0,  # mandatory
                'item_vendor_id': row[11] or '',
                'hsn_sac_code': row[12] or '',
                'product_qty': float(row[13]) if row[13] else 0.0,
                'des5': row[15] or '',
                'des6': row[16] or '',
            }

            self.env['purchase.order.line'].create(line_vals)
            created_lines += 1

        # Reset file_data
        self.sudo().write({'file_data': False})

        # If any errors collected, show them all at once
        if errors:
            raise ValidationError("\n".join(errors))

        return {
            'effect': {
                'fadeout': 'slow',
                'message': _("Successfully created %s purchase order lines.") % created_lines,
                'type': 'rainbow_man',
            }
        }




