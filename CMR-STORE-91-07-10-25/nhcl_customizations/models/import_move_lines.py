import openpyxl
from io import BytesIO
from odoo import api, fields, models, _
from odoo.exceptions import ValidationError
from odoo.exceptions import UserError
import logging

_logger = logging.getLogger(__name__)
import io

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class move_line_wizard(models.TransientModel):
    _name = 'move.line.wizard'
    _description = "Move Line Wizard"

    picking_file = fields.Binary(string="Select File")
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    import_prod_option = fields.Selection([('barcode', 'Barcode'), ('code', 'Code'), ('name', 'Name')],
                                          string='Import Product By ', default='name')
    product_details_option = fields.Selection(
        [('from_product', 'Take Details From The Product'), ('from_xls', 'Take Details From The XLS File'),
         ], default='from_xls')

    def import_move_line(self):
        res = False
        counter = 0
        if self.import_option == 'csv':
            keys = ['product', 'quantity']
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.picking_file)), read_only=True
                )
                ws = wb.active
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))

            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                field = list(map(str, row_no))
                values = dict(zip(keys, field))
                if values:
                    if len(row_no) == 0:
                        continue
                    else:
                        product_barcode = ''
                        lot_name = ''
                        if row_no[0] == None:
                            continue
                        else:
                            if len(row_no[0]) > 13:
                                if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                                    17] == '1':
                                    for i in range(0, len(row_no[0])):
                                        if i > 1 and i < 16:
                                            product_barcode += row_no[0][i]
                                        elif i > 17 and i < len(row_no[0]):
                                            lot_name += row_no[0][i]
                                            continue
                            else:
                                product_barcode = row_no[0]
                        if row_no[1] == None or row_no[1] <= 0:
                            raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                        # if row_no[2] == None or row_no[2] <= 0:
                        #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                        if self.product_details_option == 'from_product':
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1]
                            })
                        elif self.product_details_option == 'from_xls':
                            values.update({'product': product_barcode,
                                           'serial_no': lot_name,
                                           'quantity': row_no[1],
                                           })
                        else:
                            values.update({
                                'product': product_barcode,
                                'serial_no': lot_name,
                                'quantity': row_no[1],
                            })
                        res = self.create_move_line(values)
        else:
            try:
                wb = openpyxl.load_workbook(
                    filename=BytesIO(base64.b64decode(self.picking_file)), read_only=True
                )
                ws = wb.active
                values = {}
            except Exception:
                raise ValidationError(_("Please select any file or You have selected invalid file"))
            for row_no in ws.iter_rows(min_row=2, max_row=None, min_col=None,
                                       max_col=None, values_only=True):
                counter+=1
                product_barcode = ''
                lot_name = ''
                if row_no[0] == None:
                    continue
                else:
                    if len(row_no[0]) > 13:
                        if row_no[0][0] == '0' and row_no[0][1] == '1' and row_no[0][16] == '2' and row_no[0][
                            17] == '1':
                            for i in range(0, len(row_no[0])):
                                if i > 1 and i < 16:
                                    product_barcode += row_no[0][i]
                                elif i > 17 and i < len(row_no[0]):
                                    lot_name += row_no[0][i]
                                    continue
                    else:
                        product_barcode = row_no[0]
                if row_no[1] == None or row_no[1] <= 0:
                    raise ValidationError(_('%s Quantity must be greater than zero.') % (row_no[0]))
                # if row_no[2] == None or row_no[2] <= 0:
                #     raise ValidationError(_('%s Price Must be Greater than Zero.') % (row_no[0]))
                if self.product_details_option == 'from_product':
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name
                    })
                elif self.product_details_option == 'from_xls':
                    values.update({'product': product_barcode,
                                   'quantity': row_no[1],
                                   'serial_no': lot_name,
                                   })
                else:
                    values.update({
                        'product': product_barcode,
                        'quantity': row_no[1],
                        'serial_no': lot_name,
                    })
                res = self.create_move_line(values)
        view_id = self.env.ref('bi_import_all_orders_lines.message_wizard_popup')
        context = dict(self._context or {})
        dict_msg = str(counter) + " Records Imported Successfully."
        context['message'] = dict_msg
        return {
            'name': _('Success'),
            'type': 'ir.actions.act_window',
            'view_mode': 'form',
            'res_model': 'message.wizard',
            'views': [(view_id.id, 'form')],
            'view_id': view_id.id,
            'target': 'new',
            'context': context,
        }

    def create_move_line(self, values):
        picking_brw = self.env['stock.picking'].browse(self._context.get('active_id'))
        serial_no = self.env['stock.lot']
        product_obj_search = self.env['product.product']
        if self.product_details_option == 'from_product':
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                raise ValidationError(_('%s Product was not found in the Database.') % values.get('product'))
            if picking_brw.state == 'draft':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_product_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_product_line.dummy_lot_ids.filtered(lambda x:x.id == serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'description_picking': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                    })
            elif picking_brw.state == 'waiting':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_lot_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_lot_line.dummy_lot_ids.filtered(lambda x:x.id == serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'description_picking': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                    })
            elif picking_brw.state != 'waiting' or picking_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        elif self.product_details_option == 'from_xls':
            if self.import_prod_option == 'barcode':
                barcode = values.get('product')
                if len(barcode) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', barcode)])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database') % values.get(
                            'product'))

            if picking_brw.state == 'draft':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_lot_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_lot_line.dummy_lot_ids.filtered(lambda x:x.id == serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'name': product_id.display_name,
                        'description_picking': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                    })
            elif picking_brw.state == 'waiting':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_lot_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_lot_line.dummy_lot_ids.filtered(lambda x:x.id == serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'description_picking': product_id.display_name,
                        'name': product_id.display_name,
                        'product_uom_qty': values.get('quantity'),
                        'product_uom': product_id.uom_id.id,
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                    })
            elif picking_brw.state != 'waiting' or picking_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        else:
            if self.import_prod_option == 'barcode':
                if len(values['product']) > 13:
                    product_obj_search = self.env['product.product'].search([('barcode', '=', values['product'])])
                    serial_no = self.env['stock.lot'].search([('name', '=', values['serial_no'])])
                    if not serial_no:
                        raise ValidationError(_('The serial number for this is not found in the database.'))
                else:
                    product_barcodes = self.env['product.barcode'].search([('barcode', '=', values['product'])])
                    if len(product_barcodes) > 0:
                        product_obj_search = product_barcodes[0].product_id
            elif self.import_prod_option == 'code':
                raise ValidationError(_('Please set the import Option to Barcode.'))
            else:
                raise ValidationError(_('Please set the import Option to Barcode.'))
            if product_obj_search:
                product_id = product_obj_search
            else:
                if self.import_prod_option == 'name':
                    raise ValidationError(_('Please set the import Option to Barcode.'))
                else:
                    raise ValidationError(
                        _('%s Product was not found in the Database') % values.get(
                            'product'))

            if picking_brw.state == 'draft':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_lot_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_lot_line.dummy_lot_ids.filtered(lambda x:x.id ==serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'description_picking': product_id.display_name,
                        'name': product_id.display_name,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'product_uom_qty': values.get('quantity'),
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                    })

            elif picking_brw.state == 'waiting':
                if product_id.nhcl_product_type == 'unbranded':
                    existing_lot_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id)
                    existing_line = existing_lot_line.dummy_lot_ids.filtered(lambda x:x.id == serial_no.id)
                else:
                    existing_line = picking_brw.move_ids.filtered(
                        lambda x: x.product_id == product_id and x.prod_barcode == values.get('product'))
                if existing_line:
                    raise ValidationError(_('%s The product already exists.') % values.get('product'))
                existing_move_line = picking_brw.move_ids.filtered(
                    lambda x: x.product_id == product_id)
                if existing_move_line:
                    existing_move_line.product_uom_qty += values.get('quantity')
                    if product_id.nhcl_product_type == 'unbranded':
                        existing_move_line.dummy_lot_ids = [(4, serial_no.id)]
                else:
                    move_lines = self.env['stock.move'].create({
                        'picking_id': picking_brw.id,
                        'product_id': product_id.id,
                        'dummy_lot_ids': serial_no.ids,
                        'prod_barcode': values.get('product'),
                        'product_uom_qty': values.get('quantity'),
                        'location_id': picking_brw.location_id.id,
                        'location_dest_id': picking_brw.location_dest_id.id,
                    })

            elif picking_brw.state != 'waiting' or picking_brw.state != 'draft':
                raise UserError(_('We cannot import data in validated or confirmed order!.'))
        return True
