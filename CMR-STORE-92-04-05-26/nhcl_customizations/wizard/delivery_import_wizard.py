import re

from odoo import models, fields, _
from odoo.exceptions import ValidationError
import base64
from io import BytesIO
from openpyxl import load_workbook


class DeliveryImportWizard(models.TransientModel):
    _name = 'delivery.import.wizard'
    _description = "Delivery Import Wizard"

    picking_id = fields.Many2one('stock.picking', string="Delivery", required=True)
    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def _check_serial_availability(self, lot, location_id):
        StockMoveLine = self.env['stock.move.line']
        StockMove = self.env['stock.move']

        # IN qty
        in_qty = sum(StockMoveLine.search([
            # ('company_id.nhcl_company_bool', '=', True),
            ('lot_id', '=', lot.id),
            ('location_dest_id', '=', location_id),
        ]).mapped('quantity'))

        # OUT qty (returns)
        out_qty = sum(StockMoveLine.search([
            # ('company_id.nhcl_company_bool', '=', True),
            ('lot_id', '=', lot.id),
            ('picking_id.stock_picking_type', '=', 'goods_return'),
        ]).mapped('quantity'))

        # Already used in other sale orders
        used_qty = sum(StockMove.search([
            ('dummy_lot_ids', 'in', [lot.id]),
            ('state', 'not in', ['cancel']),
        ]).mapped('product_uom_qty'))

        available = in_qty - out_qty - used_qty
        return available

    def action_import_barcodes(self):
        """Read Excel and create/merge sale order lines based on GS1 or EAN-13 barcode format.
        - Merges quantities for the same lot in the same sale order (lot-tracked products).
        - Validates that the total requested qty for a barcode does not exceed actual available qty.
        - When merging/updating qty, always update the line's price_unit from the lot (cost_price)
          falling back to product.lst_price.
        - For serial-tracked products behavior:
            * Each serial (lot that represents a serial) is handled as a separate 1-qty line.
            * If the same serial is already used in this order or another active order it will be skipped.
        """
        self.ensure_one()
        if not self.file:
            raise ValidationError(_("Please upload an Excel file."))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise ValidationError(_("Invalid Excel file: %s") % str(e))

        # SaleOrderLine = self.env['sale.order.line']
        StockMove = self.env['stock.move']
        StockQuant = self.env['stock.quant']
        StockMoveLine = self.env['stock.move.line']
        location_id = self.env.ref('stock.stock_location_stock').id
        skipped_lines = []

        # Helper: find existing sale order line for same product + lot + serial type in this sale order
        def _find_existing_line(product, lot_id):
            return StockMove.search([
                ('picking_id', '=', self.picking_id.id),
                ('product_id', '=', product.id),
                ('dummy_lot_ids', 'in', [lot_id]),
                # ('sale_serial_type', '=', serial_type),
                ('state', 'not in', ['cancel'])
            ], limit=1)


        # Iterate rows
        for row in sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0]).strip() if row[0] else ""
            qty = row[1] or 0

            if not barcode or not qty:
                skipped_lines.append((barcode, "Missing barcode or quantity"))
                continue


            # ------------------ BRAND + 'R' BARCODE (NEW BLOCK) ------------------
            custom_serial_pattern = r'^[A-Za-z][A-Za-z0-9]{3,}$'
            if re.match(custom_serial_pattern, barcode):

                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', barcode),
                    # ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                # ONLY handle if BRAND FOUND
                if matched_quant:
                    lot_code = barcode

                    quant_domain = [
                        ('lot_id.ref', '=', lot_code),
                        # ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand')
                    ]

                    quants = StockQuant.search(quant_domain, order='id asc')
                    product = quants[:1].product_id if quants else False

                    if not product:
                        skipped_lines.append((barcode, "No matching product for R-type brand barcode"))
                        continue


                    # -------- SERIAL TRACKING --------
                    if product.tracking == 'serial':

                        created_count = 0
                        for q in quants:
                            lot = q.lot_id
                            if not lot:
                                continue

                            available_qty = self._check_serial_availability(lot, location_id)
                            if available_qty < 1:
                                continue

                            # skip already used
                            existing_lines = StockMove.search([
                                ('dummy_lot_ids', 'in', [lot.id]),
                                ('state', 'not in', ['cancel'])
                            ])
                            if existing_lines:
                                continue

                            StockMove.create({
                                'picking_id': self.picking_id.id,
                                'product_id': product.id,
                                'dummy_lot_ids': [(6, 0, [lot.id])],
                                # 'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': 1,
                                'product_uom': product.uom_id.id,
                                # 'price_unit': lot.cost_price,
                                # 'type_product': lot.type_product,
                                # 'sale_serial_type': lot.serial_type or 'regular',
                                'location_id': self.picking_id.location_id.id,
                                'location_dest_id': self.picking_id.location_dest_id.id,

                            })

                            created_count += 1
                            if created_count >= qty:
                                break

                        if created_count < qty:
                            skipped_lines.append(
                                (barcode, f"Only {created_count} serials available out of requested {qty}")
                            )

                        continue


                    # -------- LOT TRACKING --------
                    else:
                        available_quants = quants.filtered(lambda q: q.quantity > 0)

                        total_available = sum(available_quants.mapped('quantity'))
                        if qty > total_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                            continue

                        remaining_qty = qty
                        for q in available_quants:
                            lot = q.lot_id
                            actual_available_qty = q.quantity
                            if actual_available_qty <= 0:
                                continue

                            allocate_qty = min(remaining_qty, actual_available_qty)

                            existing_line = _find_existing_line(product, lot.id)

                            if existing_line:
                                existing_line.write({
                                    'product_uom_qty': existing_line.product_uom_qty + allocate_qty,
                                })
                            else:
                                StockMove.create({
                                    'picking_id': self.picking_id.id,
                                    'product_id': product.id,
                                    'dummy_lot_ids': [(6, 0, [lot.id])],
                                    'name': product.display_name,
                                    'product_uom_qty': allocate_qty,
                                    'product_uom': product.uom_id.id,
                                    'location_id': self.picking_id.location_id.id,
                                    'location_dest_id': self.picking_id.location_dest_id.id,
                                })

                            remaining_qty -= allocate_qty
                            if remaining_qty <= 0:
                                break

                        if remaining_qty > 0:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))

                        continue
            # ------------------ GS1 (identified by presence of 'R' in the barcode) ------------------
            if 'R' in barcode:
                # GS1 -> lot code starts from 'R'
                lot_code = barcode[barcode.find('R'):]
                quant_domain = [
                    ('lot_id.name', '=', lot_code),
                    # ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    # product type filter used in original code
                    ('lot_id.type_product', '=', 'un_brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                product = quants[:1].product_id if quants else False

                if not product:
                    skipped_lines.append((barcode, "No matching product for GS1 barcode"))
                    continue

                if product.tracking == 'serial':

                    if qty != 1:
                        skipped_lines.append((barcode, "Serial product: Qty must be 1"))
                        continue

                    for q in quants:
                        lot = q.lot_id
                        if not lot:
                            continue

                        # -------- LANDED COST & PRICE VALIDATION --------
                        if lot.rs_price <= 0.0:
                            skipped_lines.append((barcode, f"{lot.name} not done with landed cost"))
                            continue

                        # -------- AVAILABILITY CHECK (MOVEMENT BASED) --------
                        available_qty = self._check_serial_availability(lot, location_id)
                        if available_qty < 1:
                            skipped_lines.append((barcode, f"Serial {lot.name} not available for sale"))
                            continue

                        # -------- CREATE SALE LINE --------
                        StockMove.create({
                            'picking_id': self.picking_id.id,
                            'product_id': product.id,
                            'dummy_lot_ids': [(6, 0, [lot.id])],
                            # 'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'location_id': self.picking_id.location_id.id,
                            'location_dest_id': self.picking_id.location_dest_id.id,
                        })
                        break


                # LOT TRACKING (this is where merging is applied)
                else:
                    # Build list of lots (quants) with their actual available qty
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No stock available for GS1 lot items"))
                        continue

                    # Validate total available across all matching lots
                    total_available = sum(available_quants.mapped('quantity'))
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = q.quantity
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        # Merge or create
                        existing_line = _find_existing_line(product, lot.id)
                        price_unit = getattr(lot, 'cost_price', product.lst_price)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                                # 'price_unit': price_unit,
                            })
                        else:
                            StockMove.create({
                                'picking_id': self.picking_id.id,
                                'product_id': product.id,
                                'dummy_lot_ids': [(6, 0, [lot.id])],
                                # 'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                # 'price_unit': price_unit,
                                # 'type_product': getattr(lot, 'type_product', False),
                                # 'sale_serial_type': sale_serial_type,
                                'location_id': self.picking_id.location_id.id,
                                'location_dest_id': self.picking_id.location_dest_id.id,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))

            # ------------------ EAN-13 (length 13) ------------------
            elif len(barcode) == 13:
                lot_code = barcode
                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', lot_code),
                    # ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
                    continue

                product = matched_quant.product_id

                # Domain to gather all quants for that lot_code + product
                quant_domain = [
                    ('product_id', '=', product.id),
                    ('lot_id.ref', '=', lot_code),
                    # ('company_id.nhcl_company_bool', '=', True),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                available_quants = quants.filtered(lambda q: q.quantity > 0)

                # existing_types = StockMove.search([
                #     ('picking_id', '=', self.picking_id.id),
                #     ('state', 'not in', ['cancel']),
                #     ('product_id', '=', product.id)
                # ]).mapped('sale_serial_type')
                # sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL TRACKING (Enhanced: include regular + return serials)
                if product.tracking == 'serial':
                    # Build regular & return quants (prioritize regular then return)
                    regular_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        # ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        # '|', ('lot_id.serial_type', '=', False), ('lot_id.serial_type', '=', 'regular'),
                    ], order='id asc')

                    return_quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', lot_code),
                        # ('company_id.nhcl_company_bool', '=', True),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand'),
                        # ('lot_id.serial_type', '=', 'return'),
                    ], order='id asc')

                    combined_quants = regular_quants + return_quants
                    if not combined_quants:
                        skipped_lines.append((barcode, "No serial stock available for EAN-13"))
                        continue

                    # Validate total available serials
                    if qty > len(combined_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(combined_quants)}"))
                        continue

                    created_count = 0
                    for q in combined_quants:
                        lot = q.lot_id

                        # Create a 1-qty line for this serial
                        StockMove.create({
                            'picking_id': self.picking_id.id,
                            'product_id': product.id,
                            'dummy_lot_ids': [(6, 0, [lot.id])],
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'location_id': self.picking_id.location_id.id,
                            'location_dest_id': self.picking_id.location_dest_id.id,
                        })
                        created_count += 1
                        if created_count >= qty:
                            break

                    if created_count < qty:
                        short = qty - created_count
                        skipped_lines.append((barcode, f"Requested qty exceeds available stock. Short by {short}"))

                # LOT TRACKING -> merge/update qty and price from lot
                else:
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if not available_quants:
                        skipped_lines.append((barcode, "No lot stock available for EAN-13"))
                        continue
                    total_available = sum(available_quants.mapped('quantity'))
                    if qty > total_available:
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available stock {int(total_available)}"))
                        continue

                    remaining_qty = qty
                    for q in available_quants:
                        lot = q.lot_id
                        actual_available_qty = q.quantity
                        if actual_available_qty <= 0:
                            continue

                        allocate_qty = min(remaining_qty, actual_available_qty)
                        if allocate_qty <= 0:
                            continue

                        existing_line = _find_existing_line(product, lot.id)

                        if existing_line:
                            # Update qty and ALWAYS update price from lot (option A)
                            new_qty = (existing_line.product_uom_qty or 0.0) + allocate_qty
                            existing_line.write({
                                'product_uom_qty': new_qty,
                            })
                        else:
                            StockMove.create({
                                'picking_id': self.picking_id.id,
                                'product_id': product.id,
                                'dummy_lot_ids': [(6, 0, [lot.id])],
                                'name': product.display_name,
                                'product_uom_qty': allocate_qty,
                                'product_uom': product.uom_id.id,
                                'location_id': self.picking_id.location_id.id,
                                'location_dest_id': self.picking_id.location_dest_id.id,
                            })

                        remaining_qty -= allocate_qty
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds available stock. Short by {int(remaining_qty)}"))
            # ------------------ UNKNOWN BARCODE ------------------
            else:
                skipped_lines.append((barcode, "Unknown barcode format"))
                continue
        # After processing all rows, raise ValidationError if anything was skipped
        if skipped_lines:
            msg = "\n".join([f"{b or '<empty>'}: {reason}" for b, reason in skipped_lines])
            raise ValidationError(_("Some lines were skipped:\n%s") % msg)
