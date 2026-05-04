from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
import base64
import io
from openpyxl import load_workbook
import xlrd


class AssetMaster(models.Model):
    _name = 'cf.asset.master'
    _description = 'Asset Master'

    name = fields.Char(required=True)
    code = fields.Char()

    @api.constrains('name')
    def _check_unique_name(self):
        for rec in self:
            if rec.name:
                existing = self.search([('name', '=', rec.name),('id', '!=', rec.id)], limit=1)
                if existing:
                    raise ValidationError(_("Asset with name '%s' already exists.") % rec.name)

class CentralFashionsCount(models.Model):
    _name = 'cf.inventory.count'
    _description = 'Central Fashions Inventory Count'

    name = fields.Char(string="Reference", required=True,
                       copy=False, readonly=True, default=lambda self: _('New'))
    requester_id = fields.Many2one('res.users', string="Requester",
                                   default=lambda self: self.env.user, required=True)
    date = fields.Date(default=fields.Date.today, required=True)
    company_id = fields.Many2one(
        'res.company',
        default=lambda self: self.env.company,
        required=True
    )

    upload_file = fields.Binary(string="Upload Excel")
    upload_filename = fields.Char()

    line_ids = fields.One2many(
        'cf.inventory.count.line',
        'count_id',
        string="Lines"
    )
    mode = fields.Selection([
        ('scan', 'Scan'),
        ('upload', 'Upload')
    ], string="Mode", default='scan')
    scan_barcode = fields.Char(string="Scan Barcode")
    lot_input_qty = fields.Float(string="Lot Qty")

    # @api.onchange('scan_barcode')
    # def _onchange_scan_barcode(self):
    #     if not self.scan_barcode:
    #         return
    #
    #     barcode = self.scan_barcode.strip()
    #     product = False
    #
    #     # 🔹 STEP 1: Check MAIN barcode (product.product)
    #     product = self.env['product.product'].search([
    #         ('barcode', '=', barcode)
    #     ], limit=1)
    #
    #     # 🔹 STEP 2: If not found → MULTI BARCODE
    #     if not product:
    #         multi_barcode = self.env['product.barcode'].search([
    #             ('barcode', '=', barcode)
    #         ], limit=1)
    #
    #         if multi_barcode:
    #             product = multi_barcode.product_id
    #
    #     # ❌ STEP 3: Not found
    #     if not product:
    #         self.scan_barcode = False
    #         raise ValidationError(f"Barcode '{barcode}' not found")
    #
    #     product_tmpl = product.product_tmpl_id
    #
    #     # 🔹 STEP 4: Existing line check
    #     line = self.line_ids.filtered(
    #         lambda l: l.product_tmpl_id.id == product_tmpl.id
    #     )
    #     soh = sum(line.product_tmpl_id.product_variant_ids.mapped('qty_available'))
    #
    #     if line:
    #         #  Increment count
    #         line.global_count += 1
    #         line.soh = soh
    #     else:
    #
    #         self.line_ids += self.line_ids.new({
    #             'product_tmpl_id': product_tmpl.id,
    #             'family_id': product_tmpl.categ_id.parent_id.parent_id.parent_id.id,
    #             'cat_id': product_tmpl.categ_id.parent_id.parent_id.id,
    #             'class_id': product_tmpl.categ_id.parent_id.id,
    #             'brick_id': product_tmpl.categ_id.id,
    #             'global_count': 1,
    #             'soh': soh,
    #         })
    #
    #     # Reset field (important for next scan)
    #     self.scan_barcode = False

    @api.onchange('scan_barcode')
    def _onchange_scan_barcode(self):
        if not self.scan_barcode:
            return

        barcode = self.scan_barcode.strip()

        try:
            main_location = self.env.ref('stock.stock_location_stock')
        except:
            self.scan_barcode = False
            raise ValidationError("Configuration Error: Main stock location not found.")


        all_matching_quants = self.env['stock.quant'].search([
            '|',
            ('lot_id.ref', '=', barcode),
            ('lot_id.name', '=', barcode),
            ('location_id', '=', main_location.id),
            ('quantity', '>', 0)
        ], order='id asc')

        if not all_matching_quants:
            self.scan_barcode = False
            raise ValidationError(f"Stock for '{barcode}' not found in Main Location.")

        selected_quant = False
        for q in all_matching_quants:
            lot = q.lot_id
            scanned_this_lot_qty = 0
            for line in self.line_ids:
                if lot.name in (line.lot_name or "").split(", "):
                    if lot.product_id.tracking == 'serial':
                        scanned_this_lot_qty += 1
                    else:
                        scanned_this_lot_qty += line.global_count

            if scanned_this_lot_qty < q.quantity:
                selected_quant = q
                break

        if not selected_quant:
            self.scan_barcode = False
            raise ValidationError(f"All available lots for barcode '{barcode}' have already been scanned.")

        lot_id = selected_quant.lot_id
        product = selected_quant.product_id
        product_tmpl = product.product_tmpl_id
        tracking = product.tracking

        if lot_id.type_product == 'brand' and lot_id.ref != barcode:
            self.scan_barcode = False
            raise ValidationError("For Brand products, please scan the Barcode.")
        elif lot_id.type_product == 'un_brand' and lot_id.name != barcode:
            self.scan_barcode = False
            raise ValidationError("For Un-Brand products, please scan the Serial Number.")

        # 4. Tracking Logic
        qty_to_add = 0
        if tracking == 'serial':
            qty_to_add = 1
        else:
            if not self.lot_input_qty or self.lot_input_qty <= 0:
                self.scan_barcode = False
                raise ValidationError("Please enter 'Lot Qty' before scanning.")

            # Check if input qty exceeds what's left in this specific lot
            if (self.lot_input_qty) > (selected_quant.quantity):
                self.scan_barcode = False
                raise ValidationError(
                    f"Insufficient stock in Lot '{lot_id.name}'. Available: {selected_quant.quantity}")
            qty_to_add = self.lot_input_qty

        # 5. Update Lines (Single line per Product Template)
        self._update_inventory_lines(product_tmpl, lot_id, qty_to_add)

        self.lot_input_qty = 0.0
        self.scan_barcode = False

    def _update_inventory_lines(self, product_tmpl, lot, qty):
        soh = sum(product_tmpl.product_variant_ids.mapped('qty_available'))
        new_lot_name = lot.name
        new_ref_name = lot.ref if lot.type_product == 'brand' else lot.name

        line = self.line_ids.filtered(lambda l: l.product_tmpl_id.id == product_tmpl.id)

        if line:
            line.global_count += qty
            line.soh = soh

            lots = [l.strip() for l in (line.lot_name or "").split(",") if l.strip()]
            if new_lot_name not in lots:
                lots.append(new_lot_name)
                line.lot_name = ", ".join(lots)

            refs = [r.strip() for r in (line.ref_name or "").split(",") if r.strip()]
            if new_ref_name not in refs:
                refs.append(new_ref_name)
                line.ref_name = ", ".join(refs)
        else:
            self.line_ids += self.line_ids.new({
                'product_tmpl_id': product_tmpl.id,
                'lot_name': new_lot_name,
                'ref_name': new_ref_name,
                'family_id': product_tmpl.categ_id.parent_id.parent_id.parent_id.id,
                'cat_id': product_tmpl.categ_id.parent_id.parent_id.id,
                'class_id': product_tmpl.categ_id.parent_id.id,
                'brick_id': product_tmpl.categ_id.id,
                'global_count': qty,
                'soh': soh,
            })

    @api.constrains('date')
    def _check_unique_date(self):
        for rec in self:
            if rec.date:
                existing = self.search([('date', '=', rec.date), ('id', '!=', rec.id)], limit=1)
                if existing:
                    raise ValidationError(_(f"Global Count with date {rec.date} already exists."))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('inventory.count') or _('New')

        return super(CentralFashionsCount, self).create(vals_list)

    def action_import_excel(self):
        self.ensure_one()

        if not self.upload_file:
            raise ValidationError("Please upload an Excel file.")

        self.line_ids.unlink()

        file_data = base64.b64decode(self.upload_file)
        workbook = load_workbook(io.BytesIO(file_data))
        sheet = workbook.active

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

            asset_name = str(row[0]).strip() if row[0] else ''
            product_name = str(row[1]).strip() if row[1] else ''
            floor_value_excel = row[2]
            global_count = row[3]

            if not asset_name:
                raise ValidationError(f"Asset missing at Row {row_index}")

            if not product_name:
                raise ValidationError(f"Product missing at Row {row_index}")

            if global_count is None:
                raise ValidationError(f"Global Count missing at Row {row_index}")

            # 🔹 Asset
            asset = self.env['cf.asset.master'].search(
                [('name', '=', asset_name)], limit=1
            )
            if not asset:
                raise ValidationError(f"Asset '{asset_name}' not found (Row {row_index})")

            # 🔹 Product Template
            product_tmpl = self.env['product.template'].search(
                [('name', '=', product_name)], limit=1
            )
            if not product_tmpl:
                raise ValidationError(f"Product '{product_name}' not found (Row {row_index})")

            # 🔹 Floor conversion
            try:
                floor_key = str(int(float(floor_value_excel)))
            except Exception:
                raise ValidationError(f"Invalid Floor '{floor_value_excel}' (Row {row_index})")

            valid_keys = dict(self.env['cf.inventory.count.line']._fields['floor'].selection).keys()

            if floor_key not in valid_keys:
                raise ValidationError(f"Invalid Floor '{floor_value_excel}' (Row {row_index})")
            soh = sum(product_tmpl.product_variant_ids.mapped('qty_available'))
            diff = global_count - soh
            # 🔹 Create Line
            self.env['cf.inventory.count.line'].create({
                'count_id': self.id,
                'asset_id': asset.id,
                'product_tmpl_id': product_tmpl.id,
                'floor': floor_key,
                'family_id': product_tmpl.categ_id.parent_id.parent_id.parent_id.id,
                'cat_id': product_tmpl.categ_id.parent_id.parent_id.id,
                'class_id': product_tmpl.categ_id.parent_id.id,
                'brick_id': product_tmpl.categ_id.id,
                'global_count': global_count,
                'soh': soh,
                'diff': diff,
            })


class CentralFashionsCountLine(models.Model):
    _name = 'cf.inventory.count.line'
    _description = 'Inventory Count Line'

    count_id = fields.Many2one('cf.inventory.count', ondelete='cascade')
    asset_id = fields.Many2one('cf.asset.master')
    floor = fields.Selection([
        ('0', 'Ground Floor'),
        ('1', '1st Floor'),
        ('2', '2nd Floor'),
        ('3', '3rd Floor'),
        ('4', '4th Floor'),
        ('5', '5th Floor'),
        ('6', '6th Floor'),
        ('7', '7th Floor'),
    ], string="Floor")
    product_tmpl_id = fields.Many2one('product.template',string='Product')
    family_id = fields.Many2one('product.category',string='Family')
    cat_id = fields.Many2one('product.category',string='Category',)
    class_id = fields.Many2one('product.category',string='Class')
    brick_id = fields.Many2one('product.category',string='Brick')
    global_count = fields.Float(string='Global Count')
    soh = fields.Float(string='SOH')
    diff = fields.Float(string='Diff', compute='get_diff', store=True)
    lot_name = fields.Char(string="Serial")
    ref_name = fields.Char(string="Barcode")

    @api.depends('soh')
    def get_diff(self):
        for rec in self:
            if rec.soh:
                rec.diff = rec.global_count - rec.soh


    @api.onchange('product_tmpl_id')
    def onchange_product_tmpl_id(self):
        for rec in self:
            if self.product_tmpl_id:
                soh = sum(self.product_tmpl_id.product_variant_ids.mapped('qty_available'))
                rec.family_id = self.product_tmpl_id.categ_id.parent_id.parent_id.parent_id.id
                rec.cat_id = self.product_tmpl_id.categ_id.parent_id.parent_id.id
                rec.class_id = self.product_tmpl_id.categ_id.parent_id.id
                rec.brick_id = self.product_tmpl_id.categ_id.id
                rec.soh = soh

    @api.onchange('global_count')
    @api.onchange('global_count')
    def onchange_global_count(self):
        for rec in self:
            if rec.global_count and rec.soh:
                rec.diff = rec.global_count - rec.soh
