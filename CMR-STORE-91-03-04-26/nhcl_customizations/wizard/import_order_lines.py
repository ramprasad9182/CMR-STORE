import openpyxl
from io import BytesIO
from openpyxl import load_workbook
from odoo import fields, models, _
from odoo.exceptions import ValidationError
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
except ImportError:
    _logger.debug('Oops! Cannot `import xlrd`.')
try:
    import csv
except ImportError:
    _logger.debug('Oops! Cannot `import csv`.')
try:
    import base64
except ImportError:
    _logger.debug('Oops! Cannot `import base64`.')


class order_line_wizard(models.TransientModel):
    _name = 'order.line.wizard'
    _description = "Order Line Wizard"

    sale_order_id = fields.Many2one('sale.order', string="Sale Order", required=True)
    file = fields.Binary(string="Excel File", required=True)
    filename = fields.Char(string="Filename")

    def action_import_barcodes(self):
        """Read Excel and create sale order lines based on GS1 or EAN-13 barcode format"""
        self.ensure_one()
        if not self.file:
            raise ValidationError(_("Please upload an Excel file."))

        try:
            data = base64.b64decode(self.file)
            wb = load_workbook(filename=BytesIO(data), data_only=True)
            sheet = wb.active
        except Exception as e:
            raise ValidationError(_("Invalid Excel file: %s") % str(e))

        skipped_lines = []
        location_id = self.env.ref('stock.stock_location_stock').id
        damage_location = self.env['stock.location'].search([('name', 'like', '%-DM')]).id
        SaleOrderLine = self.env['sale.order.line']
        StockQuant = self.env['stock.quant']

        for row in sheet.iter_rows(min_row=2, values_only=True):
            barcode = str(row[0]).strip() if row[0] else ""
            qty = row[1] or 0

            if not barcode or not qty:
                skipped_lines.append((barcode, "Missing barcode or quantity"))
                continue
            # =========================================================
            #  R START → BRAND (REGULAR) → SAME AS EAN
            # =========================================================
            if barcode.startswith('R') and self.sale_order_id.transfer_type == 'regular':

                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', barcode),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    # fallback to GS1 if unbrand
                    unbrand_exists = StockQuant.search([
                        ('lot_id.name', '=', barcode),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'un_brand')
                    ], limit=1)

                    if not unbrand_exists:
                        skipped_lines.append((barcode, "Invalid R barcode"))
                        continue

                else:
                    product = matched_quant.product_id

                    quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', barcode),
                        ('quantity', '>', 0),
                        ('location_id', '=', location_id),
                        ('lot_id.type_product', '=', 'brand')
                    ], order='id asc')

                    sale_serial_type = (SaleOrderLine.search([
                        ('order_id', '=', self.sale_order_id.id),
                        ('product_id', '=', product.id),
                        ('order_id.transfer_type', '=', 'regular')
                    ]).mapped('sale_serial_type') or ['regular'])[0]

                    # -------- SERIAL --------
                    if product.tracking == 'serial':
                        remaining_qty = qty

                        for q in quants:
                            lot = q.lot_id

                            inward_exists = self.env['stock.move.line'].search_count([
                                ('picking_id.stock_picking_type', 'in', ['receipt', 'exchange']),
                                ('lot_id', '=', lot.id),
                            ]) > 0

                            if not inward_exists:
                                continue

                            already_sold = SaleOrderLine.search_count([
                                ('lot_ids', 'in', [lot.id]),
                                ('order_id.transfer_type', '=', 'regular')
                            ]) > 0

                            if already_sold:
                                continue

                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': 1,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                            remaining_qty -= 1
                            if remaining_qty <= 0:
                                break

                        if remaining_qty > 0:
                            skipped_lines.append((barcode, f"Short by {remaining_qty}"))

                    # -------- LOT --------
                    else:
                        remaining_qty = qty

                        for q in quants:
                            lot = q.lot_id

                            used_qty = sum(SaleOrderLine.search([
                                ('lot_ids', 'in', [lot.id]),
                                ('order_id.transfer_type', '=', 'regular')
                            ]).mapped('product_uom_qty'))

                            in_qty = q.quantity
                            actual_available = in_qty - used_qty

                            if actual_available <= 0:
                                continue

                            allocate = min(remaining_qty, actual_available)

                            existing_line = SaleOrderLine.search([
                                ('order_id', '=', self.sale_order_id.id),
                                ('product_id', '=', product.id),
                                ('lot_ids', 'in', [lot.id]),
                                ('order_id.transfer_type', '=', 'regular')
                            ], limit=1)

                            if existing_line:
                                existing_line.product_uom_qty += allocate
                                existing_line.price_unit = getattr(lot, 'cost_price', product.lst_price)
                            else:
                                SaleOrderLine.create({
                                    'order_id': self.sale_order_id.id,
                                    'product_id': product.id,
                                    'lot_ids': [(6, 0, [lot.id])],
                                    'branded_barcode': lot.ref,
                                    'name': product.display_name,
                                    'product_uom_qty': allocate,
                                    'product_uom': product.uom_id.id,
                                    'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                    'type_product': getattr(lot, 'type_product', False),
                                    'sale_serial_type': sale_serial_type,
                                })

                            remaining_qty -= allocate
                            if remaining_qty <= 0:
                                break

                        if remaining_qty > 0:
                            skipped_lines.append((barcode, f"Short by {remaining_qty}"))

                    continue

            # =========================================================
            #  R START → BRAND (DAMAGE)
            # =========================================================
            if barcode.startswith('R') and self.sale_order_id.transfer_type == 'damage':

                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', barcode),
                    ('quantity', '>', 0),
                    ('location_id', '=', damage_location),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    unbrand_exists = StockQuant.search([
                        ('lot_id.name', '=', barcode),
                        ('location_id', '=', damage_location),
                        ('lot_id.type_product', '=', 'un_brand')
                    ], limit=1)

                    if not unbrand_exists:
                        skipped_lines.append((barcode, "Invalid R barcode"))
                        continue

                else:
                    product = matched_quant.product_id

                    quants = StockQuant.search([
                        ('product_id', '=', product.id),
                        ('lot_id.ref', '=', barcode),
                        ('quantity', '>', 0),
                        ('location_id', '=', damage_location),
                        ('lot_id.type_product', '=', 'brand')
                    ], order='id asc')

                    sale_serial_type = (SaleOrderLine.search([
                        ('order_id', '=', self.sale_order_id.id),
                        ('product_id', '=', product.id),
                        ('order_id.transfer_type', '=', 'damage')
                    ]).mapped('sale_serial_type') or ['regular'])[0]

                    # SERIAL
                    if product.tracking == 'serial':
                        for q in quants[:qty]:
                            lot = q.lot_id

                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': 1,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                    # LOT
                    else:
                        remaining_qty = qty

                        for q in quants:
                            lot = q.lot_id

                            allocate = min(remaining_qty, q.quantity)

                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                            remaining_qty -= allocate
                            if remaining_qty <= 0:
                                break

                    continue

            # ==============================
            # GS1 BARCODE
            # ==============================
            if 'R' in barcode and self.sale_order_id.transfer_type == 'regular':
                lot_code = barcode[barcode.find('R'):]
                quant_domain = [
                    ('lot_id.name', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'un_brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                product = quants[:1].product_id

                if not product:
                    skipped_lines.append((barcode, "No matching product for GS1 barcode"))
                    continue

                # Get serial type
                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id),('order_id.transfer_type','=','regular')
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL PRODUCT
                if product.tracking == 'serial':
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if qty > len(available_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(available_quants)}"))
                        continue

                    created = 0
                    for q in available_quants:
                        lot = q.lot_id
                        # Already used?
                        existing_lines = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','regular')
                        ])
                        if existing_lines:
                            raise ValidationError(f"Not available to add {lot.name}.")

                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': getattr(lot, 'cost_price', product.lst_price),
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': sale_serial_type,
                        })
                        created += 1
                        if created >= qty:
                            break

                # LOT PRODUCT (MERGING LOGIC)
                else:

                    remaining_qty = qty
                    for q in quants:
                        lot = q.lot_id

                        # -----------------------------
                        # Compute already-used and in-qty
                        # -----------------------------
                        used_qty = sum(SaleOrderLine.search([
                            ('order_id.state', 'not in', ['cancel']),
                            ('lot_ids', 'in', [lot.id]),('order_id.transfer_type','=','regular')
                        ]).mapped('product_uom_qty'))

                        in_qty = sum(self.env['stock.move.line'].search([
                            ('picking_id.stock_picking_type', 'in', ['receipt', 'exchange']),
                            ('company_id.nhcl_company_bool', '=', False),
                            ('lot_id', '=', lot.id)
                        ]).mapped('quantity'))

                        actual_available = in_qty - used_qty
                        if actual_available <= 0:
                            continue

                        allocate = min(remaining_qty, actual_available)
                        if allocate <= 0:
                            continue

                        # -----------------------------
                        # MERGE WITH EXISTING LINE
                        # -----------------------------
                        existing_line = SaleOrderLine.search([
                            ('order_id', '=', self.sale_order_id.id),
                            ('product_id', '=', product.id),
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','regular')
                        ], limit=1)

                        # Validate globally before merging
                        if allocate > actual_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available stock for LOT {lot.name}")
                            )
                            continue

                        if existing_line:
                            existing_line.product_uom_qty += allocate
                            existing_line.price_unit = getattr(lot, 'cost_price', product.lst_price)
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds total available stock. Short by {remaining_qty}")
                        )

            elif 'R' in barcode and self.sale_order_id.transfer_type == 'damage':
                lot_code = barcode[barcode.find('R'):]
                quant_domain = [
                    ('lot_id.name', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', damage_location),
                    ('lot_id.type_product', '=', 'un_brand')
                ]
                quants = StockQuant.search(quant_domain, order='id asc')
                product = quants[:1].product_id

                if not product:
                    skipped_lines.append((barcode, "No matching product for GS1 barcode"))
                    continue

                # Get serial type
                existing_types = SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id),('order_id.transfer_type','=','damage')
                ]).mapped('sale_serial_type')
                sale_serial_type = existing_types[0] if existing_types else 'regular'

                # SERIAL PRODUCT
                if product.tracking == 'serial':
                    available_quants = quants.filtered(lambda q: q.quantity > 0)
                    if qty > len(available_quants):
                        skipped_lines.append(
                            (barcode, f"Requested qty {qty} exceeds available serials {len(available_quants)}"))
                        continue

                    created = 0
                    for q in available_quants:
                        lot = q.lot_id
                        # Already used?
                        existing_lines = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','damage')
                        ])
                        if existing_lines:
                            continue

                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': getattr(lot, 'cost_price', product.lst_price),
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': sale_serial_type,
                        })
                        created += 1
                        if created >= qty:
                            break

                # LOT PRODUCT (MERGING LOGIC)
                else:

                    remaining_qty = qty
                    for q in quants:
                        lot = q.lot_id

                        # -----------------------------
                        # Compute already-used and in-qty
                        # -----------------------------
                        used_qty = sum(SaleOrderLine.search([
                            ('order_id.state', 'not in', ['cancel']),
                            ('lot_ids', 'in', [lot.id]),('order_id.transfer_type','=','damage')
                        ]).mapped('product_uom_qty'))

                        in_qty = q.quantity

                        actual_available = in_qty - used_qty
                        if actual_available <= 0:
                            continue

                        allocate = min(remaining_qty, actual_available)
                        if allocate <= 0:
                            continue

                        # -----------------------------
                        # MERGE WITH EXISTING LINE
                        # -----------------------------
                        existing_line = SaleOrderLine.search([
                            ('order_id', '=', self.sale_order_id.id),
                            ('product_id', '=', product.id),
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','damage')
                        ], limit=1)

                        # Validate globally before merging
                        if allocate > actual_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available stock for LOT {lot.name}")
                            )
                            continue

                        if existing_line:
                            existing_line.product_uom_qty += allocate
                            existing_line.price_unit = getattr(lot, 'cost_price', product.lst_price)
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds total available stock. Short by {remaining_qty}")
                        )

            # ==============================
            # EAN-13 BARCODE
            # ==============================
            elif len(barcode) == 13 and self.sale_order_id.transfer_type == 'regular':
                lot_code = barcode
                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
                    continue

                product = matched_quant.product_id
                quants = StockQuant.search([
                    ('product_id', '=', product.id),
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', location_id),
                    ('lot_id.type_product', '=', 'brand')
                ], order='id asc')

                sale_serial_type = (SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id),('order_id.transfer_type','=','regular')
                ]).mapped('sale_serial_type') or ['regular'])[0]

                # SERIAL PRODUCT
                if product.tracking == 'serial':
                    remaining_qty = qty

                    for q in quants:
                        lot = q.lot_id

                        # Check inward existence (serial must have at least one inward)
                        inward_exists = self.env['stock.move.line'].search_count([
                            ('picking_id.stock_picking_type', 'in', ['receipt', 'exchange']),
                            ('company_id.nhcl_company_bool', '=', False),
                            ('lot_id', '=', lot.id),
                        ]) > 0

                        if not inward_exists:
                            continue  # No inward → invalid serial

                        # Check if already used in any non-cancelled regular sale
                        already_sold = SaleOrderLine.search_count([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),
                            ('order_id.transfer_type', '=', 'regular'),
                        ]) > 0

                        if already_sold:
                            continue  # IMPORTANT: serial already consumed

                        # Allocate this serial
                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': getattr(lot, 'cost_price', product.lst_price),
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': sale_serial_type,
                        })

                        remaining_qty -= 1
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        raise ValidationError(
                            f"No available serials for barcode {barcode}. "
                            f"Short by {remaining_qty}"
                        )

                # LOT PRODUCT (MERGING LOGIC)
                else:

                    remaining_qty = qty
                    for q in quants:
                        lot = q.lot_id

                        used_qty = sum(SaleOrderLine.search([
                            ('order_id.state', 'not in', ['cancel']),
                            ('lot_ids', 'in', [lot.id]),('order_id.transfer_type','=','regular')
                        ]).mapped('product_uom_qty'))

                        in_qty = sum(self.env['stock.move.line'].search([
                            ('picking_id.stock_picking_type', 'in', ['receipt', 'exchange']),
                            ('company_id.nhcl_company_bool', '=', False),
                            ('lot_id', '=', lot.id)
                        ]).mapped('quantity'))

                        actual_available = in_qty - used_qty
                        if actual_available <= 0:
                            continue

                        allocate = min(remaining_qty, actual_available)

                        existing_line = SaleOrderLine.search([
                            ('order_id', '=', self.sale_order_id.id),
                            ('product_id', '=', product.id),
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','regular')
                        ], limit=1)

                        if allocate > actual_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available for LOT {lot.name}")
                            )
                            continue

                        if existing_line:
                            existing_line.product_uom_qty += allocate
                            existing_line.price_unit = getattr(lot, 'cost_price', product.lst_price)
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds total available stock. Short by {remaining_qty}")
                        )

            elif len(barcode) == 13 and self.sale_order_id.transfer_type == 'damage':
                lot_code = barcode
                matched_quant = StockQuant.search([
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', damage_location),
                    ('lot_id.type_product', '=', 'brand')
                ], limit=1)

                if not matched_quant:
                    skipped_lines.append((barcode, "No stock found for EAN-13 barcode"))
                    continue

                product = matched_quant.product_id
                quants = StockQuant.search([
                    ('product_id', '=', product.id),
                    ('lot_id.ref', '=', lot_code),
                    ('company_id.nhcl_company_bool', '=', False),
                    ('quantity', '>', 0),
                    ('location_id', '=', damage_location),
                    ('lot_id.type_product', '=', 'brand')
                ], order='id asc')

                sale_serial_type = (SaleOrderLine.search([
                    ('order_id', '=', self.sale_order_id.id),
                    ('order_id.state', 'not in', ['cancel']),
                    ('product_id', '=', product.id),('order_id.transfer_type','=','damage')
                ]).mapped('sale_serial_type') or ['regular'])[0]

                # SERIAL PRODUCT
                if product.tracking == 'serial':
                    available_quants = quants
                    if qty > len(available_quants):
                        short = qty - len(available_quants)
                        skipped_lines.append((barcode, f"Requested qty exceeds available stock. Short by {short}"))
                        continue

                    created = 0
                    for q in available_quants:
                        lot = q.lot_id
                        used = SaleOrderLine.search([
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','damage')
                        ])
                        if used:
                            continue

                        SaleOrderLine.create({
                            'order_id': self.sale_order_id.id,
                            'product_id': product.id,
                            'lot_ids': [(6, 0, [lot.id])],
                            'branded_barcode': lot.ref,
                            'name': product.display_name,
                            'product_uom_qty': 1,
                            'product_uom': product.uom_id.id,
                            'price_unit': getattr(lot, 'cost_price', product.lst_price),
                            'type_product': getattr(lot, 'type_product', False),
                            'sale_serial_type': sale_serial_type,
                        })
                        created += 1
                        if created >= qty:
                            break

                # LOT PRODUCT (MERGING LOGIC)
                else:

                    remaining_qty = qty
                    for q in quants:
                        lot = q.lot_id

                        used_qty = sum(SaleOrderLine.search([
                            ('order_id.state', 'not in', ['cancel']),
                            ('lot_ids', 'in', [lot.id]),('order_id.transfer_type','=','damage')
                        ]).mapped('product_uom_qty'))

                        in_qty = q.quantity

                        actual_available = in_qty - used_qty
                        if actual_available <= 0:
                            continue

                        allocate = min(remaining_qty, actual_available)

                        existing_line = SaleOrderLine.search([
                            ('order_id', '=', self.sale_order_id.id),
                            ('product_id', '=', product.id),
                            ('lot_ids', 'in', [lot.id]),
                            ('order_id.state', 'not in', ['cancel']),('order_id.transfer_type','=','damage')
                        ], limit=1)

                        if allocate > actual_available:
                            skipped_lines.append(
                                (barcode, f"Requested qty exceeds available for LOT {lot.name}")
                            )
                            continue

                        if existing_line:
                            existing_line.product_uom_qty += allocate
                            existing_line.price_unit = getattr(lot, 'cost_price', product.lst_price)
                        else:
                            SaleOrderLine.create({
                                'order_id': self.sale_order_id.id,
                                'product_id': product.id,
                                'lot_ids': [(6, 0, [lot.id])],
                                'branded_barcode': lot.ref,
                                'name': product.display_name,
                                'product_uom_qty': allocate,
                                'product_uom': product.uom_id.id,
                                'price_unit': getattr(lot, 'cost_price', product.lst_price),
                                'type_product': getattr(lot, 'type_product', False),
                                'sale_serial_type': sale_serial_type,
                            })

                        remaining_qty -= allocate
                        if remaining_qty <= 0:
                            break

                    if remaining_qty > 0:
                        skipped_lines.append(
                            (barcode, f"Requested qty exceeds total available stock. Short by {remaining_qty}")
                        )

            else:
                skipped_lines.append((barcode, "Unknown barcode format"))
                continue

        # FINAL ERROR REPORT
        if skipped_lines:
            msg = "\n".join([f"{b}: {reason}" for b, reason in skipped_lines])
            raise ValidationError(_("Some lines were skipped:\n%s") % msg)
