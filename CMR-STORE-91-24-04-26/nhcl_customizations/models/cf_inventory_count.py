from odoo import fields, models, api, _
from odoo.exceptions import ValidationError
import base64
import io
from openpyxl import load_workbook
import xlrd


class AssetMaster(models.Model):
    _name = 'cf.asset.master'
    _description = 'Asset Master'

    name = fields.Char(required=True)
    code = fields.Char()

    @api.constrains('name')
    def _check_unique_name(self):
        for rec in self:
            if rec.name:
                existing = self.search([('name', '=', rec.name),('id', '!=', rec.id)], limit=1)
                if existing:
                    raise ValidationError(_("Asset with name '%s' already exists.") % rec.name)

class CentralFashionsCount(models.Model):
    _name = 'cf.inventory.count'
    _description = 'Central Fashions Inventory Count'

    name = fields.Char(string="Reference", required=True,
                       copy=False, readonly=True, default=lambda self: _('New'))
    requester_id = fields.Many2one('res.users', string="Requester",
                                   default=lambda self: self.env.user, required=True)
    date = fields.Date(default=fields.Date.today, required=True)
    company_id = fields.Many2one(
        'res.company',
        default=lambda self: self.env.company,
        required=True
    )

    upload_file = fields.Binary(string="Upload Excel")
    upload_filename = fields.Char()

    line_ids = fields.One2many(
        'cf.inventory.count.line',
        'count_id',
        string="Lines"
    )

    mode = fields.Selection([
        ('scan', 'Scan'),
        ('upload', 'Upload')
    ], string="Mode", default='scan')
    scan_barcode = fields.Char(string="Scan Barcode")

    @api.onchange('scan_barcode')
    def _onchange_scan_barcode(self):
        if not self.scan_barcode:
            return

        barcode = self.scan_barcode.strip()
        product = False

        # 🔹 STEP 1: Check MAIN barcode (product.product)
        product = self.env['product.product'].search([
            ('barcode', '=', barcode)
        ], limit=1)

        # 🔹 STEP 2: If not found → MULTI BARCODE
        if not product:
            multi_barcode = self.env['product.barcode'].search([
                ('barcode', '=', barcode)
            ], limit=1)

            if multi_barcode:
                product = multi_barcode.product_id

        # ❌ STEP 3: Not found
        if not product:
            self.scan_barcode = False
            raise ValidationError(f"Barcode '{barcode}' not found")

        product_tmpl = product.product_tmpl_id

        # 🔹 STEP 4: Existing line check
        line = self.line_ids.filtered(
            lambda l: l.product_tmpl_id.id == product_tmpl.id
        )

        if line:
            #  Increment count
            line.global_count += 1
        else:

            self.line_ids += self.line_ids.new({
                'product_tmpl_id': product_tmpl.id,
                'family_id': product_tmpl.categ_id.parent_id.parent_id.parent_id.id,
                'cat_id': product_tmpl.categ_id.parent_id.parent_id.id,
                'class_id': product_tmpl.categ_id.parent_id.id,
                'brick_id': product_tmpl.categ_id.id,
                'global_count': 1,
            })

        # Reset field (important for next scan)
        self.scan_barcode = False

    @api.constrains('date')
    def _check_unique_date(self):
        for rec in self:
            if rec.date:
                existing = self.search([('date', '=', rec.date), ('id', '!=', rec.id)], limit=1)
                if existing:
                    raise ValidationError(_(f"Global Count with date {rec.date} already exists."))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get('name', _('New')) == _('New'):
                vals['name'] = self.env['ir.sequence'].next_by_code('inventory.count') or _('New')

        return super(CentralFashionsCount, self).create(vals_list)

    def action_import_excel(self):
        self.ensure_one()

        if not self.upload_file:
            raise ValidationError("Please upload an Excel file.")

        self.line_ids.unlink()

        file_data = base64.b64decode(self.upload_file)
        workbook = load_workbook(io.BytesIO(file_data))
        sheet = workbook.active

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

            asset_name = str(row[0]).strip() if row[0] else ''
            product_name = str(row[1]).strip() if row[1] else ''
            floor_value_excel = row[2]
            global_count = row[3]

            if not asset_name:
                raise ValidationError(f"Asset missing at Row {row_index}")

            if not product_name:
                raise ValidationError(f"Product missing at Row {row_index}")

            if global_count is None:
                raise ValidationError(f"Global Count missing at Row {row_index}")

            # 🔹 Asset
            asset = self.env['cf.asset.master'].search(
                [('name', '=', asset_name)], limit=1
            )
            if not asset:
                raise ValidationError(f"Asset '{asset_name}' not found (Row {row_index})")

            # 🔹 Product Template
            product_tmpl = self.env['product.template'].search(
                [('name', '=', product_name)], limit=1
            )
            if not product_tmpl:
                raise ValidationError(f"Product '{product_name}' not found (Row {row_index})")

            # 🔹 Floor conversion
            try:
                floor_key = str(int(float(floor_value_excel)))
            except Exception:
                raise ValidationError(f"Invalid Floor '{floor_value_excel}' (Row {row_index})")

            valid_keys = dict(self.env['cf.inventory.count.line']._fields['floor'].selection).keys()

            if floor_key not in valid_keys:
                raise ValidationError(f"Invalid Floor '{floor_value_excel}' (Row {row_index})")
            soh = sum(product_tmpl.product_variant_ids.mapped('qty_available'))
            diff = global_count - soh
            # 🔹 Create Line
            self.env['cf.inventory.count.line'].create({
                'count_id': self.id,
                'asset_id': asset.id,
                'product_tmpl_id': product_tmpl.id,
                'floor': floor_key,
                'family_id': product_tmpl.categ_id.parent_id.parent_id.parent_id.id,
                'cat_id': product_tmpl.categ_id.parent_id.parent_id.id,
                'class_id': product_tmpl.categ_id.parent_id.id,
                'brick_id': product_tmpl.categ_id.id,
                'global_count': global_count,
                'soh': soh,
                'diff': diff,
            })


class CentralFashionsCountLine(models.Model):
    _name = 'cf.inventory.count.line'
    _description = 'Inventory Count Line'

    count_id = fields.Many2one('cf.inventory.count', ondelete='cascade')
    asset_id = fields.Many2one('cf.asset.master')
    floor = fields.Selection([
        ('0', 'Ground Floor'),
        ('1', '1st Floor'),
        ('2', '2nd Floor'),
        ('3', '3rd Floor'),
        ('4', '4th Floor'),
        ('5', '5th Floor'),
        ('6', '6th Floor'),
        ('7', '7th Floor'),
    ], string="Floor")
    product_tmpl_id = fields.Many2one('product.template' ,string='Product')
    family_id = fields.Many2one('product.category',string='Family')
    cat_id = fields.Many2one('product.category',string='Category',)
    class_id = fields.Many2one('product.category',string='Class')
    brick_id = fields.Many2one('product.category',string='Brick')
    global_count = fields.Float(string='Global Count')
    soh = fields.Float(string='SOH')
    diff = fields.Float(string='Diff', compute='get_diff', store=True)

    @api.depends('soh')
    def get_diff(self):
        for rec in self:
            if rec.soh:
                rec.diff = rec.global_count - rec.soh


    @api.onchange('product_tmpl_id')
    def onchange_product_tmpl_id(self):
        for rec in self:
            if self.product_tmpl_id:
                soh = sum(self.product_tmpl_id.product_variant_ids.mapped('qty_available'))
                rec.family_id = self.product_tmpl_id.categ_id.parent_id.parent_id.parent_id.id
                rec.cat_id = self.product_tmpl_id.categ_id.parent_id.parent_id.id
                rec.class_id = self.product_tmpl_id.categ_id.parent_id.id
                rec.brick_id = self.product_tmpl_id.categ_id.id
                rec.soh = soh

    @api.onchange('global_count')
    def onchange_global_count(self):
        for rec in self:
            if rec.global_count and rec.soh:
                rec.diff = rec.global_count - rec.soh
